"""
trend_analyzer.py
分析股票趨勢變化（今天 vs 昨天）
"""
import pandas as pd
import glob
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def get_recent_files(days=2):
    """取得最近 N 天的數據檔案"""
    files = glob.glob("stock_data_*.csv")
    if len(files) < days:
        return None
    
    # 按日期排序
    files_sorted = sorted(files, reverse=True)
    return files_sorted[:days]

print("=" * 70)
print("  股票趨勢分析器")
print("=" * 70)
print()

# 取得今天和昨天的數據
files = get_recent_files(2)

if not files or len(files) < 2:
    print("❌ 需要至少 2 天的數據才能分析趨勢")
    print("   請先執行：python test_trend.py")
    exit(1)

today_file = files[0]
yesterday_file = files[1]

print(f"📊 今天數據：{today_file}")
print(f"📊 昨天數據：{yesterday_file}")
print()

# 載入數據
df_today = pd.read_csv(today_file, encoding="utf-8-sig")
df_yesterday = pd.read_csv(yesterday_file, encoding="utf-8-sig")

print(f"✅ 今天：{len(df_today)} 筆")
print(f"✅ 昨天：{len(df_yesterday)} 筆")
print()

# 合併數據（以股票代號為 key）
df_merged = pd.merge(
    df_today, 
    df_yesterday, 
    on='股票代號', 
    suffixes=('_今', '_昨'),
    how='inner'
)

print(f"🔗 成功比對：{len(df_merged)} 筆")
print()

# 計算變化
print("⏳ 計算變化中...")

df_merged['價格漲跌'] = df_merged['收盤價_今'] - df_merged['收盤價_昨']
df_merged['漲跌幅%'] = ((df_merged['收盤價_今'] - df_merged['收盤價_昨']) / df_merged['收盤價_昨'] * 100).round(2)
df_merged['K值變化'] = (df_merged['K值_今'] - df_merged['K值_昨']).round(2)
df_merged['D值變化'] = (df_merged['D值_今'] - df_merged['D值_昨']).round(2)
df_merged['RSI變化'] = (df_merged['RSI_今'] - df_merged['RSI_昨']).round(2)
df_merged['BB位置變化'] = (df_merged['BB位置_今'] - df_merged['BB位置_昨']).round(2)
df_merged['量比變化'] = (df_merged['量比_今'] - df_merged['量比_昨']).round(2)

# 判斷趨勢
df_merged['價格趨勢'] = df_merged['漲跌幅%'].apply(
    lambda x: '大漲' if x >= 3 else ('上漲' if x > 0 else ('下跌' if x < -3 else '平穩'))
)

df_merged['KD趨勢'] = df_merged.apply(
    lambda row: 'KD轉強' if row['K值變化'] > 3 and row['D值變化'] > 3 
    else ('KD轉弱' if row['K值變化'] < -3 and row['D值變化'] < -3 else 'KD平穩'),
    axis=1
)

df_merged['RSI趨勢'] = df_merged['RSI變化'].apply(
    lambda x: 'RSI轉強' if x > 5 else ('RSI轉弱' if x < -5 else 'RSI平穩')
)

df_merged['量能趨勢'] = df_merged['量比變化'].apply(
    lambda x: '量能放大' if x > 0.5 else ('量能縮小' if x < -0.5 else '量能平穩')
)

# 選擇輸出欄位
output_cols = [
    '股票代號', '股票名稱_今',
    '收盤價_今', '收盤價_昨', '價格漲跌', '漲跌幅%', '價格趨勢',
    'K值_今', 'K值_昨', 'K值變化', 'KD趨勢',
    'D值_今', 'D值_昨', 'D值變化',
    'RSI_今', 'RSI_昨', 'RSI變化', 'RSI趨勢',
    'BB位置_今', 'BB位置_昨', 'BB位置變化',
    '量比_今', '量比_昨', '量比變化', '量能趨勢',
    '殖利率_今', '本益比_今', '市值億_今'
]

df_output = df_merged[output_cols].copy()

# 重新命名欄位
df_output.columns = [
    '股票代號', '股票名稱',
    '今收盤', '昨收盤', '漲跌', '漲跌幅%', '價格趨勢',
    '今K值', '昨K值', 'K變化', 'KD趨勢',
    '今D值', '昨D值', 'D變化',
    '今RSI', '昨RSI', 'RSI變化', 'RSI趨勢',
    '今BB位置', '昨BB位置', 'BB變化',
    '今量比', '昨量比', '量比變化', '量能趨勢',
    '殖利率', '本益比', '市值億'
]

# 按漲跌幅排序
df_output = df_output.sort_values('漲跌幅%', ascending=False)

# 儲存 Excel
timestamp = datetime.now().strftime("%Y%m%d")
excel_file = f"trend_report_{timestamp}.xlsx"

df_output.to_excel(excel_file, index=False, engine='openpyxl')

print("⏳ 美化 Excel 格式...")

wb = load_workbook(excel_file)
ws = wb.active

# 調整欄寬
print("⏳ 調整欄寬...")

for column in ws.columns:
    max_length = 0
    column_letter = column[0].column_letter
    header_value = str(column[0].value) if column[0].value else ""
    
    for cell in column:
        try:
            if cell.value is None:
                continue
            val = str(cell.value)
            length = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in val)
            if length > max_length:
                max_length = length
        except:
            pass
    
    adjusted_width = max_length + 2
    
    if header_value in ['漲跌', 'K變化', 'D變化', 'RSI變化', 'BB變化']:
        final_width = max(9, min(adjusted_width, 12))
    elif header_value in ['股票代號', '今收盤', '昨收盤']:
        final_width = max(10, min(adjusted_width, 12))
    elif header_value in ['股票名稱']:
        final_width = max(16, min(adjusted_width, 22))
    elif header_value in ['價格趨勢', 'KD趨勢', 'RSI趨勢', '量能趨勢']:
        final_width = max(11, min(adjusted_width, 16))
    else:
        final_width = max(10, min(adjusted_width, 20))
    
    ws.column_dimensions[column_letter].width = final_width

ws.freeze_panes = "A2"

header_fill = PatternFill(start_color="4472C4", 
                          end_color="4472C4", 
                          fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

from openpyxl.worksheet.table import Table, TableStyleInfo

last_column = ws.max_column
last_row = ws.max_row

def col_letter(n):
    string = ""
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        string = chr(65 + remainder) + string
    return string

table_ref = f"A1:{col_letter(last_column)}{last_row}"
tab = Table(displayName="TrendTable", ref=table_ref)

style = TableStyleInfo(
    name="TableStyleLight1",
    showFirstColumn=False,
    showLastColumn=False,
    showRowStripes=True,
    showColumnStripes=False
)
tab.tableStyleInfo = style

ws.add_table(tab)
wb.save(excel_file)

print()
print("=" * 70)
print(f"✅ 趨勢報表生成完成！")
print(f"   檔案：{excel_file}")
print(f"   • 深淺交錯格式（篩選排序不影響）✓")
print(f"   • 凍結標題列 ✓")
print(f"   • 自動篩選 ✓")
print()

# 統計資訊
print("📊 趨勢統計：")
print(f"   上漲股票：{len(df_output[df_output['漲跌幅%'] > 0])} 筆")
print(f"   下跌股票：{len(df_output[df_output['漲跌幅%'] < 0])} 筆")
print(f"   漲幅 > 3%：{len(df_output[df_output['漲跌幅%'] >= 3])} 筆")
print(f"   跌幅 > 3%：{len(df_output[df_output['漲跌幅%'] <= -3])} 筆")
print()
print(f"   KD轉強：{len(df_output[df_output['KD趨勢'] == 'KD轉強'])} 筆")
print(f"   RSI轉強：{len(df_output[df_output['RSI趨勢'] == 'RSI轉強'])} 筆")
print(f"   量能放大：{len(df_output[df_output['量能趨勢'] == '量能放大'])} 筆")
print()

# 顯示前 10 名
print("📈 今日漲幅前 10 名：")
top10 = df_output.head(10)
for idx, row in top10.iterrows():
    print(f"   {row['股票代號']} {row['股票名稱']:8s} {row['漲跌幅%']:6.2f}% ({row['價格趨勢']})")

print()
print("=" * 70)