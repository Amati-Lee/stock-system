"""
============================================================
 台股即時看盤系統 — stock_system.py
 支持篩選模式：漲停 / 殖利率 / RSI+MA策略
 日期區間可自選，閾值可參數化調整
============================================================
"""

import os
import sys
import time
import json
import logging
import warnings

# 修正 Windows 終端編碼（避免 emoji 導致 cp950 UnicodeEncodeError）
if sys.stdout and hasattr(sys.stdout, 'reconfigure'):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
from datetime import datetime, timedelta
from dataclasses import dataclass, field

# --- 壓制 yfinance 內部的 logging / warning 訊息 ---
# （如 "possibly delisted"、HTTP 404 等）
logging.getLogger("yfinance").setLevel(logging.CRITICAL)
warnings.filterwarnings("ignore")

import yfinance as yf
import pandas as pd


# ============================================================
# 篩選參數（全部可自定義）
# ============================================================

@dataclass
class ScreenerConfig:
    """所有篩選參數集中管理"""

    # --- 日期設定 ---
    start_date: str = ""          # 開始日期 YYYY-MM-DD（空字串 = 自動算）
    end_date:   str = ""          # 結束日期 YYYY-MM-DD（空字串 = 今天）
    lookback_days: int = 10       # 若未指定日期，往前推幾天

    # --- 篩選模式旗標 ---
    filter_limit_up:   bool = False   # 漲停篩選
    filter_dividend:   bool = False   # 殖利率篩選
    filter_rsi_ma:     bool = False   # RSI+MA策略篩選
    filter_volume:     bool = False   # 成交量篩選
    filter_market_cap: bool = False   # 市值篩選
    filter_pe_ratio:   bool = False   # 本益比篩選
    filter_bbands:     bool = False   # 布林通道開口向上篩選
    filter_resonance:  bool = False   # 技術指標共振策略（KD + RSI + BB + 量 + MACD）
    filter_kd_golden:  bool = False   # KD 黃金交叉（做多訊號）
    filter_bb_upper:   bool = False   # 布林通道接近上軌（強勢突破）

    # --- 漲停參數 ---
    limit_up_pct: float = 10.0        # 台股普通股漲停幅度 %（近年改為 10%）

    # --- 殖利率參數 ---
    min_dividend_yield: float = 5.0   # 最低殖利率 %

    # --- RSI / MA 參數 ---
    rsi_period: int   = 14
    ma_period:  int   = 20
    rsi_oversold: float = 30.0        # RSI 低於此值視為超賣

    # --- KD 參數 ---
    kd_period: int = 9                # KD 週期
    kd_oversold_k: float = 20.0       # K 值超賣閾值
    kd_oversold_d: float = 20.0       # D 值超賣閾值

    # --- MACD 參數 ---
    macd_fast: int = 12               # 快線 EMA 週期
    macd_slow: int = 26               # 慢線 EMA 週期
    macd_signal: int = 9              # 訊號線 EMA 週期

    # --- 布林通道參數 ---
    bbands_period: int = 20           # 布林通道週期
    bbands_std: float = 2.0           # 標準差倍數
    bbands_expand_min: float = 0.05   # 開口擴大最小幅度（5%）

    # --- 成交量參數（共振用）---
    min_volume: int = 1000            # 最低日均量（張）
    volume_surge_ratio: float = 1.5   # 量能放大倍數（今日量 >= 5日均量 * 1.5）

    # --- 市值參數 ---
    min_market_cap: float = 10.0      # 最低市值（億元）
    max_market_cap: float = 10000.0   # 最高市值（億元）

    # --- 本益比參數 ---
    min_pe_ratio: float = 0.0         # 最低本益比
    max_pe_ratio: float = 50.0        # 最高本益比

    # --- 掃描範圍 ---
    max_stocks: int = 50              # 最多掃描幾筆股票
    
    # --- 排序設定 ---
    sort_by: str = ""                 # 排序欄位：dividend_yield, market_cap, pe_ratio, volume 等
    sort_desc: bool = True            # True=由大到小, False=由小到大


# ============================================================
# 日期工具
# ============================================================

def resolve_dates(cfg: ScreenerConfig) -> tuple[str, str]:
    """根據 config 計算 start / end 日期字串"""
    today = datetime.today()

    end   = datetime.strptime(cfg.end_date, "%Y-%m-%d") if cfg.end_date   else today
    start = datetime.strptime(cfg.start_date, "%Y-%m-%d") if cfg.start_date else (end - timedelta(days=cfg.lookback_days))

    # yfinance 的 end 日期是不含的，所以 +1 天
    return start.strftime("%Y-%m-%d"), (end + timedelta(days=1)).strftime("%Y-%m-%d")


# ============================================================
# 股票清單（離線主清單 + yfinance 驗證緩存）
# ============================================================
# 策略：
#   - 內建一張離線的台股代號清單（約 1700 筆候選）
#   - 第一次執行時用 yfinance 逐一快速驗證哪些是活著的股票
#   - 驗證後緩存到 tw_stock_verified.txt（永久有效）
#   - 之後直接讀緩存，除非手動刪除或使用 deep=True 重新驗證
# ============================================================

VERIFIED_CACHE = "tw_stock_verified.txt"

# --- 離線主清單：台股上市板常見代號範圍 ---
def _get_master_codes(deep: bool = False) -> list[str]:
    """
    生成候選股票代號清單。

    deep=False（預設，快速模式）：
      用已知的活躍範圍，約 6600 筆候選，驗證約 22 分鐘。
      涵蓋上市＋上櫃所有主要號段。

    deep=True（深度掃描）：
      開啟 1100-9999 全部範圍，約 8900 筆候選，驗證約 30 分鐘。
      確保上市＋上櫃全部都會被掃到，不會漏掉任何股票。
      建議第一次緩存過期後用這個模式跑一次。
    """
    if deep:
        # 全範圍掃描：台股所有可能的 4 位數字代號
        codes = list(range(1100, 10000))
    else:
        # 快速模式：涵蓋上市＋上櫃所有主要號段
        ranges = [
            (1101, 1799),   # 傳統產業（上市：水泥、食品、塑膠、紡織、機電等）
            (2001, 2999),   # 電子/金融（上市＋部分上櫃）
            (3001, 3999),   # 電子/生技（上市＋上櫃）
            (4100, 4999),   # 上櫃（生技、光電、IT服務、半導體等）
            (5001, 5880),   # 上市新股＋上櫃電子
            (6001, 6999),   # 上市＋上櫃（新掛牌、綠能、IC設計等）
            (8001, 8999),   # 上櫃（金融、營建、觀光、貿易等）
            (9001, 9958),   # 上市/上櫃（觀光、貿易、綜合等）
        ]
        codes = []
        for start, end in ranges:
            codes.extend(range(start, end + 1))

    return [str(c) for c in sorted(set(codes))]


def _resolve_suffix(ticker: str) -> str | None:
    """判斷股票後綴：上市 .TW 或上櫃 .TWO，回傳完整 symbol 或 None"""
    import sys, io
    _orig_out, _orig_err = sys.stdout, sys.stderr
    sys.stdout = io.StringIO()
    sys.stderr = io.StringIO()
    try:
        for suffix in (".TW", ".TWO"):
            df = yf.download(f"{ticker}{suffix}", period="1d", progress=False)
            if not df.empty:
                return f"{ticker}{suffix}"
        return None
    except Exception:
        return None
    finally:
        sys.stdout, sys.stderr = _orig_out, _orig_err


def _verify_stock(ticker: str) -> bool:
    """用 yfinance 快速驗證一個股票是否存在（上市+上櫃）"""
    return _resolve_suffix(ticker) is not None


def _run_verification(deep: bool = False) -> list[str]:
    """
    對主清單逐一驗證，印出進度。
    deep=True 時會開啟全範圍掃描（約 30 分鐘）。
    """
    master = _get_master_codes(deep=deep)
    mode_label = "深度掃描（全範圍，含上櫃）" if deep else "快速模式（已知範圍）"
    print(f"📥 開始驗證股票清單 [{mode_label}]")
    print(f"   候選：{len(master)} 筆")
    print(f"   預估時間：約 {len(master) * 0.2 / 60:.0f} 分鐘，驗證完後緩存 7 天")
    print(f"   你可以先去泡杯咖啡 ☕")
    print()

    verified = []
    dead = 0
    total = len(master)

    for i, code in enumerate(master, 1):
        alive = _verify_stock(code)
        if alive:
            verified.append(code)
        else:
            dead += 1

        # 每 50 筆印一次進度
        if i % 50 == 0 or i == total:
            pct = i / total * 100
            print(f"  進度：{i:>4}/{total} ({pct:5.1f}%)  有效：{len(verified)}  無效：{dead}")

        time.sleep(0.2)

    print(f"\n✅ 驗證完成：有效 {len(verified)} 筆 / 無效 {dead} 筆")
    return verified


# 註：不再使用日期檢查，緩存永久有效
# def _cache_is_fresh() -> bool:
#     if not os.path.exists(VERIFIED_CACHE):
#         return False
#     mtime = os.path.getmtime(VERIFIED_CACHE)
#     age_days = (time.time() - mtime) / 86400
#     return age_days < 7


def _save_verified(stocks: list[str]):
    with open(VERIFIED_CACHE, "w", encoding="utf-8") as f:
        for s in stocks:
            f.write(s + "\n")
    print(f"💾 緩存寫入 {VERIFIED_CACHE}（{len(stocks)} 筆有效股票）")


def _load_verified() -> list[str]:
    with open(VERIFIED_CACHE, "r", encoding="utf-8") as f:
        return [line.strip() for line in f if line.strip()]


def get_stock_list(cfg: ScreenerConfig, deep: bool = False) -> list[str]:
    """
    取股票清單的主入口。

    優先使用 stock_names_all.json（TWSE/TPEx API 來源，排除興櫃），
    若 JSON 不存在或為空則 fallback 到 tw_stock_verified.txt。
    deep=True 時強制重新驗證全範圍並更新 tw_stock_verified.txt。
    """
    if deep:
        print("🔍 深度掃描模式：強制重新驗證全範圍股票清單 (1100-9999)")
        stocks = _run_verification(deep=True)
        if stocks:
            _save_verified(stocks)
            print(f"💾 清單已更新：{len(stocks)} 筆股票已存入 tw_stock_verified.txt")
        else:
            print("❌ 驗證結果為空")
            return []
    else:
        # 優先用 stock_names_all.json（排除興櫃）
        stocks = _get_stocks_from_names_json()
        if stocks:
            print(f"📦 使用 stock_names_all.json 清單：{len(stocks)} 筆（上市+上櫃+興櫃）")
        elif os.path.exists(VERIFIED_CACHE):
            stocks = _load_verified()
            print(f"📦 Fallback 到 tw_stock_verified.txt：{len(stocks)} 筆")
        else:
            print("⚠️  找不到股票清單，自動建立...")
            stocks = _run_verification(deep=False)
            if stocks:
                _save_verified(stocks)
                print(f"💾 清單已建立：{len(stocks)} 筆股票")
            else:
                print("❌ 驗證結果為空")
                return []

    count = min(len(stocks), cfg.max_stocks)
    est_min = count * 0.3 / 60
    print(f"📊 本次掃描：{count} 筆，預估時間：約 {est_min:.1f} 分鐘")

    return stocks[:cfg.max_stocks]


def _get_stocks_from_names_json() -> list[str]:
    """從 stock_names_all.json 取得上市+上櫃股票清單（排除興櫃）。"""
    json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stock_names_all.json")
    market_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stock_market_type.json")
    if not os.path.exists(json_path):
        return []
    try:
        with open(json_path, "r", encoding="utf-8") as f:
            names = json.load(f)
        market = {}
        if os.path.exists(market_path):
            with open(market_path, "r", encoding="utf-8") as f:
                market = json.load(f)
        stocks = sorted(k for k in names if market.get(k, "") in ("上市", "上櫃", "興櫃"))
        return stocks if len(stocks) >= 500 else []  # 安全閾值：少於 500 筆視為異常
    except Exception:
        return []


# ============================================================
# yfinance 數據下載（統一處理 MultiIndex）
# ============================================================

# 後綴快取：避免同一支股票重複嘗試 .TW / .TWO
_suffix_cache: dict[str, str] = {}


def _get_symbol(ticker: str) -> str:
    """回傳帶後綴的 yfinance symbol（優先用快取）"""
    if ticker in _suffix_cache:
        return _suffix_cache[ticker]
    # 預設 .TW，download_stock 裡會自動 fallback
    return f"{ticker}.TW"


def download_stock(ticker: str, start: str, end: str) -> pd.DataFrame | None:
    """
    下載單股數據並統一為扁平 DataFrame。
    return 的 columns 全部小寫：open, high, low, close, volume
    自動偵測上市(.TW)/上櫃(.TWO)，並快取後綴。
    """
    import sys, io

    try:
        _orig_out, _orig_err = sys.stdout, sys.stderr
        sys.stdout = io.StringIO()
        sys.stderr = io.StringIO()
        try:
            # 如果已知後綴就直接用，否則依序嘗試
            if ticker in _suffix_cache:
                suffixes = [_suffix_cache[ticker].replace(ticker, "")]
            else:
                suffixes = [".TW", ".TWO"]

            df = None
            for suffix in suffixes:
                symbol = f"{ticker}{suffix}"
                df = yf.download(
                    symbol,
                    start=start,
                    end=end,
                    progress=False,
                    auto_adjust=True
                )
                if not df.empty:
                    _suffix_cache[ticker] = symbol
                    break
                df = None
        finally:
            sys.stdout, sys.stderr = _orig_out, _orig_err

        if df is None or df.empty:
            return None

        # --- 處理 yfinance 新版本的 MultiIndex columns ---
        if isinstance(df.columns, pd.MultiIndex):
            df.columns = df.columns.get_level_values(0)

        df.columns = df.columns.str.lower()

        # 去除重複欄位名（yfinance 偶爾回傳重複 columns）
        df = df.loc[:, ~df.columns.duplicated()]

        # 確認必要欄位存在
        required = {"open", "high", "low", "close", "volume"}
        if not required.issubset(set(df.columns)):
            return None

        return df

    except Exception:
        return None


def _fetch_ticker_info(ticker: str) -> dict | None:
    """取得 yfinance Ticker info，自動偵測 .TW/.TWO"""
    import sys, io
    try:
        _orig_out, _orig_err = sys.stdout, sys.stderr
        sys.stdout = io.StringIO()
        sys.stderr = io.StringIO()
        try:
            # 優先用快取的後綴
            if ticker in _suffix_cache:
                return yf.Ticker(_suffix_cache[ticker]).info
            for suffix in (".TW", ".TWO"):
                info = yf.Ticker(f"{ticker}{suffix}").info
                if info and info.get("regularMarketPrice"):
                    _suffix_cache[ticker] = f"{ticker}{suffix}"
                    return info
            return None
        finally:
            sys.stdout, sys.stderr = _orig_out, _orig_err
    except Exception:
        return None


# ============================================================
# 篩選邏輯
# ============================================================

def calc_rsi(series: pd.Series, period: int = 14) -> pd.Series:
    """計算 RSI"""
    delta = series.diff()
    gain  = delta.clip(lower=0)
    loss  = -delta.clip(upper=0)
    avg_gain = gain.rolling(period).mean()
    avg_loss = loss.rolling(period).mean()
    rs = avg_gain / avg_loss
    return 100 - (100 / (1 + rs))


def calc_kd(df: pd.DataFrame, period: int = 9) -> tuple[pd.Series, pd.Series]:
    """
    計算 KD 指標（隨機指標）
    回傳：(K值, D值)
    """
    # 計算 RSV (Raw Stochastic Value)
    low_min = df["low"].rolling(period).min()
    high_max = df["high"].rolling(period).max()
    rsv = 100 * (df["close"] - low_min) / (high_max - low_min)
    
    # K 值 = RSV 的加權移動平均（權重 1/3）
    k = rsv.ewm(alpha=1/3, adjust=False).mean()
    
    # D 值 = K 值的加權移動平均（權重 1/3）
    d = k.ewm(alpha=1/3, adjust=False).mean()
    
    return k, d


def calc_macd(series: pd.Series, fast: int = 12, slow: int = 26, signal: int = 9) -> tuple[pd.Series, pd.Series, pd.Series]:
    """
    計算 MACD 指標
    回傳：(DIF, MACD, 柱狀圖)
    """
    # DIF (快線) = EMA12 - EMA26
    ema_fast = series.ewm(span=fast, adjust=False).mean()
    ema_slow = series.ewm(span=slow, adjust=False).mean()
    dif = ema_fast - ema_slow
    
    # MACD (慢線/訊號線) = DIF 的 9 日 EMA
    macd = dif.ewm(span=signal, adjust=False).mean()
    
    # 柱狀圖 = DIF - MACD
    histogram = dif - macd
    
    return dif, macd, histogram


def check_limit_up(df: pd.DataFrame, cfg: ScreenerConfig) -> dict | None:
    """
    檢查在篩選日期範圍內是否出現漲停。
    漲停判斷：當日收盤價相對前日收盤價的漲幅 >= limit_up_pct
    """
    if len(df) < 2:
        return None

    df = df.copy()
    df["prev_close"] = df["close"].shift(1)
    df["chg_pct"]    = (df["close"] - df["prev_close"]) / df["prev_close"] * 100

    # 篩選出漲停日
    limit_up_days = df[df["chg_pct"] >= cfg.limit_up_pct]

    if limit_up_days.empty:
        return None

    # 回傳最近一次漲停的資訊
    latest = limit_up_days.iloc[-1]
    return {
        "date":     str(latest.name.date()) if hasattr(latest.name, "date") else str(latest.name),
        "close":    round(float(latest["close"]), 2),
        "chg_pct":  round(float(latest["chg_pct"]), 2),
        "count":    len(limit_up_days)   # 期間內漲停次數
    }


_stock_names_cache: dict[str, str] | None = None

def _load_stock_names() -> dict[str, str]:
    """從 stock_names_all.json 載入股票名稱對照表。"""
    global _stock_names_cache
    if _stock_names_cache is not None:
        return _stock_names_cache
    json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stock_names_all.json")
    if os.path.exists(json_path):
        with open(json_path, "r", encoding="utf-8") as f:
            _stock_names_cache = json.load(f)
    else:
        _stock_names_cache = {}
    return _stock_names_cache

def refresh_stock_names() -> int:
    """從 TWSE/TPEx API 更新 stock_names_all.json，回傳總筆數。"""
    import requests, urllib3
    urllib3.disable_warnings()
    all_names: dict[str, str] = {}
    # 上市
    try:
        r = requests.get("https://openapi.twse.com.tw/v1/exchangeReport/STOCK_DAY_ALL",
                         timeout=30, verify=False)
        if r.ok:
            for item in r.json():
                code = item.get("Code", "").strip()
                name = item.get("Name", "").strip()
                if code and name and len(code) == 4:
                    all_names[code] = name
    except Exception:
        pass
    # 上櫃
    try:
        r = requests.get("https://www.tpex.org.tw/openapi/v1/tpex_mainboard_daily_close_quotes",
                         timeout=30, verify=False)
        if r.ok:
            for item in r.json():
                code = item.get("SecuritiesCompanyCode", "").strip()
                name = item.get("CompanyName", "").strip()
                if code and name and len(code) == 4:
                    all_names[code] = name
    except Exception:
        pass
    if all_names:
        # 合併舊有（保留 API 查不到但之前有的）
        old = _load_stock_names()
        merged = {**old, **all_names}
        json_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stock_names_all.json")
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(merged, f, ensure_ascii=False, indent=2)
        global _stock_names_cache
        _stock_names_cache = merged
        return len(merged)
    return 0

def get_stock_name(ticker: str) -> str:
    """
    取得股票中文名稱。
    從 stock_names_all.json 載入（由 TWSE/TPEx API 產生）。
    """
    names = _load_stock_names()
    return names.get(ticker, ticker)


# --- 以下為舊硬編碼備份，已改用 JSON 檔 ---
_STOCK_NAMES_LEGACY = {
        "1101": "台泥", "1102": "亞泥", "1103": "嘉泥", "1104": "環泥",
        "1108": "幸福", "1109": "信大", "1110": "東泥", "1201": "味全",
        "1203": "味王", "1210": "大成", "1213": "大飲", "1215": "卜蜂",
        "1216": "統一", "1217": "愛之味", "1218": "泰山", "1219": "福壽",
        "1220": "台榮", "1225": "福懋油", "1227": "佳格", "1229": "聯華",
        "1231": "聯華食", "1232": "大統益", "1233": "天仁", "1234": "黑松",
        "1301": "台塑", "1303": "南亞", "1304": "台聚", "1305": "華夏",
        "1307": "三芳", "1308": "亞聚", "1309": "台達化", "1310": "台苯",
        "1312": "國喬", "1313": "聯成", "1314": "中石化", "1315": "達新",
        "1316": "上曜", "1319": "東陽", "1321": "大洋", "1323": "永裕",
        "1324": "地球", "1325": "恒大", "1326": "台化", "1337": "再生-KY",
        "1402": "遠東新", "1409": "新纖", "1410": "南染", "1413": "宏洲",
        "1414": "東和", "1416": "廣豐", "1417": "嘉裕", "1418": "東華",
        "1419": "新紡", "1423": "利華", "1432": "大魯閣", "1434": "福懋",
        "1436": "華友聯", "1437": "勤益控", "1438": "裕豐", "1439": "中和",
        "1440": "南紡", "1441": "大東", "1442": "名軒", "1443": "立益",
        "1444": "力麗", "1445": "大宇", "1446": "宏和", "1447": "力鵬",
        "1449": "佳和", "1451": "年興", "1452": "宏益", "1453": "大將",
        "1454": "台富", "1455": "集盛", "1456": "怡華", "1457": "宜進",
        "1459": "聯發", "1460": "宏遠", "1463": "強盛", "1464": "得力",
        "1465": "偉全", "1466": "聚隆", "1467": "南緯", "1468": "昶和",
        "1469": "正上", "1470": "大統染", "1471": "首利", "1472": "三洋紡",
        "1473": "台南", "1474": "弘裕", "1475": "本盟", "1476": "儒鴻",
        "1477": "聚陽", "1503": "士電", "1504": "東元", "1506": "正道",
        "1507": "永大", "1512": "瑞利", "1513": "中興電", "1514": "亞力",
        "1515": "力山", "1516": "川飛", "1517": "利奇", "1519": "華城",
        "1521": "大億", "1522": "堤維西", "1524": "耿鼎", "1525": "江申",
        "1526": "日馳", "1527": "鑽全", "1528": "恩德", "1529": "樂士",
        "1530": "亞崴", "1531": "高林股", "1532": "勤美", "1533": "車王電",
        "1535": "中宇", "1536": "和大", "1537": "廣隆", "1538": "正峰新",
        "1539": "巨庭", "1540": "喬福", "1541": "錩泰", "1558": "伸興",
        "1560": "中砂", "1568": "倉佑", "1582": "信錦", "1583": "程泰",
        "1589": "永冠-KY", "1590": "亞德客-KY", "1592": "英瑞-KY",
        "1598": "岱宇", "1603": "華電", "1604": "聲寶", "1605": "華新",
        "1608": "華榮", "1609": "大亞", "1611": "中電", "1612": "宏泰",
        "1613": "台一", "1614": "三洋電", "1615": "大山", "1616": "億泰",
        "1617": "榮星", "1618": "合機", "2002": "中鋼", "2006": "東和鋼鐵",
        "2008": "高興昌", "2009": "第一銅", "2010": "春源", "2012": "春雨",
        "2013": "中鋼構", "2014": "中鴻", "2015": "豐興", "2017": "官田鋼",
        "2020": "美亞", "2022": "聚亨", "2023": "燁輝", "2024": "志聯",
        "2025": "千興", "2027": "大成鋼", "2028": "威致", "2029": "盛餘",
        "2030": "彰源", "2031": "新光鋼", "2032": "新鋼", "2033": "佳大",
        "2034": "允強", "2038": "海光", "2059": "川湖", "2062": "橋椿",
        "2101": "南港", "2102": "泰豐", "2103": "台橡", "2104": "中橡",
        "2105": "正新", "2106": "建大", "2107": "厚生", "2108": "南帝",
        "2109": "華豐", "2114": "鑫永銓", "2115": "六暉-KY", "2201": "裕隆",
        "2204": "中華", "2206": "三陽工業", "2207": "和泰車", "2208": "台船",
        "2227": "裕日車", "2231": "為升", "2301": "光寶科", "2302": "麗正",
        "2303": "聯電", "2305": "全友", "2308": "台達電", "2312": "金寶",
        "2313": "華通", "2314": "台揚", "2316": "楠梓電", "2317": "鴻海",
        "2321": "東訊", "2323": "中環", "2324": "仁寶", "2327": "國巨",
        "2328": "廣宇", "2329": "華泰", "2330": "台積電", "2331": "精英",
        "2332": "友訊", "2337": "旺宏", "2338": "光罩", "2340": "光磊",
        "2342": "茂矽", "2344": "華邦電", "2345": "智邦", "2347": "聯強",
        "2349": "錸德", "2351": "順達科", "2352": "佳世達", "2353": "宏碁",
        "2354": "鴻準", "2355": "敬鵬", "2356": "英業達", "2357": "華碩",
        "2358": "廷鑫", "2359": "所羅門", "2360": "致茂", "2362": "藍天",
        "2363": "矽統", "2364": "倫飛", "2365": "昆盈", "2367": "燿華",
        "2368": "金像電", "2369": "菱生", "2371": "大同", "2373": "震旦行",
        "2374": "佳能", "2375": "智寶", "2376": "技嘉", "2377": "微星",
        "2379": "瑞昱", "2380": "虹光", "2382": "廣達", "2383": "台光電",
        "2385": "群光", "2387": "精元", "2388": "威盛", "2390": "云辰",
        "2392": "正崴", "2393": "億光", "2395": "研華", "2397": "友通",
        "2401": "凌陽", "2402": "毅嘉", "2404": "漢唐", "2405": "浩鑫",
        "2406": "國碩", "2408": "南亞科", "2409": "友達", "2412": "中華電",
        "2413": "環科", "2414": "精技", "2415": "錩新", "2417": "圓剛",
        "2419": "仲琦", "2420": "新巨", "2423": "固緯", "2424": "隴華",
        "2425": "承啟", "2426": "鼎元", "2427": "三商電", "2428": "興勤",
        "2429": "銘旺科", "2430": "燦坤", "2431": "聯昌", "2432": "倚天",
        "2433": "互盛電", "2434": "統懋", "2436": "偉詮電", "2438": "翔耀",
        "2439": "美律", "2440": "太空梭", "2441": "超豐", "2442": "新美齊",
        "2443": "新利虹", "2444": "兆勁", "2449": "京元電子", "2450": "神腦",
        "2451": "創見", "2453": "凌群", "2454": "聯發科", "2455": "全新",
        "2456": "奇力新", "2457": "飛宏", "2458": "義隆", "2459": "敦吉",
        "2460": "建通", "2461": "光群雷", "2462": "良得電", "2463": "雅捷",
        "2464": "盟立", "2465": "麗臺", "2466": "冠西電", "2467": "志聖",
        "2468": "華經", "2471": "資通", "2472": "立隆電", "2474": "可成",
        "2475": "華映", "2476": "鉅祥", "2477": "美隆電", "2478": "大毅",
        "2480": "敦陽科", "2481": "強茂", "2482": "連宇", "2483": "百容",
        "2484": "希華", "2485": "兆赫", "2486": "一詮", "2488": "漢平",
        "2489": "瑞軒", "2491": "吉祥全", "2492": "華新科", "2493": "揚博",
        "2495": "普安", "2496": "卓越", "2497": "怡利電", "2498": "宏達電",
        "2499": "東貝", "2501": "國建", "2504": "國產", "2505": "國揚",
        "2506": "太設", "2509": "全坤建", "2511": "太子", "2514": "龍邦",
        "2515": "中工", "2516": "新建", "2520": "冠德", "2524": "京城",
        "2527": "宏璟", "2528": "皇普", "2530": "華建", "2534": "宏盛",
        "2535": "達欣工", "2536": "宏普", "2537": "聯上發", "2538": "基泰",
        "2539": "櫻花建", "2540": "愛山林", "2542": "興富發", "2543": "皇昌",
        "2545": "皇翔", "2546": "根基", "2547": "日勝生", "2548": "華固",
        "2597": "潤弘", "2601": "益航", "2603": "長榮", "2605": "新興",
        "2606": "裕民", "2607": "榮運", "2608": "嘉里大榮", "2609": "陽明",
        "2610": "華航", "2611": "志信", "2612": "中航", "2613": "中櫃",
        "2614": "東森", "2615": "萬海", "2616": "山隆", "2617": "台航",
        "2618": "長榮航", "2630": "亞航", "2633": "台灣高鐵", "2634": "漢翔",
        "2636": "台驊投控", "2637": "慧洋-KY", "2642": "宅配通",
        "2701": "萬企", "2702": "華園", "2704": "國賓", "2705": "六福",
        "2706": "第一店", "2707": "晶華", "2712": "遠雄來", "2722": "夏都",
        "2723": "美食-KY", "2727": "王品", "2731": "雄獅", "2739": "寒舍",
        "2801": "彰銀", "2809": "京城銀", "2812": "台中銀", "2820": "華票",
        "2823": "中壽", "2832": "台產", "2834": "臺企銀", "2836": "高雄銀",
        "2838": "聯邦銀", "2845": "遠東銀", "2849": "安泰銀", "2850": "新產",
        "2851": "中再保", "2852": "第一保", "2855": "統一證", "2856": "元富證",
        "2867": "三商壽", "2880": "華南金", "2881": "富邦金", "2882": "國泰金",
        "2883": "開發金", "2884": "玉山金", "2885": "元大金", "2886": "兆豐金",
        "2887": "台新金", "2888": "新光金", "2889": "國票金", "2890": "永豐金",
        "2891": "中信金", "2892": "第一金", "2901": "欣欣", "2903": "遠百",
        "2904": "匯僑", "2905": "三商", "2906": "高林", "2908": "特力",
        "2910": "統領", "2911": "麗嬰房", "2912": "統一超", "2913": "農林",
        "2915": "潤泰全", "2923": "鼎固-KY", "2929": "淘帝-KY",
        "3002": "歐格", "3003": "健和興", "3004": "豐達科", "3005": "神基",
        "3006": "晶豪科", "3008": "大立光", "3010": "華立", "3011": "今皓",
        "3013": "晟銘電", "3014": "聯陽", "3015": "全漢", "3016": "嘉晶",
        "3017": "奇鋐", "3018": "同開", "3019": "亞光", "3021": "鴻名",
        "3022": "威強電", "3023": "信邦", "3024": "憶聲", "3025": "星通",
        "3026": "禾伸堂", "3027": "盛達", "3028": "增你強", "3029": "零壹",
        "3030": "德律", "3031": "佰鴻", "3032": "偉訓", "3033": "威健",
        "3034": "聯詠", "3035": "智原", "3036": "文曄", "3037": "欣興",
        "3038": "全台", "3040": "遠見", "3041": "揚智", "3042": "晶技",
        "3043": "科風", "3044": "健鼎", "3045": "台灣大", "3046": "建碁",
        "3047": "訊舟", "3048": "益登", "3049": "和鑫", "3050": "鈺德",
        "3051": "力特", "3052": "夆典", "3054": "立萬利", "3055": "蔚華科",
        "3056": "總太", "3057": "喬鼎", "3058": "立德", "3059": "華晶科",
        "3060": "銘異", "3062": "建漢", "3090": "日電貿", "3092": "鴻碩",
        "3093": "港建", "3094": "聯傑", "3130": "一零四", "3149": "正達",
        "3164": "景岳", "3167": "大量", "3189": "景碩", "3231": "緯創",
        "3257": "虹冠電", "3266": "昇陽", "3296": "勝德", "3305": "昇貿",
        "3308": "聯德", "3312": "弘憶股", "3338": "泰碩", "3346": "麗清",
        "3356": "奇偶", "3376": "新日興", "3380": "明泰", "3406": "玉晶光",
        "3443": "創意", "3454": "晶睿", "3481": "群創", "3501": "維熹",
        "3504": "揚明光", "3515": "華擎", "3528": "安馳", "3530": "晶相光",
        "3532": "台勝科", "3533": "嘉澤", "3535": "晶彩科", "3545": "敦泰",
        "3557": "嘉威", "3583": "辛耘", "3583": "辛耘", "3588": "通嘉",
        "3591": "艾笛森", "3592": "瑞鼎", "3593": "力銘", "3594": "磐儀",
        "3596": "智易", "3605": "宏致", "3607": "谷崧", "3617": "碩天",
        "3653": "健策", "3665": "貿聯-KY", "3669": "圓展", "3673": "TPK-KY",
        "3679": "新至陞", "3682": "亞太電", "3686": "達能", "3691": "碩禾",
        "3694": "海華", "3701": "大眾控", "3702": "大聯大", "3703": "欣陸",
        "3704": "合勤控", "3705": "永信", "3706": "神達", "3708": "上緯投控",
        "3711": "日月光投控", "3714": "富采", "4104": "佳醫", "4106": "雃博",
        "4108": "懷特", "4119": "旭富", "4133": "亞諾法", "4137": "麗豐-KY",
        "4142": "國光生", "4144": "康聯-KY", "4306": "炎洲", "4415": "台原藥",
        "4438": "廣越", "4526": "東台", "4529": "淳紳", "4530": "宏易",
        "4532": "瑞智", "4555": "氣立", "4712": "南璋", "4722": "國精化",
        "4725": "信昌化", "4735": "豪展", "4737": "華廣", "4746": "台耀",
        "4904": "遠傳", "4906": "正文", "4915": "致伸", "4916": "事欣科",
        "4919": "新唐", "4927": "泰鼎-KY", "4935": "茂林-KY", "4938": "和碩",
        "4952": "凌通", "4956": "光鋐", "4958": "臻鼎-KY", "4960": "誠美材",
        "4968": "立積", "4976": "佳凌", "4989": "榮科", "4994": "傳奇",
        "4999": "鑫禾", "5007": "三星", "5203": "訊連", "5215": "科嘉-KY",
        "5225": "東科-KY", "5234": "達興材料", "5243": "乙盛-KY",
        "5258": "虹堡", "5269": "祥碩", "5285": "界霖", "5289": "宜鼎",
        "5305": "敦南", "5347": "世界", "5371": "中光電", "5388": "中磊",
        "5434": "崇越", "5469": "瀚宇博", "5471": "松翰", "5484": "慧友",
        "5515": "建國", "5519": "隆大", "5522": "遠雄", "5525": "順天",
        "5531": "鄉林", "5533": "皇鼎", "5534": "長虹", "5536": "聖暉",
        "5607": "遠雄港", "5608": "四維航", "5871": "中租-KY", "5876": "上海商銀",
        "5880": "合庫金", "6005": "群益證", "6024": "群益期", "6115": "鎰勝",
        "6116": "彩晶", "6117": "迎廣", "6133": "金橋", "6136": "富爾特",
        "6139": "亞翔", "6141": "柏承", "6142": "友勁", "6152": "百一",
        "6153": "嘉聯益", "6155": "鈞寶", "6164": "華興", "6165": "捷泰",
        "6166": "凌華", "6168": "宏齊", "6176": "瑞儀", "6177": "達麗",
        "6183": "關貿", "6184": "大豐電", "6191": "精成科", "6197": "佳必琪",
        "6202": "盛群", "6205": "詮欣", "6206": "飛捷", "6209": "今國光",
        "6213": "聯茂", "6214": "精誠", "6215": "和椿", "6239": "力成",
        "6243": "迅杰", "6257": "矽格", "6269": "台郡", "6271": "同欣電",
        "6277": "宏正", "6278": "台表科", "6281": "全國電", "6285": "啟碁",
        "6405": "悅城", "6412": "群電", "6414": "樺漢", "6415": "矽力-KY",
        "6416": "瑞祺電通", "6443": "元晶", "6451": "訊芯-KY", "6477": "安集",
        "6504": "南六", "6505": "台塑化", "6531": "愛普", "6533": "晶心科",
        "6538": "倉和", "6552": "易華電", "6573": "虹揚-KY", "6579": "研揚",
        "6605": "帝寶", "6669": "緯穎", "6702": "興航", "8011": "台通",
        "8016": "矽創", "8021": "尖點", "8028": "昇陽半導體", "8033": "雷虎",
        "8046": "南電", "8070": "長華", "8072": "陞泰", "8081": "致新",
        "8101": "華冠", "8103": "瀚荃", "8105": "凌巨", "8110": "華東",
        "8131": "福懋科", "8150": "南茂", "8163": "達方", "8171": "天宇",
        "8201": "無敵", "8210": "勤誠", "8213": "志超", "8215": "明基材",
        "8261": "富鼎", "8271": "宇瞻", "8299": "群聯", "8341": "日友",
        "8349": "恒耀", "8358": "金居", "8367": "建新國際", "8374": "羅昇",
        "8380": "昱泉", "8404": "百和興業-KY", "8411": "福貞-KY",
        "8422": "可寧衛", "8427": "基勝-KY", "8429": "金麗-KY",
        "8454": "富邦媒", "8463": "潤泰材", "8464": "億豐", "8466": "美吉吉-KY",
        "8478": "東哥遊艇", "8480": "泰昇-KY", "8481": "政伸", "8488": "吉源-KY",
        "8926": "台汽電", "8940": "新天地", "9103": "美德醫療-DR",
        "9105": "泰金寶-DR", "9110": "九豪", "9136": "巨騰-DR",
        "9188": "精熙-DR", "9802": "鈺齊-KY", "9904": "寶成",
        "9905": "大華", "9906": "欣巴巴", "9907": "統一實", "9908": "大台北",
        "9910": "豐泰", "9911": "櫻花", "9912": "偉聯", "9914": "美利達",
        "9917": "中保科", "9918": "欣天然", "9919": "康那香", "9921": "巨大",
        "9924": "福興", "9925": "新保", "9926": "新海", "9927": "泰銘",
        "9928": "中視", "9929": "秋雨", "9930": "中聯資源", "9931": "欣高",
        "9933": "中鼎", "9934": "成霖", "9935": "慶豐富", "9937": "全國",
        "9938": "百和", "9939": "宏全", "9940": "信義", "9941": "裕融",
        "9942": "茂順", "9943": "好樂迪", "9944": "新麗", "9945": "潤泰新",
        "9946": "三發地產", "9955": "佳龍", "9958": "世紀鋼",
        # 使用者補充的股票名稱
        "1235": "興泰", "1236": "宏亞", "1256": "鮮活果汁-KY", "1339": "昭輝", "1341": "富林-KY",
        "1342": "八貫", "1435": "中福", "1623": "大東電", "1702": "南僑", "1707": "葡萄王",
        "1708": "東鹼", "1711": "永光", "1712": "興農", "1713": "國化", "1723": "中碳",
        "1726": "永記", "1730": "花仙子", "1731": "美吾華", "1735": "日勝化", "1736": "喬山",
        "1776": "展宇", "1795": "美時", "1806": "冠軍", "1808": "潤隆", "1809": "中釉",
        "1817": "凱撒衛", "2211": "長榮鋼", "2228": "劍麟", "2233": "宇隆", "2239": "英利-KY",
        "2243": "宏旭-KY", "2247": "汎德永業", "2248": "華勝-KY", "2348": "海悅", "2421": "建準",
        "2753": "八方雲集", "2762": "世界健身-KY", "2816": "旺旺保", "2897": "王道銀行",
        "2945": "三商家購", "3138": "耀登", "3150": "鈺寶-創", "3168": "眾福科", "3209": "全科",
        "3229": "晟鈦", "3413": "京鼎", "3419": "譁裕", "3550": "聯穎", "3576": "聯合再生",
        "3661": "世芯-KY", "3712": "永崴投控", "4148": "全宇生技-KY", "4439": "冠星-KY",
        "4440": "宜新實業", "4536": "拓凱", "4551": "智伸科", "4557": "永新-KY", "4560": "強信-KY",
        "4564": "元翎", "4569": "六方科-KY", "4572": "駐龍", "4581": "光隆精密-KY", "4720": "德淵",
        "4736": "泰博", "4763": "材料-KY", "4766": "南寶", "4930": "燦星網", "4942": "嘉彰",
        "4949": "有成精密", "4961": "天鈺", "4977": "眾達-KY", "5284": "jpp-KY", "5288": "豐祥-KY",
        "5546": "永固-KY", "5706": "鳳凰", "6108": "競國", "6128": "上福", "6189": "豐藝",
        "6201": "亞弘電", "6216": "居易", "6235": "華孚", "6272": "驊陞", "6409": "旭隼",
        "6426": "統新", "6431": "光麗-KY", "6442": "光聖", "6446": "藥華藥", "6456": "GIS-KY",
        "6464": "台數科", "6472": "保瑞", "6491": "晶碩", "6515": "穎崴", "6518": "康科特", "6525": "捷敏-KY",
        "6581": "鋼聯", "6585": "鼎基", "6591": "動力-KY", "6606": "建德工業", "6655": "科定",
        "6670": "復盛應用", "6671": "三能-KY", "6691": "洋基工程", "6715": "嘉基", "6719": "力智",
        "6722": "輝創", "6743": "安普新", "6754": "匯橋設計", "6757": "台灣虎航", "6768": "志強-KY",
        "6771": "平和環保-創", "6782": "視陽", "6790": "永豐實", "6796": "晉弘", "6805": "富世達",
        "6807": "峰源-KY", "6830": "汎銓", "6835": "圓裕", "6854": "錼創科技-KY創", "6863": "永道-KY",
        "6885": "全福生技", "6887": "寶綠特-KY", "6902": "GOGOLOOK", "6918": "愛派司", "6919": "康霈",
        "6921": "嘉雨思-創", "6923": "中台", "6924": "榮惠-KY創", "6933": "AMAX-KY", "6934": "心誠鎂",
        "6937": "天虹", "6944": "兆聯實業", "6949": "沛爾生醫-創", "6951": "青新-創", "6955": "邦睿生技-創",
        "6957": "裕慶-KY", "6965": "中傑-KY", "6988": "威力暘-創", "7631": "聚賢研發-創",
        "7705": "三商餐飲", "7711": "永擎", "7732": "金興精密", "7749": "意騰-KY", "7750": "新代",
        "7765": "中華資安", "7769": "鴻勁", "7788": "松川精密", "7795": "長廣", "7799": "禾榮科",
        "8442": "威宏-KY", "8462": "柏文", "8467": "波力-KY", "8476": "台境", "8482": "商億-KY",
        "8487": "愛爾達-創", "8996": "高力",
        # ---- 以下為深度掃描補充（2026-02-27）----
        "1338": "廣華-KY", "1340": "勝悅-KY", "1563": "巧新", "1587": "吉茂", "1597": "直得",
        "1626": "艾美特-KY", "1709": "和益", "1710": "東聯", "1714": "和桐", "1717": "長興",
        "1718": "中纖", "1720": "生達", "1721": "三晃", "1722": "台肥", "1725": "元禎",
        "1727": "中華化", "1732": "毛寶", "1733": "五鼎", "1734": "杏輝", "1737": "臺鹽",
        "1752": "南光", "1760": "寶齡富錦", "1762": "中化生", "1773": "勝一", "1783": "和康生",
        "1786": "科妍", "1789": "神隆", "1802": "台玻", "1805": "寶徠", "1810": "和成",
        "1903": "士紙", "1904": "正隆", "1905": "華紙", "1906": "寶隆", "1907": "永豐餘",
        "1909": "榮成", "2007": "燁興", "2049": "上銀", "2069": "運錩", "2236": "百達-KY",
        "2241": "艾姆勒", "2250": "IKKA-KY", "2254": "巨鎧精密-創", "2258": "鴻華先進-創",
        "2399": "映泰", "2645": "長榮航太", "2646": "星宇航空", "2748": "雲品", "2939": "永邑-KY",
        "3135": "凌航", "3311": "閎暉", "3321": "同泰", "3416": "融程電", "3432": "台端",
        "3437": "榮創", "3447": "展達", "3450": "聯鈞", "3494": "誠研", "3518": "柏騰",
        "3543": "州巧", "3563": "牧德", "3622": "洋華", "3645": "達邁", "3652": "精聯",
        "3715": "定穎投控", "3716": "中化控股", "3717": "聯嘉投控", "4155": "訊映", "4164": "承業醫",
        "4190": "佐登-KY", "4414": "如興", "4426": "利勤", "4441": "振大環球", "4540": "全球傳動",
        "4545": "銘鈺", "4552": "力達-KY", "4562": "穎漢", "4566": "時碩工業", "4571": "鈞興-KY",
        "4576": "大銀微系統", "4583": "台灣精銳", "4585": "達明", "4588": "玖鼎電力", "4590": "富田-創",
        "4739": "康普", "4755": "三福化", "4764": "雙鍵", "4770": "上品", "4771": "望隼",
        "4807": "日成-KY", "4912": "聯德控股-KY", "4934": "太極", "4943": "康控-KY", "4967": "十銓",
        "5222": "全訊", "5244": "弘凱", "5283": "禾聯碩", "5292": "華懋", "5306": "桂盟",
        "5521": "工信", "5538": "東明-KY", "5906": "台南-KY", "5907": "大洋-KY", "6112": "邁達特",
        "6120": "達運", "6192": "巨路", "6196": "帆宣", "6224": "聚鼎", "6225": "天瀚",
        "6226": "光鼎", "6230": "尼得科超眾", "6282": "康舒", "6283": "淳安", "6423": "億而得",
        "6438": "迅得", "6449": "鈺邦", "6526": "達發", "6534": "正瀚-創", "6541": "泰福-KY",
        "6550": "北極星藥業-KY", "6558": "興能高", "6582": "申豐", "6589": "台康生技",
        "6592": "和潤企業", "6598": "ABC-KY", "6614": "資拓宏宇", "6625": "必應",
        "6641": "基士德-KY", "6645": "金萬林-創", "6657": "華安", "6658": "聯策",
        "6666": "羅麗芬-KY", "6668": "中揚光", "6672": "騰輝電子-KY", "6674": "鋐寶科技",
        "6689": "伊雲谷", "6695": "芯鼎", "6698": "旭暉應材", "6706": "惠特", "6742": "澤米",
        "6753": "龍德造船", "6756": "威鋒電子", "6770": "力積電", "6776": "展碁國際",
        "6781": "AES-KY", "6789": "采鈺", "6792": "詠業", "6794": "向榮生技", "6799": "來頡",
        "6806": "森崴能源", "6831": "邁科", "6834": "天二科技", "6838": "台新藥",
        "6861": "睿生光電", "6862": "三集瑞-KY", "6869": "雲豹能源", "6873": "泓德能源",
        "6890": "來億-KY", "6901": "鑽石投資", "6906": "現觀科", "6909": "創控",
        "6914": "阜爾運通", "6916": "華凌", "6928": "攸泰科技", "6931": "青松健康",
        "6936": "永鴻生技", "6952": "大武山", "6958": "日盛台駿", "6962": "奕力-KY",
        "6969": "成信實業-創", "6994": "富威電力", "7610": "聯友金屬-創", "7721": "微程式",
        "7722": "LINEPAY", "7730": "暉盛-創", "7736": "虎山", "7740": "熙特爾-創",
        "7780": "大研生醫", "7786": "東方風能", "7791": "皇家可口", "7823": "奧義賽博-KY創",
        "8039": "台虹", "8045": "達運光電", "8104": "錸寶", "8112": "至上", "8114": "振樺電",
        "8162": "微矽電子-創", "8222": "寶一", "8249": "菱光", "8438": "昶昕", "8443": "阿瘦",
        "8473": "山林水", "8499": "鼎炫-KY", "9902": "台火",
}


def check_dividend(cfg: ScreenerConfig, ticker: str, info: dict | None = None) -> dict | None:
    """
    計算殖利率。
    不直接用 dividendYield（格式不穩定，同一個版本裡不同股票回傳方式都不同）。
    改用 dividendRate（每股年配息金額）/ 現價 * 100 來自己算，語義明確不會亂。
    """
    try:
        if info is None:
            info = _fetch_ticker_info(ticker)
        if not info:
            return None

        price = info.get("regularMarketPrice") or info.get("previousClose")
        # dividendRate = 每股年配息金額（元），語義固定不變
        div_rate = info.get("dividendRate") or info.get("trailingAnnualDividendRate")

        if not price or not div_rate or price <= 0 or div_rate <= 0:
            return None

        yield_pct = float(div_rate) / float(price) * 100

        if yield_pct >= cfg.min_dividend_yield:
            return {
                "price":          round(float(price), 2),
                "dividend_yield": round(yield_pct, 2),
                "dividend_rate":  round(float(div_rate), 2)   # 順便顯示每股配息
            }

        return None

    except Exception:
        return None


def check_rsi_ma(df: pd.DataFrame, cfg: ScreenerConfig) -> dict | None:
    """
    RSI + MA 策略篩選。
    進場訊號：收盤價 > MA 且 RSI < 超賣閾值
    """
    if len(df) < max(cfg.rsi_period, cfg.ma_period):
        return None

    df = df.copy()
    df["ma"]    = df["close"].rolling(cfg.ma_period).mean()
    df["rsi"]   = calc_rsi(df["close"], cfg.rsi_period)

    latest = df.iloc[-1]

    signal = (latest["close"] > latest["ma"]) and (latest["rsi"] < cfg.rsi_oversold)

    if signal:
        return {
            "close":  round(float(latest["close"]), 2),
            "ma":     round(float(latest["ma"]), 2),
            "rsi":    round(float(latest["rsi"]), 2)
        }
    return None


def check_bbands(df: pd.DataFrame, cfg: ScreenerConfig) -> dict | None:
    """
    布林通道開口向上篩選。
    判斷條件：
    1. 通道寬度擴大（相對前一日增加 >= expand_min）
    2. 中軌（MA）向上
    3. 價格接近或突破上軌
    """
    if len(df) < cfg.bbands_period + 5:
        return None

    df = df.copy()
    
    # 計算布林通道
    df["bb_middle"] = df["close"].rolling(cfg.bbands_period).mean()
    df["bb_std"] = df["close"].rolling(cfg.bbands_period).std()
    df["bb_upper"] = df["bb_middle"] + (cfg.bbands_std * df["bb_std"])
    df["bb_lower"] = df["bb_middle"] - (cfg.bbands_std * df["bb_std"])
    
    # 通道寬度（上軌 - 下軌）
    df["bb_width"] = df["bb_upper"] - df["bb_lower"]
    
    # 通道寬度變化率
    df["bb_width_change"] = df["bb_width"].pct_change()
    
    # 取最近和前一天
    if len(df) < 2:
        return None
    
    latest = df.iloc[-1]
    prev = df.iloc[-2]
    
    # 檢查條件
    # 1. 通道寬度擴大
    width_expanding = latest["bb_width_change"] >= cfg.bbands_expand_min
    
    # 2. 中軌向上（最近 3 天 MA 向上）
    ma_uptrend = latest["bb_middle"] > prev["bb_middle"]
    
    # 3. 價格接近上軌（距離上軌小於通道寬度的 30%）
    price_near_upper = (latest["bb_upper"] - latest["close"]) / latest["bb_width"] < 0.3
    
    # 三個條件都滿足
    if width_expanding and ma_uptrend and price_near_upper:
        return {
            "close": round(float(latest["close"]), 2),
            "bb_upper": round(float(latest["bb_upper"]), 2),
            "bb_middle": round(float(latest["bb_middle"]), 2),
            "bb_lower": round(float(latest["bb_lower"]), 2),
            "bb_width": round(float(latest["bb_width"]), 2),
            "width_change_pct": round(float(latest["bb_width_change"]) * 100, 2)
        }
    
    return None


def check_kd_golden(df: pd.DataFrame, cfg: ScreenerConfig) -> dict | None:
    """
    KD 黃金交叉篩選（做多訊號）
    條件：K 值向上突破 D 值
    """
    if len(df) < cfg.kd_period + 5:
        return None
    
    df = df.copy()
    k, d = calc_kd(df, cfg.kd_period)
    df["k"] = k
    df["d"] = d
    
    if len(df) < 2:
        return None
    
    latest = df.iloc[-1]
    prev = df.iloc[-2]
    
    # K 向上突破 D（黃金交叉）
    golden_cross = (latest["k"] > latest["d"]) and (prev["k"] <= prev["d"])
    
    # 或者 K 已經在 D 上方且持續向上（確認趨勢）
    k_above_d = (latest["k"] > latest["d"]) and (latest["k"] > prev["k"])
    
    if golden_cross or k_above_d:
        return {
            "close": round(float(latest["close"]), 2),
            "k": round(float(latest["k"]), 2),
            "d": round(float(latest["d"]), 2),
            "k_trend": "黃金交叉" if golden_cross else "K>D且上升"
        }
    
    return None


def check_bb_upper(df: pd.DataFrame, cfg: ScreenerConfig) -> dict | None:
    """
    布林通道接近上軌篩選（強勢突破）
    條件：價格接近或突破上軌，顯示強勢
    """
    if len(df) < cfg.bbands_period + 5:
        return None
    
    df = df.copy()
    df["bb_middle"] = df["close"].rolling(cfg.bbands_period).mean()
    df["bb_std"] = df["close"].rolling(cfg.bbands_period).std()
    df["bb_upper"] = df["bb_middle"] + (cfg.bbands_std * df["bb_std"])
    df["bb_lower"] = df["bb_middle"] - (cfg.bbands_std * df["bb_std"])
    df["bb_width"] = df["bb_upper"] - df["bb_lower"]
    
    if len(df) < 2:
        return None
    
    latest = df.iloc[-1]
    prev = df.iloc[-2]
    
    # 價格接近上軌（在上軌下方 10% 以內）或已突破
    distance_to_upper = (latest["bb_upper"] - latest["close"]) / latest["bb_width"]
    price_near_upper = distance_to_upper < 0.1 or distance_to_upper < 0  # 接近或突破
    
    # 中軌向上趨勢
    ma_uptrend = latest["bb_middle"] > prev["bb_middle"]
    
    if price_near_upper and ma_uptrend:
        breakthrough = "突破上軌" if latest["close"] > latest["bb_upper"] else "接近上軌"
        return {
            "close": round(float(latest["close"]), 2),
            "bb_upper": round(float(latest["bb_upper"]), 2),
            "bb_middle": round(float(latest["bb_middle"]), 2),
            "bb_lower": round(float(latest["bb_lower"]), 2),
            "status": breakthrough
        }
    
    return None


def check_resonance(df: pd.DataFrame, cfg: ScreenerConfig) -> dict | None:
    """
    技術指標共振策略：KD + RSI + 布林通道 + 量能 + MACD
    五個指標必須同時出現買進訊號才算命中。
    
    買進訊號條件：
    1. KD 指標：K < 20 且 D < 20（超賣），或 K 向上突破 D（黃金交叉）
    2. RSI 指標：RSI < 30（超賣）
    3. 布林通道：價格接近下軌（準備反彈）且通道不縮窄
    4. 量能：今日成交量 >= 5日均量 * 1.5（量能放大）
    5. MACD：DIF 向上突破 MACD（黃金交叉）或柱狀圖由負轉正
    """
    # 需要足夠的歷史數據（MACD 需要至少 slow + signal + 10 天）
    required_days = max(cfg.kd_period, cfg.rsi_period, cfg.bbands_period, cfg.macd_slow + cfg.macd_signal) + 10
    if len(df) < required_days:
        return None

    df = df.copy()
    
    # ========== 1. 計算 KD 指標 ==========
    k, d = calc_kd(df, cfg.kd_period)
    df["k"] = k
    df["d"] = d
    
    # ========== 2. 計算 RSI 指標 ==========
    df["rsi"] = calc_rsi(df["close"], cfg.rsi_period)
    
    # ========== 3. 計算布林通道 ==========
    df["bb_middle"] = df["close"].rolling(cfg.bbands_period).mean()
    df["bb_std"] = df["close"].rolling(cfg.bbands_period).std()
    df["bb_upper"] = df["bb_middle"] + (cfg.bbands_std * df["bb_std"])
    df["bb_lower"] = df["bb_middle"] - (cfg.bbands_std * df["bb_std"])
    df["bb_width"] = df["bb_upper"] - df["bb_lower"]
    
    # ========== 4. 計算成交量 ==========
    df["vol_ma5"] = df["volume"].rolling(5).mean()
    
    # ========== 5. 計算 MACD 指標 ==========
    dif, macd, histogram = calc_macd(df["close"], cfg.macd_fast, cfg.macd_slow, cfg.macd_signal)
    df["dif"] = dif
    df["macd"] = macd
    df["histogram"] = histogram
    
    # 取最近兩天數據
    if len(df) < 2:
        return None
    
    latest = df.iloc[-1]
    prev = df.iloc[-2]
    
    # ========== 檢查五大訊號 ==========
    
    # 訊號 1：KD 超賣且黃金交叉
    kd_oversold = (latest["k"] < cfg.kd_oversold_k) and (latest["d"] < cfg.kd_oversold_d)
    kd_golden_cross = (latest["k"] > latest["d"]) and (prev["k"] <= prev["d"])  # K 突破 D
    signal_kd = kd_oversold or kd_golden_cross  # 超賣或黃金交叉其一即可
    
    # 訊號 2：RSI 超賣
    signal_rsi = latest["rsi"] < cfg.rsi_oversold
    
    # 訊號 3：布林通道 - 價格接近下軌（距離下軌 < 通道寬度的 30%）
    price_near_lower = (latest["close"] - latest["bb_lower"]) / latest["bb_width"] < 0.3
    bb_not_shrinking = latest["bb_width"] >= prev["bb_width"] * 0.9  # 通道沒有急縮
    signal_bb = price_near_lower and bb_not_shrinking
    
    # 訊號 4：量能放大
    signal_volume = latest["volume"] >= (latest["vol_ma5"] * cfg.volume_surge_ratio)
    
    # 訊號 5：MACD 黃金交叉或柱狀圖轉正
    macd_golden_cross = (latest["dif"] > latest["macd"]) and (prev["dif"] <= prev["macd"])  # DIF 突破 MACD
    histogram_positive = (latest["histogram"] > 0) and (prev["histogram"] <= 0)  # 柱狀圖由負轉正
    signal_macd = macd_golden_cross or histogram_positive
    
    # ========== 五大訊號必須全部成立 ==========
    if signal_kd and signal_rsi and signal_bb and signal_volume and signal_macd:
        return {
            "close": round(float(latest["close"]), 2),
            "k": round(float(latest["k"]), 2),
            "d": round(float(latest["d"]), 2),
            "rsi": round(float(latest["rsi"]), 2),
            "bb_upper": round(float(latest["bb_upper"]), 2),
            "bb_middle": round(float(latest["bb_middle"]), 2),
            "bb_lower": round(float(latest["bb_lower"]), 2),
            "volume": int(latest["volume"]),
            "vol_ma5": int(latest["vol_ma5"]),
            "volume_ratio": round(float(latest["volume"] / latest["vol_ma5"]), 2),
            "dif": round(float(latest["dif"]), 2),
            "macd": round(float(latest["macd"]), 2),
            "histogram": round(float(latest["histogram"]), 2),
            # 記錄各訊號狀態（除錯用）
            "signals": {
                "kd": True,
                "rsi": True,
                "bb": True,
                "volume": True,
                "macd": True
            }
        }
    
    return None


def check_market_metrics(cfg: ScreenerConfig, ticker: str, df: pd.DataFrame | None = None, info: dict | None = None) -> dict | None:
    """
    檢查市場指標：成交量、市值、本益比。
    同時從 yfinance info 和歷史數據取得。
    """
    try:
        if info is None:
            info = _fetch_ticker_info(ticker)
        if not info:
            return None

        # 市值（單位：元 → 億元）
        market_cap_raw = info.get("marketCap")
        market_cap = round(float(market_cap_raw) / 1e8, 2) if market_cap_raw else None

        # 本益比
        pe_ratio = info.get("trailingPE") or info.get("forwardPE")
        pe_ratio = round(float(pe_ratio), 2) if pe_ratio and pe_ratio > 0 else None

        # 平均成交量（張）：從最近歷史數據算
        avg_volume = None
        if df is not None and len(df) > 0:
            # volume 單位是股，除以 1000 = 張
            avg_volume = round(df["volume"].tail(5).mean() / 1000, 0)

        # 檢查篩選條件
        pass_filter = True

        if cfg.filter_market_cap and market_cap:
            if not (cfg.min_market_cap <= market_cap <= cfg.max_market_cap):
                pass_filter = False

        if cfg.filter_pe_ratio and pe_ratio:
            if not (cfg.min_pe_ratio <= pe_ratio <= cfg.max_pe_ratio):
                pass_filter = False

        if cfg.filter_volume and avg_volume:
            if avg_volume < cfg.min_volume:
                pass_filter = False

        if not pass_filter:
            return None

        # 回傳數據（即使沒開篩選也回傳，供排序用）
        return {
            "market_cap": market_cap,
            "pe_ratio": pe_ratio,
            "avg_volume": int(avg_volume) if avg_volume else None
        }

    except Exception:
        return None


# ============================================================
# 結果數據結構
# ============================================================

@dataclass
class StockResult:
    ticker:       str
    name:         str = ""            # 股票中文名稱
    limit_up:     dict | None = None
    dividend:     dict | None = None
    rsi_ma:       dict | None = None
    bbands:       dict | None = None   # 布林通道開口向上
    resonance:    dict | None = None   # 技術指標共振（KD+RSI+BB+量+MACD）
    kd_golden:    dict | None = None   # KD 黃金交叉（做多訊號）
    bb_upper:     dict | None = None   # 布林通道接近上軌（強勢突破）
    market:       dict | None = None   # 市場指標：市值、本益比、成交量


# ============================================================
# CSV 資料收集（取代 data_downloader.py）
# ============================================================

def _build_csv_row(ticker: str, name: str, df: pd.DataFrame, info: dict | None) -> dict | None:
    """從已下載的 df 和 info 建構一筆 CSV 資料列（最後一天的指標快照）"""
    try:
        if len(df) < 30:
            return None

        df = df.copy()

        # 計算技術指標
        k, d = calc_kd(df, period=9)
        df["k"] = k
        df["d"] = d
        df["rsi"] = calc_rsi(df["close"], period=14)

        dif, macd, histogram = calc_macd(df["close"], fast=12, slow=26, signal=9)
        df["dif"] = dif
        df["macd"] = macd
        df["histogram"] = histogram

        df["bb_middle"] = df["close"].rolling(20).mean()
        df["bb_std"] = df["close"].rolling(20).std()
        df["bb_upper"] = df["bb_middle"] + (2.0 * df["bb_std"])
        df["bb_lower"] = df["bb_middle"] - (2.0 * df["bb_std"])

        df["vol_ma5"] = df["volume"].rolling(5).mean()
        df["volume_ratio"] = df["volume"] / df["vol_ma5"]

        latest = df.iloc[-1]

        # 取得實際交易日（yfinance DataFrame index 是交易日期）
        trading_date = latest.name
        if hasattr(trading_date, 'strftime'):
            trading_date_str = trading_date.strftime('%Y-%m-%d')
        else:
            trading_date_str = str(trading_date)[:10]

        # 從 info 取得市場資料
        market_cap = None
        pe_ratio = None
        price = round(float(latest["close"]), 2)
        div_rate = None
        div_yield = None

        if info:
            market_cap_raw = info.get("marketCap")
            market_cap = round(float(market_cap_raw) / 1e8, 2) if market_cap_raw else None

            pe = info.get("trailingPE") or info.get("forwardPE")
            pe_ratio = round(float(pe), 2) if pe and pe > 0 else None

            p = info.get("regularMarketPrice") or info.get("previousClose")
            if p:
                price = round(float(p), 2)

            dr = info.get("dividendRate") or info.get("trailingAnnualDividendRate")
            div_rate = round(float(dr), 2) if dr and dr > 0 else None

            if price and div_rate and price > 0 and div_rate > 0:
                div_yield = round((div_rate / price) * 100, 2)

        return {
            "股票代號": ticker,
            "股票名稱": name,
            "收盤價": price,
            "開盤價": round(float(latest["open"]), 2),
            "最高價": round(float(latest["high"]), 2),
            "最低價": round(float(latest["low"]), 2),
            "K值": round(float(latest["k"]), 2) if not pd.isna(latest["k"]) else None,
            "D值": round(float(latest["d"]), 2) if not pd.isna(latest["d"]) else None,
            "RSI": round(float(latest["rsi"]), 2) if not pd.isna(latest["rsi"]) else None,
            "DIF": round(float(latest["dif"]), 2) if not pd.isna(latest["dif"]) else None,
            "MACD": round(float(latest["macd"]), 2) if not pd.isna(latest["macd"]) else None,
            "柱狀圖": round(float(latest["histogram"]), 2) if not pd.isna(latest["histogram"]) else None,
            "BB上軌": round(float(latest["bb_upper"]), 2) if not pd.isna(latest["bb_upper"]) else None,
            "BB中軌": round(float(latest["bb_middle"]), 2) if not pd.isna(latest["bb_middle"]) else None,
            "BB下軌": round(float(latest["bb_lower"]), 2) if not pd.isna(latest["bb_lower"]) else None,
            "成交量張": int(latest["volume"] / 1000),
            "5日均量張": int(latest["vol_ma5"] / 1000) if not pd.isna(latest["vol_ma5"]) else None,
            "量比": round(float(latest["volume_ratio"]), 2) if not pd.isna(latest["volume_ratio"]) else None,
            "市值億": market_cap,
            "本益比": pe_ratio,
            "每股配息": div_rate,
            "殖利率": div_yield,
            "交易日": trading_date_str,
            "更新時間": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    except Exception:
        return None


def _save_stock_csv(csv_data: list[dict]) -> str | None:
    """將收集的股票資料存成 CSV 檔案"""
    if not csv_data:
        return None

    df_all = pd.DataFrame(csv_data)
    csv_filename = f"stock_data_{datetime.now().strftime('%Y%m%d')}.csv"
    df_all.to_csv(csv_filename, index=False, encoding="utf-8-sig")

    print(f"\n📊 CSV 數據已儲存：{csv_filename}")
    print(f"   總筆數：{len(df_all)}")

    return csv_filename


# ============================================================
# 價格驗證（TWSE / TPEX 官方 API）
# ============================================================

def _query_twse(ticker: str, date_ymd: str) -> float | None:
    """從證交所 API 取得上市股收盤價"""
    import requests, urllib3
    urllib3.disable_warnings()
    try:
        url = f"https://www.twse.com.tw/exchangeReport/STOCK_DAY?response=json&date={date_ymd}&stockNo={ticker}"
        r = requests.get(url, timeout=10, verify=False,
                         headers={"User-Agent": "Mozilla/5.0"})
        if r.ok:
            data = r.json()
            if data.get("stat") == "OK" and data.get("data"):
                last_row = data["data"][-1]
                return float(last_row[6].replace(",", ""))
    except Exception:
        pass
    return None


def _query_tpex(ticker: str, date_slash: str) -> float | None:
    """從櫃買中心 API 取得上櫃股收盤價"""
    import requests, urllib3
    urllib3.disable_warnings()
    try:
        url = (f"https://www.tpex.org.tw/www/zh-tw/afterTrading/dailyQuotes"
               f"?date={date_slash}&code={ticker}&response=json")
        r = requests.get(url, timeout=10, verify=False,
                         headers={"User-Agent": "Mozilla/5.0"})
        if r.ok:
            data = r.json()
            tables = data.get("tables", [])
            if tables and tables[0].get("data"):
                # 必須確認回傳的代號與查詢的一致
                for row in tables[0]["data"]:
                    if str(row[0]).strip() == ticker:
                        return float(str(row[2]).replace(",", ""))
    except Exception:
        pass
    return None


def _query_official_price(ticker: str, date_ymd: str, date_slash: str) -> float | None:
    """依市場別查詢官方收盤價（自動偵測上市/上櫃）"""
    cached = _suffix_cache.get(ticker, "")
    is_tpex = cached.endswith(".TWO")

    if is_tpex:
        price = _query_tpex(ticker, date_slash)
        if price is not None:
            return price
        return _query_twse(ticker, date_ymd)
    else:
        price = _query_twse(ticker, date_ymd)
        if price is not None:
            return price
        return _query_tpex(ticker, date_slash)


def _verify_and_fix_prices(csv_path: str) -> int:
    """驗證 >15% 價格變動，用 TWSE/TPEX 官方 API 修正"""
    import glob as _glob

    base_dir = os.path.dirname(os.path.abspath(csv_path))
    all_csvs = sorted(_glob.glob(os.path.join(base_dir, "stock_data_*.csv")))

    abs_path = os.path.abspath(csv_path)
    try:
        idx = all_csvs.index(abs_path)
    except ValueError:
        return 0
    if idx == 0:
        print("   ℹ️ 無前日 CSV，跳過價格驗證")
        return 0

    prev_csv = all_csvs[idx - 1]

    df_today = pd.read_csv(abs_path, encoding="utf-8-sig")
    df_prev = pd.read_csv(prev_csv, encoding="utf-8-sig")

    merged = pd.merge(
        df_today[["股票代號", "收盤價"]],
        df_prev[["股票代號", "收盤價"]],
        on="股票代號", suffixes=("_today", "_prev")
    )
    merged["chg_pct"] = ((merged["收盤價_today"] - merged["收盤價_prev"]).abs()
                         / merged["收盤價_prev"])
    flagged = merged[merged["chg_pct"] > 0.10]

    if flagged.empty:
        print("   ✅ 價格驗證通過（無 >15% 異常變動）")
        return 0

    print(f"\n   🔍 發現 {len(flagged)} 檔 >15% 價格變動，向官方驗證中...")

    # 從 CSV 取交易日以確保日期正確
    trading_date_raw = str(df_today["交易日"].iloc[0])[:10]  # "2026-03-18"
    date_ymd = trading_date_raw.replace("-", "")              # "20260318"
    date_slash = trading_date_raw.replace("-", "/")           # "2026/03/18"

    fixes = []
    for _, row in flagged.iterrows():
        ticker = str(int(row["股票代號"]))
        csv_price = row["收盤價_today"]

        official_price = _query_official_price(ticker, date_ymd, date_slash)

        if official_price is None:
            print(f"      ⚠️ {ticker} 無法取得官方價格，跳過")
            continue

        if abs(official_price - csv_price) / official_price > 0.01:
            fixes.append((ticker, csv_price, official_price))
            print(f"      🔧 {ticker}: {csv_price} → {official_price}")
        else:
            print(f"      ✅ {ticker}: 價格正確 ({csv_price})")

        time.sleep(0.5)

    if fixes:
        df = pd.read_csv(abs_path, encoding="utf-8-sig")
        for ticker, old_price, new_price in fixes:
            df.loc[df["股票代號"] == int(ticker), "收盤價"] = new_price
        df.to_csv(abs_path, index=False, encoding="utf-8-sig")
        print(f"\n   📝 已修正 {len(fixes)} 筆價格（來源：TWSE/TPEX 官方）")
    else:
        print(f"\n   ✅ 驗證完成，價格皆正確")

    return len(fixes)


# ============================================================
# 主掃描流程
# ============================================================

def _backfill_watchlist(csv_path: str) -> int:
    """檢查 watchlist 中的股票是否都在 CSV 裡，漏掉的立刻補抓"""
    watchlist_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "watchlist.json")
    if not os.path.exists(watchlist_path):
        return 0

    try:
        with open(watchlist_path, "r", encoding="utf-8") as f:
            watchlist = json.load(f)
    except Exception:
        return 0

    if not watchlist:
        return 0

    watch_codes = {str(item["code"]) for item in watchlist}

    df_csv = pd.read_csv(csv_path, encoding="utf-8-sig")
    existing_codes = set(df_csv["股票代號"].astype(str).tolist())

    missing = watch_codes - existing_codes
    if not missing:
        return 0

    print(f"\n   🔄 Watchlist 補漏：發現 {len(missing)} 支追蹤股票不在 CSV 中")

    end_date = (datetime.now() + timedelta(days=1)).strftime("%Y-%m-%d")
    start_date = (datetime.now() - timedelta(days=90)).strftime("%Y-%m-%d")

    added = 0
    for code in sorted(missing):
        name_display = get_stock_name(code)
        print(f"      補抓 {code} {name_display}...", end=" ", flush=True)

        try:
            df = download_stock(code, start_date, end_date)
            if df is None or len(df) < 5:
                print("資料不足，跳過")
                continue

            info = _fetch_ticker_info(code)
            csv_row = _build_csv_row(code, name_display, df, info)
            if csv_row:
                df_csv = pd.concat([df_csv, pd.DataFrame([csv_row])], ignore_index=True)
                added += 1
                print(f"OK (收盤 {csv_row['收盤價']})")
            else:
                print("建立資料失敗")
        except Exception as e:
            print(f"失敗: {e}")

        time.sleep(0.3)

    if added > 0:
        df_csv.to_csv(csv_path, index=False, encoding="utf-8-sig")
        print(f"   ✅ 已補齊 {added} 支 watchlist 股票")

    return added


def run_screener(cfg: ScreenerConfig | None = None, deep: bool = False, save_csv: bool = True) -> list[StockResult]:
    """
    main entry — 可傳入 config，不傳則用預設值。
    """
    if cfg is None:
        cfg = ScreenerConfig()

    # 至少要開啟一種篩選
    if not any([cfg.filter_limit_up, cfg.filter_dividend, cfg.filter_rsi_ma, cfg.filter_bbands, 
                cfg.filter_resonance, cfg.filter_kd_golden, cfg.filter_bb_upper]):
        print("⚠️  沒有開啟任何篩選模式，請至少開啟一種")
        print("    limit_up / dividend / rsi_ma / bbands / resonance / kd_golden / bb_upper")
        return []

    print("=" * 55)
    print(" 台股看盤系統 啟動")
    print("=" * 55)

    # 顯示當前設定
    start_date, end_date = resolve_dates(cfg)
    print(f"📅 篩選日期：{start_date} ~ {end_date}")
    modes = []
    if cfg.filter_limit_up:  modes.append(f"漲停(≥{cfg.limit_up_pct}%)")
    if cfg.filter_dividend:  modes.append(f"殖利率(≥{cfg.min_dividend_yield}%)")
    if cfg.filter_rsi_ma:    modes.append(f"RSI+MA策略")
    if cfg.filter_bbands:    modes.append(f"布林通道開口")
    if cfg.filter_resonance: modes.append(f"技術指標共振(KD+RSI+BB+量+MACD)")
    if cfg.filter_kd_golden: modes.append(f"KD黃金交叉(短線)")
    if cfg.filter_bb_upper:  modes.append(f"布林上軌突破(短線)")
    print(f"🎯 篩選模式：{' | '.join(modes)}")
    print("-" * 55)

    # 讀取股票清單
    stocks = get_stock_list(cfg, deep=deep)
    if not stocks:
        print("❌ 無股票清單可掃描")
        return []
    print(f"📦 掃描股票數：{len(stocks)}")
    print("-" * 55)

    results: list[StockResult] = []

    # 計算統一下載的起始日期（90天回溯，足夠所有指標計算）
    unified_start = (datetime.strptime(start_date, "%Y-%m-%d") - timedelta(days=90)).strftime("%Y-%m-%d")

    # CSV 收集
    csv_data = [] if save_csv else None
    csv_success = 0

    for i, ticker in enumerate(stocks, 1):
        print(f"⏳ ({i:>3}/{len(stocks)}) 掃描 {ticker} ...", end=" ", flush=True)

        result = StockResult(ticker=ticker)
        hit = False

        # === 統一下載一次（90天歷史，足夠所有指標計算）===
        df = download_stock(ticker, unified_start, end_date)
        if df is None or len(df) < 2:
            print("❌ 數據不足")
            continue

        # 取得股票名稱（dict lookup，無 API 呼叫）
        name = get_stock_name(ticker)

        # 如果需要 info 資料（殖利率篩選/市場篩選/CSV），統一取得一次
        info = None
        need_info = save_csv or cfg.filter_dividend or cfg.filter_volume or cfg.filter_market_cap or cfg.filter_pe_ratio
        if need_info:
            info = _fetch_ticker_info(ticker)

        # --- 漲停篩選 ---
        if cfg.filter_limit_up:
            lu = check_limit_up(df, cfg)
            if lu:
                result.limit_up = lu
                hit = True

        # --- 殖利率篩選 ---
        if cfg.filter_dividend:
            div = check_dividend(cfg, ticker, info=info)
            if div:
                result.dividend = div
                hit = True

        # --- RSI+MA 策略篩選 ---
        if cfg.filter_rsi_ma:
            rm = check_rsi_ma(df, cfg)
            if rm:
                result.rsi_ma = rm
                hit = True

        # --- 布林通道開口向上篩選 ---
        if cfg.filter_bbands:
            bb = check_bbands(df, cfg)
            if bb:
                result.bbands = bb
                hit = True

        # --- 技術指標共振策略（KD + RSI + BB + 量）---
        if cfg.filter_resonance:
            res = check_resonance(df, cfg)
            if res:
                result.resonance = res
                hit = True

        # --- KD 黃金交叉篩選（短線做多訊號）---
        if cfg.filter_kd_golden:
            kdg = check_kd_golden(df, cfg)
            if kdg:
                result.kd_golden = kdg
                hit = True

        # --- 布林通道接近上軌篩選（短線突破訊號）---
        if cfg.filter_bb_upper:
            bbu = check_bb_upper(df, cfg)
            if bbu:
                result.bb_upper = bbu
                hit = True

        # --- 市場指標篩選（成交量、市值、本益比）---
        if cfg.filter_volume or cfg.filter_market_cap or cfg.filter_pe_ratio or hit:
            if info is None:
                info = _fetch_ticker_info(ticker)
            market = check_market_metrics(cfg, ticker, df, info=info)
            if market:
                result.market = market
                if cfg.filter_volume or cfg.filter_market_cap or cfg.filter_pe_ratio:
                    hit = True
            elif cfg.filter_volume or cfg.filter_market_cap or cfg.filter_pe_ratio:
                hit = False

        # === CSV 資料收集 ===
        if save_csv and len(df) >= 30:
            csv_row = _build_csv_row(ticker, name, df, info)
            if csv_row:
                csv_data.append(csv_row)
                csv_success += 1

        if hit:
            result.name = name
            results.append(result)
            print("✅ 命中")
        else:
            print("—")

        time.sleep(0.3)   # 避免打太快被 yfinance 限制

    # ============================================================
    # 排序結果
    # ============================================================
    if results and cfg.sort_by:
        def get_sort_key(r: StockResult):
            """取得排序鍵值"""
            if cfg.sort_by == "dividend_yield" and r.dividend:
                return r.dividend.get("dividend_yield", 0)
            elif cfg.sort_by == "market_cap" and r.market:
                return r.market.get("market_cap") or 0
            elif cfg.sort_by == "pe_ratio" and r.market:
                return r.market.get("pe_ratio") or 0
            elif cfg.sort_by == "volume" and r.market:
                return r.market.get("avg_volume") or 0
            elif cfg.sort_by == "chg_pct" and r.limit_up:
                return r.limit_up.get("chg_pct", 0)
            return 0
        
        results.sort(key=get_sort_key, reverse=cfg.sort_desc)
        print(f"\n📊 已按 {cfg.sort_by} {'由大到小' if cfg.sort_desc else '由小到大'} 排序")

    # ============================================================
    # 輸出結果
    # ============================================================
    print("\n" + "=" * 55)
    print(f" 篩選結果：{len(results)} 筆")
    print("=" * 55)

    if not results:
        print("  沒有股票符合篩選條件")
    else:
        for r in results:
            name_display = f" {r.name}" if r.name else ""
            print(f"\n  📌 {r.ticker}{name_display}")
            if r.limit_up:
                print(f"     🔴 漲停  日期：{r.limit_up['date']}  "
                      f"收盤：{r.limit_up['close']}  "
                      f"漲幅：{r.limit_up['chg_pct']}%  "
                      f"(期間漲停次數：{r.limit_up['count']})")
            if r.dividend:
                print(f"     💰 殖利率：{r.dividend['dividend_yield']}%  "
                      f"現價：{r.dividend['price']}  "
                      f"每股配息：{r.dividend['dividend_rate']}元")
            if r.market:
                metrics = []
                if r.market.get("market_cap"):
                    metrics.append(f"市值：{r.market['market_cap']}億")
                if r.market.get("pe_ratio"):
                    metrics.append(f"本益比：{r.market['pe_ratio']}")
                if r.market.get("avg_volume"):
                    metrics.append(f"均量：{r.market['avg_volume']}張")
                if metrics:
                    print(f"     📊 {' / '.join(metrics)}")
            if r.rsi_ma:
                print(f"     📈 RSI+MA  收盤：{r.rsi_ma['close']}  "
                      f"MA{cfg.ma_period}：{r.rsi_ma['ma']}  "
                      f"RSI：{r.rsi_ma['rsi']}")
            if r.bbands:
                print(f"     📉 布林通道  收盤：{r.bbands['close']}  "
                      f"中軌：{r.bbands['bb_middle']}  "
                      f"上軌：{r.bbands['bb_upper']}  "
                      f"通道擴張：{r.bbands['width_change_pct']}%")
            if r.resonance:
                print(f"     🎯 【共振訊號】五大指標")
                print(f"        ├ KD: K={r.resonance['k']:.1f} / D={r.resonance['d']:.1f}")
                print(f"        ├ RSI: {r.resonance['rsi']:.1f}")
                print(f"        ├ BB: 價格={r.resonance['close']} / 下軌={r.resonance['bb_lower']}")
                print(f"        ├ MACD: DIF={r.resonance['dif']:.2f} / MACD={r.resonance['macd']:.2f} / 柱={r.resonance['histogram']:.2f}")
                print(f"        └ 量: {r.resonance['volume']//1000}張 (5日均{r.resonance['vol_ma5']//1000}張, 放大{r.resonance['volume_ratio']}倍)")
            if r.kd_golden:
                print(f"     📈 【KD黃金交叉】短線做多")
                print(f"        收盤：{r.kd_golden['close']}  K={r.kd_golden['k']:.1f} / D={r.kd_golden['d']:.1f}")
                print(f"        狀態：{r.kd_golden['k_trend']}")
            if r.bb_upper:
                print(f"     🚀 【布林上軌突破】短線強勢")
                print(f"        收盤：{r.bb_upper['close']}  上軌：{r.bb_upper['bb_upper']}")
                print(f"        狀態：{r.bb_upper['status']}")

    print("\n" + "=" * 55)

    # ============================================================
    # 存 CSV（取代 data_downloader.py 的功能）
    # ============================================================
    if save_csv and csv_data:
        csv_path = _save_stock_csv(csv_data)
        if csv_path:
            print(f"   成功：{csv_success} 筆")
            _verify_and_fix_prices(csv_path)
            _backfill_watchlist(csv_path)
            print(f"   💡 可使用 quick_filter.py 快速篩選")

    return results


# ============================================================
# Excel 報表輸出
# ============================================================

def export_to_excel(results: list[StockResult], cfg: ScreenerConfig, filename: str = "stock_screening_report.xlsx"):
    """
    將篩選結果輸出成 Excel 報表。
    每種篩選模式一個 sheet：漲停、殖利率、RSI+MA。
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from datetime import datetime

    wb = Workbook()
    wb.remove(wb.active)  # 移除預設空白 sheet

    # --- 標題樣式 ---
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # --- Sheet 1：漲停篩選 ---
    if cfg.filter_limit_up:
        ws1 = wb.create_sheet("漲停篩選")
        ws1.append(["股票代號", "股票名稱", "漲停日期", "收盤價", "漲幅 (%)", "期間漲停次數", "市值 (億)", "本益比", "均量 (張)"])
        
        for cell in ws1[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for r in results:
            if r.limit_up:
                market_cap = r.market.get("market_cap") if r.market else None
                pe_ratio = r.market.get("pe_ratio") if r.market else None
                avg_volume = r.market.get("avg_volume") if r.market else None
                
                ws1.append([
                    r.ticker,
                    r.name,
                    r.limit_up["date"],
                    r.limit_up["close"],
                    r.limit_up["chg_pct"],
                    r.limit_up["count"],
                    market_cap,
                    pe_ratio,
                    avg_volume
                ])

        # 設定欄寬
        ws1.column_dimensions['A'].width = 12
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 12
        ws1.column_dimensions['D'].width = 10
        ws1.column_dimensions['E'].width = 12
        ws1.column_dimensions['F'].width = 16
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['H'].width = 10
        ws1.column_dimensions['I'].width = 12

    # --- Sheet 2：殖利率篩選 ---
    if cfg.filter_dividend:
        ws2 = wb.create_sheet("殖利率篩選")
        ws2.append(["股票代號", "股票名稱", "現價", "每股配息", "殖利率 (%)", "市值 (億)", "本益比", "均量 (張)"])
        
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for r in results:
            if r.dividend:
                market_cap = r.market.get("market_cap") if r.market else None
                pe_ratio = r.market.get("pe_ratio") if r.market else None
                avg_volume = r.market.get("avg_volume") if r.market else None
                
                ws2.append([
                    r.ticker,
                    r.name,
                    r.dividend["price"],
                    r.dividend["dividend_rate"],
                    r.dividend["dividend_yield"],
                    market_cap,
                    pe_ratio,
                    avg_volume
                ])

        ws2.column_dimensions['A'].width = 12
        ws2.column_dimensions['B'].width = 20
        ws2.column_dimensions['C'].width = 10
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 14
        ws2.column_dimensions['F'].width = 12
        ws2.column_dimensions['G'].width = 10
        ws2.column_dimensions['H'].width = 12

    # --- Sheet 3：RSI+MA 策略篩選 ---
    if cfg.filter_rsi_ma:
        ws3 = wb.create_sheet("RSI_MA策略")
        ws3.append(["股票代號", "股票名稱", "收盤價", f"MA{cfg.ma_period}", "RSI", "市值 (億)", "本益比", "均量 (張)"])
        
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for r in results:
            if r.rsi_ma:
                market_cap = r.market.get("market_cap") if r.market else None
                pe_ratio = r.market.get("pe_ratio") if r.market else None
                avg_volume = r.market.get("avg_volume") if r.market else None
                
                ws3.append([
                    r.ticker,
                    r.name,
                    r.rsi_ma["close"],
                    r.rsi_ma["ma"],
                    r.rsi_ma["rsi"],
                    market_cap,
                    pe_ratio,
                    avg_volume
                ])

        ws3.column_dimensions['A'].width = 12
        ws3.column_dimensions['B'].width = 20
        ws3.column_dimensions['C'].width = 10
        ws3.column_dimensions['D'].width = 12
        ws3.column_dimensions['E'].width = 10
        ws3.column_dimensions['F'].width = 12
        ws3.column_dimensions['G'].width = 10
        ws3.column_dimensions['H'].width = 12

    # --- Sheet 4：布林通道開口向上篩選 ---
    if cfg.filter_bbands:
        ws4 = wb.create_sheet("布林通道")
        ws4.append(["股票代號", "股票名稱", "收盤價", "上軌", "中軌", "下軌", "通道擴張(%)", "市值 (億)", "本益比"])
        
        for cell in ws4[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for r in results:
            if r.bbands:
                market_cap = r.market.get("market_cap") if r.market else None
                pe_ratio = r.market.get("pe_ratio") if r.market else None
                
                ws4.append([
                    r.ticker,
                    r.name,
                    r.bbands["close"],
                    r.bbands["bb_upper"],
                    r.bbands["bb_middle"],
                    r.bbands["bb_lower"],
                    r.bbands["width_change_pct"],
                    market_cap,
                    pe_ratio
                ])

        ws4.column_dimensions['A'].width = 12
        ws4.column_dimensions['B'].width = 20
        ws4.column_dimensions['C'].width = 10
        ws4.column_dimensions['D'].width = 10
        ws4.column_dimensions['E'].width = 10
        ws4.column_dimensions['F'].width = 10
        ws4.column_dimensions['G'].width = 14
        ws4.column_dimensions['H'].width = 12
        ws4.column_dimensions['I'].width = 10

    # --- Sheet 5：技術指標共振策略 ---
    if cfg.filter_resonance:
        ws5 = wb.create_sheet("技術指標共振")
        ws5.append(["股票代號", "股票名稱", "收盤價", "K值", "D值", "RSI", "DIF", "MACD", "柱狀圖", "下軌", "量(張)", "量放大", "市值(億)", "本益比"])
        
        for cell in ws5[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for r in results:
            if r.resonance:
                market_cap = r.market.get("market_cap") if r.market else None
                pe_ratio = r.market.get("pe_ratio") if r.market else None
                
                ws5.append([
                    r.ticker,
                    r.name,
                    r.resonance["close"],
                    r.resonance["k"],
                    r.resonance["d"],
                    r.resonance["rsi"],
                    r.resonance["dif"],
                    r.resonance["macd"],
                    r.resonance["histogram"],
                    r.resonance["bb_lower"],
                    r.resonance["volume"] // 1000,  # 轉成張
                    r.resonance["volume_ratio"],
                    market_cap,
                    pe_ratio
                ])

        ws5.column_dimensions['A'].width = 12
        ws5.column_dimensions['B'].width = 20
        ws5.column_dimensions['C'].width = 10
        ws5.column_dimensions['D'].width = 8
        ws5.column_dimensions['E'].width = 8
        ws5.column_dimensions['F'].width = 8
        ws5.column_dimensions['G'].width = 10
        ws5.column_dimensions['H'].width = 10
        ws5.column_dimensions['I'].width = 10
        ws5.column_dimensions['J'].width = 10
        ws5.column_dimensions['K'].width = 10
        ws5.column_dimensions['L'].width = 10
        ws5.column_dimensions['M'].width = 12
        ws5.column_dimensions['N'].width = 10
        ws5.column_dimensions['F'].width = 8
        ws5.column_dimensions['G'].width = 10
        ws5.column_dimensions['H'].width = 12
        ws5.column_dimensions['I'].width = 12
        ws5.column_dimensions['J'].width = 12
        ws5.column_dimensions['K'].width = 10

    # --- Sheet 6：KD 黃金交叉（短線做多）---
    if cfg.filter_kd_golden:
        ws6 = wb.create_sheet("KD黃金交叉")
        ws6.append(["股票代號", "股票名稱", "收盤價", "K值", "D值", "狀態", "市值 (億)", "本益比", "殖利率(%)"])
        
        for cell in ws6[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for r in results:
            if r.kd_golden:
                market_cap = r.market.get("market_cap") if r.market else None
                pe_ratio = r.market.get("pe_ratio") if r.market else None
                div_yield = r.dividend.get("dividend_yield") if r.dividend else None
                
                ws6.append([
                    r.ticker,
                    r.name,
                    r.kd_golden["close"],
                    r.kd_golden["k"],
                    r.kd_golden["d"],
                    r.kd_golden["k_trend"],
                    market_cap,
                    pe_ratio,
                    div_yield
                ])

        ws6.column_dimensions['A'].width = 12
        ws6.column_dimensions['B'].width = 20
        ws6.column_dimensions['C'].width = 10
        ws6.column_dimensions['D'].width = 8
        ws6.column_dimensions['E'].width = 8
        ws6.column_dimensions['F'].width = 12
        ws6.column_dimensions['G'].width = 12
        ws6.column_dimensions['H'].width = 10
        ws6.column_dimensions['I'].width = 12

    # --- Sheet 7：布林通道上軌突破（短線強勢）---
    if cfg.filter_bb_upper:
        ws7 = wb.create_sheet("布林上軌突破")
        ws7.append(["股票代號", "股票名稱", "收盤價", "上軌", "中軌", "下軌", "狀態", "市值 (億)", "本益比"])
        
        for cell in ws7[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = border

        for r in results:
            if r.bb_upper:
                market_cap = r.market.get("market_cap") if r.market else None
                pe_ratio = r.market.get("pe_ratio") if r.market else None
                
                ws7.append([
                    r.ticker,
                    r.name,
                    r.bb_upper["close"],
                    r.bb_upper["bb_upper"],
                    r.bb_upper["bb_middle"],
                    r.bb_upper["bb_lower"],
                    r.bb_upper["status"],
                    market_cap,
                    pe_ratio
                ])

        ws7.column_dimensions['A'].width = 12
        ws7.column_dimensions['B'].width = 20
        ws7.column_dimensions['C'].width = 10
        ws7.column_dimensions['D'].width = 10
        ws7.column_dimensions['E'].width = 10
        ws7.column_dimensions['F'].width = 10
        ws7.column_dimensions['G'].width = 12
        ws7.column_dimensions['H'].width = 12
        ws7.column_dimensions['I'].width = 10

    # --- Sheet 8：摘要 ---
    ws_summary = wb.create_sheet("摘要", 0)  # 插入到最前面
    ws_summary.append(["台股篩選報表"])
    ws_summary.append([])
    ws_summary.append(["生成時間", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    ws_summary.append(["篩選條件", ""])
    
    if cfg.filter_limit_up:
        ws_summary.append(["  漲停門檻", f"{cfg.limit_up_pct}%"])
    if cfg.filter_dividend:
        ws_summary.append(["  最低殖利率", f"{cfg.min_dividend_yield}%"])
    if cfg.filter_rsi_ma:
        ws_summary.append(["  RSI 超賣閾值", cfg.rsi_oversold])
        ws_summary.append(["  MA 週期", cfg.ma_period])
    if cfg.filter_bbands:
        ws_summary.append(["  布林通道週期", cfg.bbands_period])
        ws_summary.append(["  通道擴張門檻", f"{cfg.bbands_expand_min*100}%"])
    if cfg.filter_resonance:
        ws_summary.append(["  【共振策略】", "KD + RSI + BB + 量能"])
        ws_summary.append(["    KD 超賣閾值", f"K<{cfg.kd_oversold_k}, D<{cfg.kd_oversold_d}"])
        ws_summary.append(["    RSI 超賣閾值", f"<{cfg.rsi_oversold}"])
        ws_summary.append(["    量能放大倍數", f"≥{cfg.volume_surge_ratio}x"])
    
    ws_summary.append([])
    ws_summary.append(["篩選結果統計", ""])
    ws_summary.append(["  總命中數", len(results)])
    
    limit_up_count = sum(1 for r in results if r.limit_up)
    dividend_count = sum(1 for r in results if r.dividend)
    rsi_ma_count   = sum(1 for r in results if r.rsi_ma)
    bbands_count   = sum(1 for r in results if r.bbands)
    resonance_count = sum(1 for r in results if r.resonance)
    kd_golden_count = sum(1 for r in results if r.kd_golden)
    bb_upper_count = sum(1 for r in results if r.bb_upper)
    
    if cfg.filter_limit_up:
        ws_summary.append(["  漲停股票數", limit_up_count])
    if cfg.filter_dividend:
        ws_summary.append(["  高殖利率股票數", dividend_count])
    if cfg.filter_rsi_ma:
        ws_summary.append(["  RSI+MA 訊號數", rsi_ma_count])
    if cfg.filter_bbands:
        ws_summary.append(["  布林通道訊號數", bbands_count])
    if cfg.filter_resonance:
        ws_summary.append(["  技術指標共振數", resonance_count])
    if cfg.filter_kd_golden:
        ws_summary.append(["  KD 黃金交叉數", kd_golden_count])
    if cfg.filter_bb_upper:
        ws_summary.append(["  布林上軌突破數", bb_upper_count])

    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.column_dimensions['A'].width = 18
    ws_summary.column_dimensions['B'].width = 25

    # 儲存到當前目錄
    wb.save(filename)
    print(f"\n📊 Excel 報表已生成：{filename}")
    return filename


# ============================================================
# 進入點 — 命令行模式選擇
# ============================================================
# 使用方法：
#   python stock_system.py              → 組合篩選（快速模式）
#   python stock_system.py limit        → 只篩漲停
#   python stock_system.py div          → 只篩殖利率
#   python stock_system.py rsi          → 只篩 RSI+MA
#   python stock_system.py div deep     → 殖利率篩選 + 深度掃描全範圍股票
#   python stock_system.py deep         → 組合篩選 + 深度掃描
#
# deep 旗標說明：
#   不加 deep  → 快速模式，掃已知範圍（約 1700 筆有效股票）
#   加 deep    → 深度掃描，掃 1100-9999 全範圍（確保上市＋上櫃都不漏）
#               第一次跑 deep 會花約 30 分鐘驗證，之後緩存 7 天
# ============================================================

import sys

def parse_args() -> tuple[str, bool, bool]:
    """解析命令行參數，回傳 (mode, deep, save_csv)"""
    args = [a.strip().lower().lstrip("-") for a in sys.argv[1:]]
    mode = "all"
    deep = False
    save_csv = True
    for a in args:
        if a == "deep":
            deep = True
        elif a in ("no-csv", "nocsv", "no_csv"):
            save_csv = False
        elif a in ("limit", "div", "rsi", "all"):
            mode = a
    return mode, deep, save_csv

if __name__ == "__main__":

    # 自動更新股票名稱對照表
    try:
        n = refresh_stock_names()
        if n:
            print(f"  📋 股票名稱對照表已更新：{n} 筆")
    except Exception:
        pass

    mode, deep, save_csv = parse_args()

    if mode == "limit":
        print("\n【模式】近10天內漲停篩選\n")
        cfg = ScreenerConfig(
            lookback_days=10,
            filter_limit_up=True,
            limit_up_pct=8.0,
            max_stocks=9999
        )

    elif mode == "div":
        print("\n【模式】殖利率 >= 5% 篩選\n")
        cfg = ScreenerConfig(
            filter_dividend=True,
            min_dividend_yield=5.0,
            max_stocks=9999,
            sort_by="dividend_yield",  # 按殖利率排序
            sort_desc=True              # 由高到低
        )

    elif mode == "rsi":
        print("\n【模式】RSI+MA 策略篩選\n")
        cfg = ScreenerConfig(
            lookback_days=60,
            filter_rsi_ma=True,
            rsi_period=14,
            ma_period=20,
            rsi_oversold=30.0,
            max_stocks=9999
        )

    else:
        print("\n【模式】組合篩選（漲停 + 殖利率 + RSI+MA）\n")
        cfg = ScreenerConfig(
            lookback_days=10,
            filter_limit_up=True,
            filter_dividend=True,
            filter_rsi_ma=True,
            limit_up_pct=8.0,
            min_dividend_yield=5.0,
            rsi_period=14,
            ma_period=20,
            rsi_oversold=30.0,
            max_stocks=9999
        )

    results = run_screener(cfg, deep=deep, save_csv=save_csv)
    
    # 自動輸出 Excel 報表
    if results:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"stock_report_{timestamp}.xlsx"
        excel_path = export_to_excel(results, cfg, filename)
        print(f"✅ 報表已儲存：{excel_path}")
