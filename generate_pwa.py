"""
generate_pwa.py
讀取最新 stock_data CSV，產出完整 PWA 到 pwa/ 目錄
"""
import pandas as pd
import json
import os
import glob
import struct
import zlib
import shutil
from datetime import datetime


def get_latest_csv():
    """找最新的 stock_data_*.csv"""
    files = glob.glob("stock_data_*.csv")
    return max(files) if files else None


def get_sorted_csvs(n=10):
    """取得最近 n 個 stock_data_*.csv（依檔名排序，新→舊）"""
    files = sorted(glob.glob("stock_data_*.csv"), reverse=True)
    return files[:n]


def compute_trend(df_today, csv_files):
    """趨勢比對：今天 vs 昨天的指標變化"""
    # 找前一個「完整」的 CSV（至少 100 筆）
    df_yesterday = None
    for f in csv_files[1:]:
        try:
            tmp = pd.read_csv(f, encoding="utf-8-sig")
            if len(tmp) >= 100:
                df_yesterday = tmp
                break
        except Exception:
            continue

    if df_yesterday is None:
        # 沒有可比對的完整 CSV
        for col in ['漲跌幅', '價格趨勢', 'KD趨勢', 'RSI趨勢', '量能趨勢', 'K值變化', 'RSI變化']:
            df_today[col] = None
        return df_today

    # 合併今昨資料
    merged = pd.merge(
        df_today[['股票代號', '收盤價', 'K值', 'D值', 'RSI', '量比', 'BB上軌', 'BB下軌']],
        df_yesterday[['股票代號', '收盤價', 'K值', 'D值', 'RSI', '量比', 'BB上軌', 'BB下軌']],
        on='股票代號', suffixes=('', '_昨')
    )

    # 計算變化
    merged['漲跌幅'] = ((merged['收盤價'] - merged['收盤價_昨']) / merged['收盤價_昨'] * 100).round(2)
    merged['K值變化'] = (merged['K值'] - merged['K值_昨']).round(2)
    merged['D值變化'] = (merged['D值'] - merged['D值_昨']).round(2)
    merged['RSI變化'] = (merged['RSI'] - merged['RSI_昨']).round(2)
    merged['量比變化'] = (merged['量比'] - merged['量比_昨']).round(2)

    # BB 位置
    def bb_pos(close, upper, lower):
        denom = upper - lower
        return ((close - lower) / denom * 100).where(denom > 0)

    merged['BB位置_今'] = bb_pos(merged['收盤價'], merged['BB上軌'], merged['BB下軌']).round(2)
    merged['BB位置_昨'] = bb_pos(merged['收盤價_昨'], merged['BB上軌_昨'], merged['BB下軌_昨']).round(2)
    merged['BB位置變化'] = (merged['BB位置_今'] - merged['BB位置_昨']).round(2)

    # 趨勢分類
    def price_trend(x):
        if pd.isna(x): return '平穩'
        if x >= 3: return '大漲'
        if x > 0: return '上漲'
        if x <= -3: return '下跌'
        return '平穩'

    def kd_trend(k_chg, d_chg):
        if pd.isna(k_chg) or pd.isna(d_chg): return 'KD平穩'
        if k_chg > 3 and d_chg > 3: return 'KD轉強'
        if k_chg < -3 and d_chg < -3: return 'KD轉弱'
        return 'KD平穩'

    def rsi_trend(x):
        if pd.isna(x): return 'RSI平穩'
        if x > 5: return 'RSI轉強'
        if x < -5: return 'RSI轉弱'
        return 'RSI平穩'

    def vol_trend(x):
        if pd.isna(x): return '量能平穩'
        if x > 0.5: return '量能放大'
        if x < -0.5: return '量能縮小'
        return '量能平穩'

    merged['價格趨勢'] = merged['漲跌幅'].apply(price_trend)
    merged['KD趨勢'] = merged.apply(lambda r: kd_trend(r['K值變化'], r['D值變化']), axis=1)
    merged['RSI趨勢'] = merged['RSI變化'].apply(rsi_trend)
    merged['量能趨勢'] = merged['量比變化'].apply(vol_trend)

    # 只取需要的欄位 merge 回主 df
    trend_cols = ['股票代號', '漲跌幅', '價格趨勢', 'KD趨勢', 'RSI趨勢', '量能趨勢', 'K值變化', 'RSI變化']
    df_today = df_today.merge(merged[trend_cols], on='股票代號', how='left')
    return df_today


def compute_breakout(df_today, csv_files):
    """突破訊號 + 強度評分（跨天統計）"""
    # 讀取所有 CSV 並堆疊（跳過不完整的檔案）
    all_dfs = []
    for f in csv_files:
        try:
            tmp = pd.read_csv(f, encoding="utf-8-sig")
            if len(tmp) >= 100:
                all_dfs.append(tmp)
        except Exception:
            continue

    if not all_dfs:
        for col in ['強度評分', '訊號數量', '訊號']:
            df_today[col] = None
        return df_today

    combined = pd.concat(all_dfs, ignore_index=True)

    # 計算每支股票的跨天統計
    num_cols = ['收盤價', 'RSI', 'D值', '量比']
    existing_cols = [c for c in num_cols if c in combined.columns]

    stats = combined.groupby('股票代號').agg(
        平均價格=('收盤價', 'mean'),
        最低價=('收盤價', 'min'),
        最高價=('收盤價', 'max'),
        平均RSI=('RSI', 'mean'),
        平均D=('D值', 'mean'),
        平均量比=('量比', 'mean'),
    ).round(2)
    stats.reset_index(inplace=True)

    # 合併最新一天資料
    latest_cols = ['股票代號', '收盤價', 'K值', 'D值', 'RSI', '量比']
    latest_cols = [c for c in latest_cols if c in df_today.columns]
    merged = stats.merge(df_today[latest_cols], on='股票代號', suffixes=('_stat', ''))

    # 計算衍生指標
    merged['漲跌幅%'] = ((merged['收盤價'] - merged['平均價格']) / merged['平均價格'] * 100).round(2)
    merged['振幅%'] = (((merged['最高價'] - merged['最低價']) / merged['平均價格']) * 100).round(2)
    merged['反彈幅度%'] = (((merged['收盤價'] - merged['最低價']) / merged['最低價']) * 100).round(2)

    # 偵測 14 種訊號
    results = []
    for _, r in merged.iterrows():
        signals = []
        score = 0

        # 1. 量能突破
        if pd.notna(r.get('量比')) and pd.notna(r.get('平均量比')) and r['平均量比'] > 0:
            if r['量比'] > 1.5 * r['平均量比']:
                signals.append('量能突破')
                score += 2

        # 2. KD黃金交叉
        if pd.notna(r.get('K值')) and pd.notna(r.get('D值')):
            if r['K值'] > r['D值'] and r['K值'] < 50:
                signals.append('KD黃金交叉')
                score += 3

        # 3. RSI轉強
        if pd.notna(r.get('RSI')) and pd.notna(r.get('平均RSI')):
            if r['RSI'] > 50 and r['平均RSI'] < 50:
                signals.append('RSI轉強')
                score += 2

        # 4. 價格突破
        if pd.notna(r.get('收盤價')) and pd.notna(r.get('最高價')) and r['最高價'] > 0:
            if r['收盤價'] >= 0.95 * r['最高價']:
                signals.append('價格突破')
                score += 3

        # 5. 站上均線
        if pd.notna(r.get('收盤價')) and pd.notna(r.get('平均價格')) and r['平均價格'] > 0:
            if r['收盤價'] > r['平均價格'] and r['收盤價'] < 1.05 * r['平均價格']:
                signals.append('站上均線')
                score += 2

        # 6. 連續上漲
        if pd.notna(r.get('漲跌幅%')) and r['漲跌幅%'] > 3:
            signals.append('連續上漲')
            score += 2

        # 7. 低檔起漲
        if pd.notna(r.get('反彈幅度%')) and pd.notna(r.get('收盤價')) and pd.notna(r.get('平均價格')):
            if r['反彈幅度%'] > 10 and r['平均價格'] > 0 and r['收盤價'] < 1.10 * r['平均價格']:
                signals.append('低檔起漲')
                score += 4

        # 8. 急漲訊號
        if pd.notna(r.get('漲跌幅%')) and r['漲跌幅%'] > 5:
            signals.append('急漲訊號')
            score += 3

        # 9. 量價齊揚
        if pd.notna(r.get('漲跌幅%')) and pd.notna(r.get('量比')):
            if r['漲跌幅%'] > 0 and r['量比'] > 1.2:
                signals.append('量價齊揚')
                score += 3

        # 10. 突破整理
        if pd.notna(r.get('振幅%')) and pd.notna(r.get('收盤價')) and pd.notna(r.get('最高價')):
            if r['振幅%'] < 15 and r['最高價'] > 0 and r['收盤價'] >= 0.95 * r['最高價']:
                signals.append('突破整理')
                score += 4

        # 11. KD超賣反彈
        if pd.notna(r.get('平均D')) and pd.notna(r.get('K值')) and pd.notna(r.get('D值')):
            if r['平均D'] < 30 and r['K值'] > r['D值']:
                signals.append('KD超賣反彈')
                score += 3

        # 12. 布林突破
        if pd.notna(r.get('收盤價')) and pd.notna(r.get('最高價')) and pd.notna(r.get('K值')):
            if r['最高價'] > 0 and r['收盤價'] >= 0.95 * r['最高價'] and r['K值'] > 70:
                signals.append('布林突破')
                score += 2

        # 13. MACD翻多
        if pd.notna(r.get('平均RSI')) and pd.notna(r.get('RSI')) and pd.notna(r.get('K值')) and pd.notna(r.get('D值')):
            if r['平均RSI'] < 45 and r['RSI'] > 50 and r['K值'] > r['D值']:
                signals.append('MACD翻多')
                score += 3

        # 14. 多重訊號（≥3 個訊號自動觸發）
        if len(signals) >= 3:
            signals.append('多重訊號')
            score += 5

        results.append({
            '股票代號': r['股票代號'],
            '強度評分': score,
            '訊號數量': len(signals),
            '訊號': ','.join(signals) if signals else ''
        })

    breakout_df = pd.DataFrame(results)
    df_today = df_today.merge(breakout_df, on='股票代號', how='left')
    return df_today


def create_png_icon(size, bg=(102, 126, 234)):
    """純 Python 產生 PNG icon（紫色底色）"""
    w = h = size
    r, g, b = bg
    row = b'\x00' + struct.pack('BBBB', r, g, b, 255) * w
    raw = row * h

    def chunk(ctype, data):
        c = ctype + data
        return struct.pack('>I', len(data)) + c + struct.pack('>I', zlib.crc32(c) & 0xffffffff)

    png = b'\x89PNG\r\n\x1a\n'
    png += chunk(b'IHDR', struct.pack('>IIBBBBB', w, h, 8, 6, 0, 0, 0))
    png += chunk(b'IDAT', zlib.compress(raw, 9))
    png += chunk(b'IEND', b'')
    return png


def generate_icons(icons_dir):
    """產生 PWA icons — 優先用 icons/ 手製圖檔，否則自動產生"""
    src_dir = 'icons'
    sizes = [144, 192, 512]
    copied = 0
    for size in sizes:
        src = os.path.join(src_dir, f'icon-{size}.png')
        dst = os.path.join(icons_dir, f'icon-{size}.png')
        if os.path.exists(src):
            shutil.copy2(src, dst)
            copied += 1
    if copied == len(sizes):
        print("  icons: custom (from icons/)")
        return

    # Fallback: Pillow or solid color
    try:
        from PIL import Image, ImageDraw
        for size in sizes:
            dst = os.path.join(icons_dir, f'icon-{size}.png')
            if os.path.exists(dst):
                continue
            img = Image.new('RGBA', (size, size), (102, 126, 234, 255))
            draw = ImageDraw.Draw(img)
            bars = [0.3, 0.5, 0.4, 0.7, 0.6, 0.85, 0.75]
            bw = size // 14
            margin = (size - bw * 13) // 2
            for i, h in enumerate(bars):
                x = margin + i * bw * 2
                y_top = int(size * (1 - h) * 0.6 + size * 0.15)
                y_bot = int(size * 0.85)
                draw.rounded_rectangle(
                    [x, y_top, x + bw, y_bot],
                    radius=max(bw // 4, 1),
                    fill=(255, 255, 255, 220)
                )
            img.save(dst)
        print("  icons: Pillow fallback")
    except Exception:
        for size in sizes:
            dst = os.path.join(icons_dir, f'icon-{size}.png')
            if os.path.exists(dst):
                continue
            data = create_png_icon(size)
            with open(dst, 'wb') as f:
                f.write(data)
        print("  icons: solid color fallback")


# ==================== HTML Template ====================
HTML_TEMPLATE = r'''<!DOCTYPE html>
<html lang="zh-TW">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=5.0">
<meta name="theme-color" content="#667eea">
<meta name="apple-mobile-web-app-capable" content="yes">
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent">
<link rel="manifest" href="manifest.json">
<link rel="apple-touch-icon" sizes="144x144" href="icons/icon-144.png">
<link rel="apple-touch-icon" sizes="192x192" href="icons/icon-192.png">
<title>台股篩選器</title>
<style>
*{margin:0;padding:0;box-sizing:border-box;-webkit-tap-highlight-color:transparent}
body{font-family:"PingFang TC","Heiti TC","Noto Sans TC","Microsoft JhengHei",-apple-system,BlinkMacSystemFont,sans-serif;background:linear-gradient(135deg,#667eea 0%,#764ba2 100%);min-height:100vh;padding:12px;padding-top:max(12px,env(safe-area-inset-top))}
.container{max-width:800px;margin:0 auto;padding-bottom:env(safe-area-inset-bottom)}
.card{background:white;border-radius:16px;padding:16px;margin-bottom:12px;box-shadow:0 4px 15px rgba(0,0,0,0.1)}
h1{font-size:22px;color:#333;margin-bottom:4px}
.info{color:#888;font-size:13px}
.search-input{width:100%;padding:14px 16px;border:2px solid #e8e8e8;border-radius:12px;font-size:16px;outline:none;transition:border-color .2s}
.search-input:focus{border-color:#667eea}
.section-title{font-size:14px;font-weight:bold;color:#667eea;margin:16px 0 8px}
.section-title:first-child{margin-top:0}
.cat-label{font-size:13px;color:#888;margin:12px 0 6px}
.cat-label:first-of-type{margin-top:0}
.btn{display:inline-block;background:#f0f0f5;color:#555;padding:10px 16px;margin:4px 4px 4px 0;border-radius:10px;border:2px solid transparent;font-size:14px;cursor:pointer;user-select:none;touch-action:manipulation;transition:all .15s}
.btn:active{transform:scale(.95)}
.btn.active{background:#667eea;color:white;border-color:#5a6fd6}
.logic-box{background:#f5f5fa;padding:12px;border-radius:10px;margin:12px 0}
.logic-box label{display:flex;align-items:center;margin:6px 0;cursor:pointer;font-size:14px}
.logic-box input[type="radio"]{margin-right:8px;width:18px;height:18px;accent-color:#667eea}
.input-group{margin:8px 0}
.input-group label{display:block;font-size:13px;color:#888;margin-bottom:4px}
.input-group input[type="number"]{width:100%;padding:12px;border:2px solid #e8e8e8;border-radius:10px;font-size:16px;outline:none;transition:border-color .2s}
.input-group input:focus{border-color:#667eea}
.input-row{display:grid;grid-template-columns:1fr 1fr;gap:10px}
.cb-group{margin:8px 0}
.cb-group label{display:flex;align-items:center;padding:8px 0;font-size:14px;color:#555;cursor:pointer}
.cb-group input[type="checkbox"]{margin-right:10px;width:20px;height:20px;accent-color:#667eea}
.divider{height:1px;background:#eee;margin:16px 0}
.result-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:12px}
.result-count{font-weight:bold;font-size:17px;color:#333}
.view-toggle{display:flex;gap:4px}
.view-btn{padding:8px 16px;background:#f0f0f5;border:none;border-radius:8px;font-size:13px;cursor:pointer;transition:all .15s}
.view-btn.active{background:#667eea;color:white}
.stock-item{background:#f8f8fc;border-radius:12px;padding:12px;margin-bottom:10px;border-left:4px solid #667eea}
.stock-header{display:flex;justify-content:space-between;align-items:center;margin-bottom:8px}
.stock-name{font-size:15px;font-weight:bold;color:#333}
.stock-price{font-size:17px;font-weight:bold;color:#e74c3c}
.stock-tags{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:8px}
.tag{font-size:11px;padding:3px 8px;border-radius:6px;font-weight:500}
.tag-emg{font-size:10px;padding:1px 5px;border-radius:4px;background:#fff3e0;color:#e65100;font-weight:bold;vertical-align:middle}
.tg{background:#e8f5e9;color:#2e7d32}
.tr{background:#fce4ec;color:#c62828}
.tb{background:#e3f2fd;color:#1565c0}
.stock-details{display:grid;grid-template-columns:repeat(3,1fr);gap:4px 8px;font-size:12px;color:#888}
.stock-details .v{color:#333;font-weight:500}
.hidden{display:none}
.table-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch;margin:0 -16px;padding:0 16px}
table{width:100%;border-collapse:collapse;font-size:12px;min-width:860px}
th{background:#667eea;color:white;padding:10px 6px;text-align:left;position:sticky;top:0;cursor:pointer;user-select:none;white-space:nowrap;font-size:12px}
th.sa::after{content:" \2191"}
th.sd::after{content:" \2193"}
th:not(.sa):not(.sd)::after{content:" \2195";opacity:.4}
td{padding:8px 6px;border-bottom:1px solid #f0f0f0;white-space:nowrap}
tr:nth-child(even){background:#fafafa}
.main-btn{width:100%;padding:16px;background:#667eea;color:white;border:none;border-radius:12px;font-size:16px;font-weight:bold;cursor:pointer;margin-top:12px}
.main-btn:active{background:#5a6fd6;transform:scale(.98)}
.empty{text-align:center;padding:40px 20px;color:#999}
.ghint{font-size:12px;color:#aaa;font-style:italic;margin-bottom:8px}
.trend-row{display:flex;gap:4px;flex-wrap:wrap;margin-bottom:6px}
.trend-badge{font-size:11px;padding:2px 7px;border-radius:5px;font-weight:500}
.tb-up{background:#fce4ec;color:#c62828}
.tb-down{background:#e8f5e9;color:#2e7d32}
.tb-flat{background:#f5f5f5;color:#999}
.strength-bar{margin:6px 0;position:relative;height:14px;background:#e8e8e8;border-radius:7px;overflow:hidden}
.strength-fill{height:100%;border-radius:7px;background:linear-gradient(90deg,#66bb6a,#43a047);transition:width .3s}
.strength-text{position:absolute;right:6px;top:0;line-height:14px;font-size:10px;color:#333;font-weight:bold}
.signal-tags{display:flex;gap:4px;flex-wrap:wrap;margin-top:4px}
.signal-tag{font-size:10px;padding:2px 6px;border-radius:4px;background:#e3f2fd;color:#1565c0}
.search-wrap{position:relative}
.ac-list{position:absolute;top:100%;left:0;right:0;background:white;border:2px solid #667eea;border-top:none;border-radius:0 0 12px 12px;max-height:260px;overflow-y:auto;z-index:100;display:none}
.ac-list.show{display:block}
.ac-item{padding:10px 16px;cursor:pointer;font-size:14px;border-bottom:1px solid #f0f0f0;display:flex;justify-content:space-between;align-items:center}
.ac-item:last-child{border-bottom:none}
.ac-item:active,.ac-item.hi{background:#f0f0ff}
.ac-item .ac-code{font-weight:bold;color:#333}
.ac-item .ac-name{color:#888;font-size:13px}
.ac-item .ac-price{color:#e74c3c;font-weight:500;font-size:13px}
@media(max-width:400px){
.stock-details{grid-template-columns:repeat(2,1fr)}
.input-row{grid-template-columns:1fr}
}
.filter-panel{background:#fff;border-radius:16px;margin-bottom:12px;box-shadow:0 4px 15px rgba(0,0,0,0.1)}
.filter-header{padding:14px 16px;font-weight:bold;cursor:pointer;display:flex;justify-content:space-between;align-items:center;font-size:15px}
#filterBody{padding:8px 16px 16px}
.slider-row{display:grid;grid-template-columns:56px 1fr 68px;align-items:center;gap:8px;margin:10px 0}
.slider-label{font-size:12px;color:#666}
.slider-val{font-size:12px;color:#667eea;text-align:right;font-variant-numeric:tabular-nums}
.slider-track{position:relative;height:36px}
.slider-track::before{content:'';position:absolute;width:100%;height:4px;background:#e0e0e0;top:50%;transform:translateY(-50%);border-radius:2px}
.slider-range{position:absolute;height:4px;background:#667eea;top:50%;transform:translateY(-50%);border-radius:2px}
.slider-track input[type=range]{position:absolute;width:100%;top:50%;transform:translateY(-50%);pointer-events:none;-webkit-appearance:none;appearance:none;background:transparent;margin:0;height:22px}
.slider-track input[type=range]::-webkit-slider-thumb{pointer-events:auto;-webkit-appearance:none;width:22px;height:22px;border-radius:50%;background:#667eea;cursor:pointer;border:2px solid #fff;box-shadow:0 1px 4px rgba(0,0,0,.3)}
.slider-track input[type=range]::-moz-range-thumb{pointer-events:auto;width:18px;height:18px;border-radius:50%;background:#667eea;cursor:pointer;border:2px solid #fff;box-shadow:0 1px 4px rgba(0,0,0,.3)}
.reset-btn{width:100%;padding:12px;background:#f0f0f5;color:#667eea;border:none;border-radius:10px;font-size:14px;font-weight:bold;cursor:pointer;margin-top:8px}
.reset-btn:active{background:#e0e0ea;transform:scale(.98)}
.chart-overlay{position:fixed;top:0;left:0;right:0;bottom:0;background:rgba(0,0,0,0.6);z-index:200;display:none;align-items:center;justify-content:center}
.chart-overlay.show{display:flex}
.chart-modal{background:white;border-radius:16px;width:95%;max-width:700px;max-height:90vh;overflow:hidden;position:relative}
.chart-header{padding:12px 16px;display:flex;justify-content:space-between;align-items:center;border-bottom:1px solid #eee}
.chart-title{font-size:16px;font-weight:bold;color:#333}
.chart-close{width:32px;height:32px;border:none;background:#f0f0f5;border-radius:8px;font-size:18px;cursor:pointer;display:flex;align-items:center;justify-content:center}
.chart-legend{padding:6px 16px 0;display:flex;gap:12px;font-size:12px;font-weight:bold}
.chart-body{padding:8px}
#chartContainer{width:100%;height:400px}
</style>
</head>
<body>
<div class="container">

<div class="card">
<h1>📊 台股篩選器</h1>
<div class="info">更新：__UPDATE_TIME__ ｜ 共 __TOTAL_COUNT__ 檔</div>
<div style="margin-top:8px;display:flex;align-items:center;gap:8px;flex-wrap:wrap">
<span id="todayDate" style="font-size:18px;font-weight:bold;color:#333"></span>
<span style="font-size:13px;color:#aaa;margin-left:4px">vs</span>
<select id="compareDate" onchange="updateCompareDate()" style="padding:6px 10px;border:2px solid #e8e8e8;border-radius:8px;font-size:14px;outline:none;background:#fff"></select>
<span id="compareInfo" style="font-size:12px;color:#aaa"></span>
<span style="margin-left:auto;font-size:18px">📈</span>
</div>
</div>

<div class="card">
<div class="search-wrap">
<input type="text" class="search-input" id="searchInput" placeholder="🔍 搜尋代號或名稱..." autocomplete="off">
<div class="ac-list" id="acList"></div>
</div>
</div>

<div class="card">
<div class="section-title" style="margin-top:0">📍 策略選擇（可複選）</div>

<div class="cat-label">長線投資</div>
<div class="btn" data-st="dividend">高殖利率</div>
<div class="btn" data-st="value">價值投資</div>

<div class="cat-label">短線操作</div>
<div class="btn" data-st="kd_golden">KD黃金交叉</div>
<div class="btn" data-st="bb_upper">布林上軌</div>
<div class="btn" data-st="macd_golden">MACD黃金交叉</div>

<div class="cat-label">超賣反彈</div>
<div class="btn" data-st="kd_oversold">KD超賣</div>
<div class="btn" data-st="rsi_oversold">RSI超賣</div>
<div class="btn" data-st="bb_lower">布林下軌</div>

<div class="cat-label">突破訊號</div>
<div class="btn" data-st="breakout_high">高強度</div>
<div class="btn" data-st="breakout_mid">中強度</div>
<div class="btn" data-st="breakout_multi">多重訊號</div>

<div class="cat-label">漲跌停</div>
<div class="btn" data-st="limit_up">今日漲停</div>
<div class="btn" data-st="limit_down">今日跌停</div>

<div class="cat-label">趨勢轉強</div>
<div class="btn" data-st="trend_price">價格上漲</div>
<div class="btn" data-st="trend_kd">KD轉強</div>
<div class="btn" data-st="trend_rsi">RSI轉強</div>
<div class="btn" data-st="trend_vol">量能放大</div>

<div id="stParams" class="hidden">
<div class="logic-box">
<div style="font-weight:bold;margin-bottom:6px;font-size:14px">策略組合</div>
<label><input type="radio" name="logic" value="or" checked> OR — 任一符合</label>
<label><input type="radio" name="logic" value="and"> AND — 全部符合</label>
</div>

<div class="section-title">⚙️ 策略參數</div>
<div class="input-row">
<div class="input-group"><label>殖利率閾值（%）</label><input type="number" id="pDiv" value="5" step="0.5"></div>
<div class="input-group"><label>本益比上限</label><input type="number" id="pPE" value="20" step="1"></div>
</div>
<div class="input-row">
<div class="input-group"><label>K值上限（超賣）</label><input type="number" id="pK" value="20" step="5"></div>
<div class="input-group"><label>RSI上限（超賣）</label><input type="number" id="pRSI" value="30" step="5"></div>
</div>
</div>

</div>

<div class="filter-panel">
<div class="filter-header" onclick="togglePanel()">📊 篩選面板 <span id="panelArrow">▶</span></div>
<div id="filterBody" style="display:none">
<div class="slider-row" data-key="收盤價" data-min="0" data-max="3000" data-step="10"><span class="slider-label">收盤價</span><div class="slider-track"><div class="slider-range"></div><input type="range" class="slider-lo" min="0" max="3000" step="10" value="0"><input type="range" class="slider-hi" min="0" max="3000" step="10" value="3000"></div><span class="slider-val">0~3000</span></div>
<div class="slider-row" data-key="K值" data-min="0" data-max="100" data-step="1"><span class="slider-label">K值</span><div class="slider-track"><div class="slider-range"></div><input type="range" class="slider-lo" min="0" max="100" step="1" value="0"><input type="range" class="slider-hi" min="0" max="100" step="1" value="100"></div><span class="slider-val">0~100</span></div>
<div class="slider-row" data-key="D值" data-min="0" data-max="100" data-step="1"><span class="slider-label">D值</span><div class="slider-track"><div class="slider-range"></div><input type="range" class="slider-lo" min="0" max="100" step="1" value="0"><input type="range" class="slider-hi" min="0" max="100" step="1" value="100"></div><span class="slider-val">0~100</span></div>
<div class="slider-row" data-key="RSI" data-min="0" data-max="100" data-step="1"><span class="slider-label">RSI</span><div class="slider-track"><div class="slider-range"></div><input type="range" class="slider-lo" min="0" max="100" step="1" value="0"><input type="range" class="slider-hi" min="0" max="100" step="1" value="100"></div><span class="slider-val">0~100</span></div>
<div class="slider-row" data-key="殖利率" data-min="0" data-max="20" data-step="0.5"><span class="slider-label">殖利率</span><div class="slider-track"><div class="slider-range"></div><input type="range" class="slider-lo" min="0" max="20" step="0.5" value="0"><input type="range" class="slider-hi" min="0" max="20" step="0.5" value="20"></div><span class="slider-val">0~20</span></div>
<div class="slider-row" data-key="本益比" data-min="0" data-max="200" data-step="5"><span class="slider-label">本益比</span><div class="slider-track"><div class="slider-range"></div><input type="range" class="slider-lo" min="0" max="200" step="5" value="0"><input type="range" class="slider-hi" min="0" max="200" step="5" value="200"></div><span class="slider-val">0~200</span></div>
<div class="slider-row" data-key="市值億" data-min="0" data-max="5000" data-step="50"><span class="slider-label">市值(億)</span><div class="slider-track"><div class="slider-range"></div><input type="range" class="slider-lo" min="0" max="5000" step="50" value="0"><input type="range" class="slider-hi" min="0" max="5000" step="50" value="5000"></div><span class="slider-val">0~5000</span></div>
<div class="slider-row" data-key="量比" data-min="0" data-max="10" data-step="0.1"><span class="slider-label">量比</span><div class="slider-track"><div class="slider-range"></div><input type="range" class="slider-lo" min="0" max="10" step="0.1" value="0"><input type="range" class="slider-hi" min="0" max="10" step="0.1" value="10"></div><span class="slider-val">0~10</span></div>
<div class="slider-row" data-key="強度評分" data-min="0" data-max="40" data-step="1"><span class="slider-label">強度</span><div class="slider-track"><div class="slider-range"></div><input type="range" class="slider-lo" min="0" max="40" step="1" value="0"><input type="range" class="slider-hi" min="0" max="40" step="1" value="40"></div><span class="slider-val">0~40</span></div>
<button class="reset-btn" onclick="resetSliders()">重置篩選</button>
</div>
</div>

<div class="card hidden" id="resultsCard">
<div class="result-header">
<div class="result-count">找到 <span id="count">0</span> 筆</div>
<div class="view-toggle">
<button class="view-btn active" id="cardViewBtn">📱 卡片</button>
<button class="view-btn" id="tableViewBtn">📊 表格</button>
</div>
</div>

<div id="cardView"><div id="cardList"></div></div>

<div id="tableView" class="hidden">
<div class="table-wrap">
<table>
<thead><tr>
<th data-col="股票代號">代號</th>
<th data-col="股票名稱">名稱</th>
<th data-col="收盤價">收盤</th>
<th data-col="K值">K</th>
<th data-col="D值">D</th>
<th data-col="kd_status">KD</th>
<th data-col="RSI">RSI</th>
<th data-col="macd_status">MACD</th>
<th data-col="bb_pos">BB%</th>
<th data-col="成交量張">成交量</th>
<th data-col="量比">量比</th>
<th data-col="殖利率">殖利率</th>
<th data-col="本益比">本益比</th>
<th data-col="市值億">市值億</th>
<th data-col="漲跌幅">漲跌%</th>
<th data-col="強度評分">強度</th>
<th data-col="訊號數量">訊號數</th>
</tr></thead>
<tbody id="tableBody"></tbody>
</table>
</div>
</div>

</div>

</div>

<div class="chart-overlay" id="chartOverlay" onclick="closeChart(event)">
<div class="chart-modal" onclick="event.stopPropagation()">
<div class="chart-header">
<span class="chart-title" id="chartTitle"></span>
<button class="chart-close" onclick="closeChart()">&#x2715;</button>
</div>
<div class="chart-legend">
<span style="color:#ff9800">MA5</span>
<span style="color:#2196f3">MA10</span>
<span style="color:#e91e63">MA20</span>
<span style="color:#9c27b0">MA60</span>
<span style="color:#4caf50">MA240</span>
</div>
<div class="chart-body">
<div id="chartContainer"></div>
</div>
</div>
</div>

<script src="https://unpkg.com/lightweight-charts@4.1.0/dist/lightweight-charts.standalone.production.js"></script>
<script>
var RAW = __STOCK_DATA__;
var HIST = __HIST_DATA__;
var OHLC_CACHE = {};
var COMPARE_DEFAULT = '__COMPARE_DEFAULT__';
var CURRENT_TRADE_DATE = '__CURRENT_TRADE_DATE__';
var data = [];
var filtered = [];
var strategies = [];
var sortCol = null;
var sortDir = 'asc';
var viewMode = 'card';
var debounceT = null;

function ok(v) { return v != null && v !== ''; }
function num(v) { var x = parseFloat(v); return isNaN(x) ? 0 : x; }

document.addEventListener('DOMContentLoaded', function() {
    // 顯示當日日期
    var tm = parseInt(CURRENT_TRADE_DATE.substring(4, 6));
    var td = parseInt(CURRENT_TRADE_DATE.substring(6, 8));
    document.getElementById('todayDate').textContent = tm + '/' + td;

    // 計算衍生欄位
    for (var i = 0; i < RAW.length; i++) {
        var s = RAW[i];
        s.kd_status = (ok(s.K值) && ok(s.D值)) ? (s.K值 > s.D值 ? '黃金交叉' : '死亡交叉') : '';
        s.macd_status = (ok(s.DIF) && ok(s.MACD)) ? (s.DIF > s.MACD ? '黃金交叉' : '死亡交叉') : '';
        s.bb_pos = (ok(s.BB上軌) && ok(s.BB下軌) && s.BB上軌 !== s.BB下軌)
            ? Math.round((s.收盤價 - s.BB下軌) / (s.BB上軌 - s.BB下軌) * 1000) / 10
            : null;
        data.push(s);
    }

    // 策略按鈕（只切換外觀）
    var btns = document.querySelectorAll('.btn[data-st]');
    for (var i = 0; i < btns.length; i++) {
        btns[i].addEventListener('click', function() {
            var st = this.getAttribute('data-st');
            var idx = strategies.indexOf(st);
            if (idx >= 0) { strategies.splice(idx, 1); this.classList.remove('active'); }
            else { strategies.push(st); this.classList.add('active'); }
            document.getElementById('stParams').className = strategies.length ? '' : 'hidden';
            doFilter();
            document.getElementById('resultsCard').className = 'card';
        });
    }

    // 初始化滑桿
    initSliders();

    // 初始化比較基準選單
    initCompareSelect();

    // 搜尋 + 自動建議
    var acEl = document.getElementById('acList');
    var searchEl = document.getElementById('searchInput');
    var acIdx = -1;

    function showAC(q) {
        if (!q) { acEl.className = 'ac-list'; acIdx = -1; return; }
        q = q.toLowerCase();
        var hits = [];
        for (var i = 0; i < data.length && hits.length < 15; i++) {
            var s = data[i];
            var code = String(s.股票代號).toLowerCase();
            var name = String(s.股票名稱 || '').toLowerCase();
            if (code.indexOf(q) >= 0 || name.indexOf(q) >= 0) hits.push(s);
        }
        if (hits.length === 0) { acEl.className = 'ac-list'; acIdx = -1; return; }
        var h = '';
        for (var i = 0; i < hits.length; i++) {
            var s = hits[i];
            h += '<div class="ac-item" data-code="' + s.股票代號 + '">';
            h += '<span><span class="ac-code">' + s.股票代號 + '</span> <span class="ac-name">' + (s.股票名稱 || '') + '</span></span>';
            h += '<span class="ac-price">$' + (ok(s.收盤價) ? s.收盤價 : '-') + '</span>';
            h += '</div>';
        }
        acEl.innerHTML = h;
        acEl.className = 'ac-list show';
        acIdx = -1;

        var items = acEl.querySelectorAll('.ac-item');
        for (var i = 0; i < items.length; i++) {
            items[i].addEventListener('mousedown', function(e) {
                e.preventDefault();
                pickAC(this.getAttribute('data-code'));
            });
        }
    }

    function pickAC(code) {
        searchEl.value = code;
        acEl.className = 'ac-list';
        acIdx = -1;
        doFilter();
        document.getElementById('resultsCard').className = 'card';
    }

    function highlightAC(idx) {
        var items = acEl.querySelectorAll('.ac-item');
        for (var i = 0; i < items.length; i++) items[i].classList.remove('hi');
        if (idx >= 0 && idx < items.length) { items[idx].classList.add('hi'); items[idx].scrollIntoView({block:'nearest'}); }
    }

    searchEl.addEventListener('input', function() {
        clearTimeout(debounceT);
        var q = this.value.trim();
        debounceT = setTimeout(function() {
            showAC(q);
            doFilter();
            document.getElementById('resultsCard').className = 'card';
        }, 100);
    });

    searchEl.addEventListener('keydown', function(e) {
        var items = acEl.querySelectorAll('.ac-item');
        if (!items.length || acEl.className.indexOf('show') < 0) return;
        if (e.key === 'ArrowDown') { e.preventDefault(); acIdx = Math.min(acIdx + 1, items.length - 1); highlightAC(acIdx); }
        else if (e.key === 'ArrowUp') { e.preventDefault(); acIdx = Math.max(acIdx - 1, 0); highlightAC(acIdx); }
        else if (e.key === 'Enter' && acIdx >= 0) { e.preventDefault(); pickAC(items[acIdx].getAttribute('data-code')); }
    });

    searchEl.addEventListener('blur', function() {
        setTimeout(function() { acEl.className = 'ac-list'; acIdx = -1; }, 150);
    });


    // 檢視切換
    document.getElementById('cardViewBtn').addEventListener('click', function() { setView('card'); });
    document.getElementById('tableViewBtn').addEventListener('click', function() { setView('table'); });

    // 排序
    var ths = document.querySelectorAll('th[data-col]');
    for (var i = 0; i < ths.length; i++) {
        ths[i].addEventListener('click', function() {
            var col = this.getAttribute('data-col');
            if (sortCol === col) sortDir = sortDir === 'asc' ? 'desc' : 'asc';
            else { sortCol = col; sortDir = 'asc'; }
            sortAndDisplay();
        });
    }
});

function updateCompareDate() {
    var dateKey = document.getElementById('compareDate').value;
    var hist = HIST[dateKey] || {};
    for (var i = 0; i < data.length; i++) {
        var s = data[i];
        var code = String(s.股票代號);
        var oldPrice = hist[code];
        if (ok(s.收盤價) && oldPrice != null && oldPrice > 0) {
            s.漲跌幅 = Math.round((s.收盤價 - oldPrice) / oldPrice * 10000) / 100;
            s.漲跌額 = Math.round((s.收盤價 - oldPrice) * 100) / 100;
            if (s.漲跌幅 >= 3) s.價格趨勢 = '大漲';
            else if (s.漲跌幅 > 0) s.價格趨勢 = '上漲';
            else if (s.漲跌幅 <= -3) s.價格趨勢 = '下跌';
            else s.價格趨勢 = '平穩';
        } else {
            s.漲跌幅 = null;
            s.漲跌額 = null;
            s.價格趨勢 = '平穩';
        }
    }
    var cm = parseInt(CURRENT_TRADE_DATE.substring(4, 6));
    var cd = parseInt(CURRENT_TRADE_DATE.substring(6, 8));
    var m = parseInt(dateKey.substring(4, 6));
    var d = parseInt(dateKey.substring(6, 8));
    document.getElementById('compareInfo').textContent = cm + '/' + cd + ' vs ' + m + '/' + d + ' 收盤價';
    if (document.getElementById('resultsCard').className.indexOf('hidden') < 0) {
        doFilter();
    }
}

function initCompareSelect() {
    var sel = document.getElementById('compareDate');
    var dates = Object.keys(HIST).sort().reverse();
    for (var i = 0; i < dates.length; i++) {
        var opt = document.createElement('option');
        opt.value = dates[i];
        var m = parseInt(dates[i].substring(4, 6));
        var d = parseInt(dates[i].substring(6, 8));
        var cnt = Object.keys(HIST[dates[i]]).length;
        opt.textContent = m + '/' + d + ' (' + cnt + '筆)';
        if (dates[i] === COMPARE_DEFAULT) opt.selected = true;
        sel.appendChild(opt);
    }
    if (dates.length) updateCompareDate();
}

function togglePanel() {
    var body = document.getElementById('filterBody');
    var arrow = document.getElementById('panelArrow');
    if (body.style.display === 'none') { body.style.display = ''; arrow.textContent = '▼'; }
    else { body.style.display = 'none'; arrow.textContent = '▶'; }
}

function initSliders() {
    var rows = document.querySelectorAll('.slider-row');
    for (var i = 0; i < rows.length; i++) {
        (function(row) {
            var lo = row.querySelector('.slider-lo');
            var hi = row.querySelector('.slider-hi');
            lo.addEventListener('input', function() { onSliderInput(row); });
            hi.addEventListener('input', function() { onSliderInput(row); });
        })(rows[i]);
        updateSliderBar(rows[i]);
    }
}

function updateSliderBar(row) {
    var lo = row.querySelector('.slider-lo');
    var hi = row.querySelector('.slider-hi');
    var loV = parseFloat(lo.value), hiV = parseFloat(hi.value);
    var mn = parseFloat(lo.min), mx = parseFloat(lo.max);
    var pctL = (Math.min(loV, hiV) - mn) / (mx - mn) * 100;
    var pctR = (Math.max(loV, hiV) - mn) / (mx - mn) * 100;
    var bar = row.querySelector('.slider-range');
    bar.style.left = pctL + '%';
    bar.style.width = (pctR - pctL) + '%';
    row.querySelector('.slider-val').textContent = Math.min(loV, hiV) + '~' + Math.max(loV, hiV);
}

function onSliderInput(row) {
    updateSliderBar(row);
    clearTimeout(debounceT);
    debounceT = setTimeout(function() { doFilter(); document.getElementById('resultsCard').className = 'card'; }, 80);
}

function resetSliders() {
    var rows = document.querySelectorAll('.slider-row');
    for (var i = 0; i < rows.length; i++) {
        var lo = rows[i].querySelector('.slider-lo');
        var hi = rows[i].querySelector('.slider-hi');
        lo.value = lo.min;
        hi.value = hi.max;
        updateSliderBar(rows[i]);
    }
    doFilter();
    document.getElementById('resultsCard').className = 'card';
}

function doFilter() {
    var search = document.getElementById('searchInput').value.trim().toLowerCase();
    var logicEl = document.querySelector('input[name="logic"]:checked');
    var logic = logicEl ? logicEl.value : 'or';

    var pDiv = num(document.getElementById('pDiv').value);
    var pPE = num(document.getElementById('pPE').value);
    var pK = num(document.getElementById('pK').value);
    var pRSI = num(document.getElementById('pRSI').value);

    // 讀取滑桿範圍
    var sliderRows = document.querySelectorAll('.slider-row');
    var sliderFilters = [];
    for (var si = 0; si < sliderRows.length; si++) {
        var row = sliderRows[si];
        var key = row.getAttribute('data-key');
        var mn = parseFloat(row.getAttribute('data-min'));
        var mx = parseFloat(row.getAttribute('data-max'));
        var lo = parseFloat(row.querySelector('.slider-lo').value);
        var hi = parseFloat(row.querySelector('.slider-hi').value);
        var loV = Math.min(lo, hi), hiV = Math.max(lo, hi);
        if (loV > mn || hiV < mx) {
            sliderFilters.push({key: key, lo: loV, hi: hiV});
        }
    }

    filtered = [];
    for (var i = 0; i < data.length; i++) {
        var s = data[i];

        // 文字搜尋
        if (search) {
            var code = String(s.股票代號).toLowerCase();
            var name = String(s.股票名稱 || '').toLowerCase();
            if (code.indexOf(search) < 0 && name.indexOf(search) < 0) continue;
        }

        // 策略篩選
        if (strategies.length > 0) {
            var mc = 0;
            for (var j = 0; j < strategies.length; j++) {
                var st = strategies[j];
                var pass = false;
                if (st === 'dividend') pass = ok(s.殖利率) && s.殖利率 >= pDiv;
                else if (st === 'value') pass = ok(s.殖利率) && s.殖利率 >= pDiv && ok(s.本益比) && s.本益比 > 0 && s.本益比 <= pPE;
                else if (st === 'kd_golden') pass = s.kd_status === '黃金交叉';
                else if (st === 'bb_upper') pass = ok(s.bb_pos) && s.bb_pos >= 80;
                else if (st === 'macd_golden') pass = s.macd_status === '黃金交叉';
                else if (st === 'kd_oversold') pass = ok(s.K值) && s.K值 <= pK;
                else if (st === 'rsi_oversold') pass = ok(s.RSI) && s.RSI <= pRSI;
                else if (st === 'bb_lower') pass = ok(s.bb_pos) && s.bb_pos <= 20;
                else if (st === 'breakout_high') pass = ok(s.強度評分) && s.強度評分 >= 20;
                else if (st === 'breakout_mid') pass = ok(s.強度評分) && s.強度評分 >= 10;
                else if (st === 'breakout_multi') pass = ok(s.訊號數量) && s.訊號數量 >= 3;
                else if (st === 'limit_up') pass = ok(s.漲跌幅) && s.漲跌幅 >= 9.5;
                else if (st === 'limit_down') pass = ok(s.漲跌幅) && s.漲跌幅 <= -9.5;
                else if (st === 'trend_price') pass = s.價格趨勢 === '上漲' || s.價格趨勢 === '大漲';
                else if (st === 'trend_kd') pass = s.KD趨勢 === 'KD轉強';
                else if (st === 'trend_rsi') pass = s.RSI趨勢 === 'RSI轉強';
                else if (st === 'trend_vol') pass = s.量能趨勢 === '量能放大';
                if (pass) mc++;
            }
            if (logic === 'or' && mc === 0) continue;
            if (logic === 'and' && mc < strategies.length) continue;
        }

        // 滑桿篩選
        var skip = false;
        for (var si = 0; si < sliderFilters.length; si++) {
            var sf = sliderFilters[si];
            if (!ok(s[sf.key]) || s[sf.key] < sf.lo || s[sf.key] > sf.hi) { skip = true; break; }
        }
        if (skip) continue;

        filtered.push(s);
    }

    sortAndDisplay();
}

function sortAndDisplay() {
    if (sortCol) {
        filtered.sort(function(a, b) {
            var va = a[sortCol], vb = b[sortCol];
            if (!ok(va)) va = sortDir === 'asc' ? '\uffff' : '';
            if (!ok(vb)) vb = sortDir === 'asc' ? '\uffff' : '';
            if (typeof va === 'number' && typeof vb === 'number') {
                return sortDir === 'asc' ? va - vb : vb - va;
            }
            va = String(va); vb = String(vb);
            return sortDir === 'asc' ? va.localeCompare(vb) : vb.localeCompare(va);
        });
    }
    document.getElementById('count').textContent = filtered.length;
    if (viewMode === 'card') renderCards();
    else renderTable();
}

function showResults() {
    document.getElementById('resultsCard').className = 'card';
    setTimeout(function() { document.getElementById('resultsCard').scrollIntoView({behavior:'smooth'}); }, 100);
}

function setView(mode) {
    viewMode = mode;
    document.getElementById('cardView').className = mode === 'card' ? '' : 'hidden';
    document.getElementById('tableView').className = mode === 'table' ? '' : 'hidden';
    document.getElementById('cardViewBtn').className = 'view-btn' + (mode === 'card' ? ' active' : '');
    document.getElementById('tableViewBtn').className = 'view-btn' + (mode === 'table' ? ' active' : '');
    sortAndDisplay();
}

function fv(v, d) { return ok(v) ? Number(v).toFixed(d) : '-'; }

function renderCards() {
    var el = document.getElementById('cardList');
    if (filtered.length === 0) {
        el.innerHTML = '<div class="empty">沒有符合條件的股票</div>';
        return;
    }
    var MAX = 200;
    var show = filtered.length > MAX ? filtered.slice(0, MAX) : filtered;
    var h = '';
    for (var i = 0; i < show.length; i++) {
        var s = show[i];
        var kc = s.kd_status === '黃金交叉' ? 'tg' : 'tr';
        var mc = s.macd_status === '黃金交叉' ? 'tg' : 'tr';
        var bc = ok(s.bb_pos) ? (s.bb_pos >= 80 ? 'tr' : s.bb_pos <= 20 ? 'tb' : 'tg') : '';
        var limitStyle = ok(s.漲跌幅) ? (s.漲跌幅 >= 9.5 ? 'background:#c62828;color:#fff;padding:2px 8px;border-radius:6px' : s.漲跌幅 <= -9.5 ? 'background:#2e7d32;color:#fff;padding:2px 8px;border-radius:6px' : '') : '';

        h += '<div class="stock-item" onclick="openChart(\'' + s.股票代號 + '\')" style="cursor:pointer">';
        h += '<div class="stock-header">';
        h += '<div class="stock-name"' + (limitStyle ? ' style="' + limitStyle + '"' : '') + '>' + s.股票代號 + ' ' + (s.股票名稱 || '') + (s.市場 === '興櫃' ? ' <span class="tag-emg">興櫃</span>' : '') + '</div>';
        h += '<div class="stock-price">$' + (ok(s.收盤價) ? s.收盤價 : '-') + '</div>';
        h += '</div>';
        h += '<div class="stock-tags">';
        if (s.kd_status) h += '<span class="tag ' + kc + '">KD ' + s.kd_status + '</span>';
        if (s.macd_status) h += '<span class="tag ' + mc + '">MACD ' + s.macd_status + '</span>';
        if (bc) h += '<span class="tag ' + bc + '">BB ' + fv(s.bb_pos, 1) + '%</span>';
        h += '</div>';

        // 趨勢指標行
        if (ok(s.價格趨勢)) {
            h += '<div class="trend-row">';
            var pc = s.價格趨勢 === '大漲' || s.價格趨勢 === '上漲' ? 'tb-up' : s.價格趨勢 === '下跌' ? 'tb-down' : 'tb-flat';
            var pLabel = '價格' + (ok(s.漲跌幅) ? (s.漲跌幅 > 0 ? '↑' : s.漲跌幅 < 0 ? '↓' : '') + Math.abs(s.漲跌幅).toFixed(1) + '%' + (ok(s.漲跌額) ? ' (' + (s.漲跌額 > 0 ? '+' : '') + s.漲跌額.toFixed(2) + ')' : '') : '');
            h += '<span class="trend-badge ' + pc + '">' + pLabel + '</span>';
            var kdc = s.KD趨勢 === 'KD轉強' ? 'tb-up' : s.KD趨勢 === 'KD轉弱' ? 'tb-down' : 'tb-flat';
            h += '<span class="trend-badge ' + kdc + '">' + s.KD趨勢 + '</span>';
            var rc = s.RSI趨勢 === 'RSI轉強' ? 'tb-up' : s.RSI趨勢 === 'RSI轉弱' ? 'tb-down' : 'tb-flat';
            h += '<span class="trend-badge ' + rc + '">' + s.RSI趨勢 + '</span>';
            var vc = s.量能趨勢 === '量能放大' ? 'tb-up' : s.量能趨勢 === '量能縮小' ? 'tb-down' : 'tb-flat';
            h += '<span class="trend-badge ' + vc + '">' + s.量能趨勢 + '</span>';
            h += '</div>';
        }

        // 突破強度
        if (ok(s.強度評分) && s.強度評分 > 0) {
            var pct = Math.min(s.強度評分 / 40 * 100, 100);
            h += '<div class="strength-bar"><div class="strength-fill" style="width:' + pct + '%"></div><span class="strength-text">強度 ' + s.強度評分 + '/40</span></div>';
            if (ok(s.訊號) && s.訊號) {
                h += '<div class="signal-tags">';
                var tags = s.訊號.split(',');
                for (var t = 0; t < tags.length; t++) {
                    if (tags[t]) h += '<span class="signal-tag">' + tags[t] + '</span>';
                }
                h += '</div>';
            }
        }

        h += '<div class="stock-details">';
        h += '<div>K <span class="v">' + fv(s.K值, 1) + '</span></div>';
        h += '<div>D <span class="v">' + fv(s.D值, 1) + '</span></div>';
        h += '<div>RSI <span class="v">' + fv(s.RSI, 1) + '</span></div>';
        h += '<div>殖利率 <span class="v">' + (ok(s.殖利率) ? s.殖利率.toFixed(2) + '%' : '-') + '</span></div>';
        h += '<div>本益比 <span class="v">' + fv(s.本益比, 1) + '</span></div>';
        h += '<div>市值 <span class="v">' + (ok(s.市值億) ? Math.round(s.市值億) + '億' : '-') + '</span></div>';
        h += '</div></div>';
    }
    if (filtered.length > MAX) {
        h += '<div class="empty">顯示前 ' + MAX + ' 筆，共 ' + filtered.length + ' 筆符合</div>';
    }
    el.innerHTML = h;
}

function renderTable() {
    var ths = document.querySelectorAll('th[data-col]');
    for (var i = 0; i < ths.length; i++) {
        ths[i].className = '';
        if (ths[i].getAttribute('data-col') === sortCol) ths[i].className = sortDir === 'asc' ? 'sa' : 'sd';
    }
    var tbody = document.getElementById('tableBody');
    if (filtered.length === 0) {
        tbody.innerHTML = '<tr><td colspan="17" class="empty">沒有符合條件的股票</td></tr>';
        return;
    }
    var h = '';
    for (var i = 0; i < filtered.length; i++) {
        var s = filtered[i];
        var kc = s.kd_status === '黃金交叉' ? '#2e7d32' : '#c62828';
        var mc = s.macd_status === '黃金交叉' ? '#2e7d32' : '#c62828';
        var limitStyle = ok(s.漲跌幅) ? (s.漲跌幅 >= 9.5 ? 'background:#c62828;color:#fff;font-weight:bold' : s.漲跌幅 <= -9.5 ? 'background:#2e7d32;color:#fff;font-weight:bold' : '') : '';
        h += '<tr onclick="openChart(\'' + s.股票代號 + '\')" style="cursor:pointer">';
        h += '<td>' + s.股票代號 + '</td>';
        h += '<td' + (limitStyle ? ' style="' + limitStyle + '"' : '') + '>' + (s.股票名稱 || '') + (s.市場 === '興櫃' ? ' <span class="tag-emg">興櫃</span>' : '') + '</td>';
        h += '<td>' + (ok(s.收盤價) ? s.收盤價 : '') + '</td>';
        h += '<td>' + fv(s.K值, 1) + '</td>';
        h += '<td>' + fv(s.D值, 1) + '</td>';
        h += '<td style="color:' + kc + '">' + (s.kd_status || '') + '</td>';
        h += '<td>' + fv(s.RSI, 1) + '</td>';
        h += '<td style="color:' + mc + '">' + (s.macd_status || '') + '</td>';
        h += '<td>' + fv(s.bb_pos, 1) + '</td>';
        h += '<td>' + (ok(s.成交量張) ? s.成交量張 : '') + '</td>';
        h += '<td>' + fv(s.量比, 2) + '</td>';
        h += '<td>' + fv(s.殖利率, 2) + '</td>';
        h += '<td>' + fv(s.本益比, 1) + '</td>';
        h += '<td>' + (ok(s.市值億) ? Math.round(s.市值億) : '') + '</td>';
        var chgColor = ok(s.漲跌幅) ? (s.漲跌幅 > 0 ? '#c62828' : s.漲跌幅 < 0 ? '#2e7d32' : '') : '';
        var chgText = fv(s.漲跌幅, 2) + (ok(s.漲跌額) ? '<br><small>' + (s.漲跌額 > 0 ? '+' : '') + s.漲跌額.toFixed(2) + '</small>' : '');
        h += '<td style="color:' + chgColor + '">' + chgText + '</td>';
        h += '<td>' + (ok(s.強度評分) ? s.強度評分 : '') + '</td>';
        h += '<td>' + (ok(s.訊號數量) ? s.訊號數量 : '') + '</td>';
        h += '</tr>';
    }
    tbody.innerHTML = h;
}

// ==================== K 線圖 ====================
var chartInstance = null;

function openChart(code) {
    if (typeof LightweightCharts === 'undefined') { alert('需要網路連線才能顯示圖表'); return; }
    var stock = null;
    for (var i = 0; i < data.length; i++) {
        if (String(data[i].股票代號) === String(code)) { stock = data[i]; break; }
    }
    if (!stock) return;

    document.getElementById('chartTitle').textContent = code + ' ' + (stock.股票名稱 || '');
    document.getElementById('chartOverlay').className = 'chart-overlay show';
    var container = document.getElementById('chartContainer');
    container.innerHTML = '<div style="display:flex;align-items:center;justify-content:center;height:400px;color:#888">載入中...</div>';

    // 從快取或遠端載入 OHLC
    if (OHLC_CACHE[code]) {
        renderChart(code, OHLC_CACHE[code]);
    } else {
        fetch('ohlc/' + code + '.json').then(function(r) {
            if (!r.ok) throw new Error('no data');
            return r.json();
        }).then(function(ohlcData) {
            OHLC_CACHE[code] = ohlcData;
            renderChart(code, ohlcData);
        }).catch(function() {
            container.innerHTML = '<div style="display:flex;align-items:center;justify-content:center;height:400px;color:#888">此股票無歷史資料</div>';
        });
    }
}

function renderChart(code, ohlcData) {
    var container = document.getElementById('chartContainer');
    container.innerHTML = '';

    var candleData = [];
    var volumeData = [];
    for (var i = 0; i < ohlcData.length; i++) {
        var d = ohlcData[i];
        var dateStr = d.t.substring(0,4) + '-' + d.t.substring(4,6) + '-' + d.t.substring(6,8);
        candleData.push({ time: dateStr, open: d.o, high: d.h, low: d.l, close: d.c });
        volumeData.push({
            time: dateStr,
            value: d.v * 1000,
            color: d.c >= d.o ? 'rgba(239,83,80,0.5)' : 'rgba(38,166,154,0.5)'
        });
    }

    var chart = LightweightCharts.createChart(container, {
        width: container.clientWidth,
        height: 400,
        layout: { background: { type: 'solid', color: '#fff' }, textColor: '#333' },
        grid: { vertLines: { color: '#f0f0f0' }, horzLines: { color: '#f0f0f0' } },
        crosshair: { mode: LightweightCharts.CrosshairMode.Normal },
        rightPriceScale: { borderColor: '#ddd' },
        timeScale: { borderColor: '#ddd', timeVisible: false }
    });

    var candleSeries = chart.addCandlestickSeries({
        upColor: '#ef5350', downColor: '#26a69a',
        borderDownColor: '#26a69a', borderUpColor: '#ef5350',
        wickDownColor: '#26a69a', wickUpColor: '#ef5350'
    });
    candleSeries.setData(candleData);

    // 均線 MA5, MA10, MA20, MA60, MA240
    var maConfigs = [
        { n: 5, color: '#ff9800', width: 1 },
        { n: 10, color: '#2196f3', width: 1 },
        { n: 20, color: '#e91e63', width: 1 },
        { n: 60, color: '#9c27b0', width: 1.5 },
        { n: 240, color: '#4caf50', width: 1.5 }
    ];
    for (var m = 0; m < maConfigs.length; m++) {
        var period = maConfigs[m].n;
        if (candleData.length < period) continue;
        var maData = [];
        var sum = 0;
        for (var i = 0; i < candleData.length; i++) {
            sum += candleData[i].close;
            if (i >= period) sum -= candleData[i - period].close;
            if (i >= period - 1) {
                maData.push({ time: candleData[i].time, value: Math.round(sum / period * 100) / 100 });
            }
        }
        var maSeries = chart.addLineSeries({
            color: maConfigs[m].color,
            lineWidth: maConfigs[m].width,
            priceLineVisible: false,
            lastValueVisible: false,
            crosshairMarkerVisible: false
        });
        maSeries.setData(maData);
    }

    var volumeSeries = chart.addHistogramSeries({
        priceFormat: { type: 'volume' },
        priceScaleId: '',
        scaleMargins: { top: 0.8, bottom: 0 }
    });
    volumeSeries.setData(volumeData);

    chart.timeScale().fitContent();
    chartInstance = chart;

    window.addEventListener('resize', chartResize);
}

function chartResize() {
    if (chartInstance) {
        var c = document.getElementById('chartContainer');
        chartInstance.applyOptions({ width: c.clientWidth });
    }
}

function closeChart(e) {
    if (e && e.target !== document.getElementById('chartOverlay')) return;
    document.getElementById('chartOverlay').className = 'chart-overlay';
    if (chartInstance) { chartInstance.remove(); chartInstance = null; }
    window.removeEventListener('resize', chartResize);
}

if ('serviceWorker' in navigator) {
    navigator.serviceWorker.register('sw.js?v=__CACHE_VERSION__').catch(function(){});
}
</script>
</body>
</html>'''


# ==================== Service Worker ====================
SW_JS_TEMPLATE = '''var CACHE_NAME = '__CACHE_VERSION__';
var ASSETS = ['./index.html', './manifest.json', './icons/icon-144.png', './icons/icon-192.png', './icons/icon-512.png'];

self.addEventListener('install', function(e) {
    e.waitUntil(caches.open(CACHE_NAME).then(function(c) { return c.addAll(ASSETS); }));
    self.skipWaiting();
});

self.addEventListener('activate', function(e) {
    e.waitUntil(
        caches.keys().then(function(keys) {
            return Promise.all(keys.filter(function(k) { return k !== CACHE_NAME; }).map(function(k) { return caches.delete(k); }));
        })
    );
    self.clients.claim();
});

self.addEventListener('fetch', function(e) {
    if (e.request.mode === 'navigate') {
        e.respondWith(
            fetch(e.request).then(function(res) {
                var clone = res.clone();
                caches.open(CACHE_NAME).then(function(c) { c.put(e.request, clone); });
                return res;
            }).catch(function() { return caches.match(e.request); })
        );
    } else {
        e.respondWith(
            caches.match(e.request).then(function(cached) { return cached || fetch(e.request); })
        );
    }
});
'''


# ==================== Manifest ====================
MANIFEST = {
    "name": "台股篩選器",
    "short_name": "台股篩選",
    "start_url": "./index.html",
    "display": "standalone",
    "background_color": "#667eea",
    "theme_color": "#667eea",
    "icons": [
        {"src": "icons/icon-144.png", "sizes": "144x144", "type": "image/png", "purpose": "any"},
        {"src": "icons/icon-192.png", "sizes": "192x192", "type": "image/png", "purpose": "any"},
        {"src": "icons/icon-512.png", "sizes": "512x512", "type": "image/png", "purpose": "any"}
    ]
}


def main():
    print("=" * 60)
    print("  產生股票篩選 PWA")
    print("=" * 60)
    print()

    # 找最新 CSV
    csv_file = get_latest_csv()
    if not csv_file:
        print("找不到 stock_data_*.csv")
        exit(1)

    csv_files = get_sorted_csvs(10)
    csv_file = csv_files[0] if csv_files else None
    if not csv_file:
        print("找不到 stock_data_*.csv")
        exit(1)

    print(f"  CSV: {csv_file} (+{len(csv_files)-1} 歷史檔)")
    df = pd.read_csv(csv_file, encoding="utf-8-sig")
    print(f"  筆數: {len(df)}")

    # 用 stock_names_all.json 補齊中文名稱
    names_json = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stock_names_all.json")
    if os.path.exists(names_json):
        with open(names_json, 'r', encoding='utf-8') as f:
            all_names = json.load(f)
        fixed = 0
        for i, row in df.iterrows():
            code = str(int(row['股票代號'])) if isinstance(row['股票代號'], (int, float)) else str(row['股票代號'])
            if str(row['股票名稱']) == code and code in all_names:
                df.at[i, '股票名稱'] = all_names[code]
                fixed += 1
        if fixed:
            print(f"  名稱修正: {fixed} 筆")

    # 用 stock_market_type.json 加入市場類型
    mkt_json = os.path.join(os.path.dirname(os.path.abspath(__file__)), "stock_market_type.json")
    if os.path.exists(mkt_json):
        with open(mkt_json, 'r', encoding='utf-8') as f:
            mkt_map = json.load(f)
        df['市場'] = df['股票代號'].astype(str).map(lambda x: mkt_map.get(str(int(x)) if x == int(x) else x, ''))
        emg_count = (df['市場'] == '興櫃').sum()
        print(f"  市場分類: 興櫃 {emg_count} 筆")

    # 計算趨勢比對
    print("  計算趨勢比對...")
    df = compute_trend(df, csv_files)
    trend_count = df['漲跌幅'].notna().sum() if '漲跌幅' in df.columns else 0
    print(f"  趨勢: {trend_count} 筆有比對資料")

    # 計算突破訊號
    print("  計算突破訊號...")
    df = compute_breakout(df, csv_files)
    breakout_count = (df['強度評分'] > 0).sum() if '強度評分' in df.columns else 0
    print(f"  突破: {breakout_count} 筆有訊號")

    # 輸出突破訊號 Excel
    if '強度評分' in df.columns:
        df_breakout = df[df['強度評分'] > 0].copy()
        if len(df_breakout) > 0:
            date_str = csv_file.replace('stock_data_', '').replace('.csv', '')
            cols = ['股票代號', '股票名稱', '收盤價', '強度評分', '訊號數量', '訊號',
                    'K值', 'D值', 'RSI', '量比', '成交量張', '市值億', '殖利率']
            cols = [c for c in cols if c in df_breakout.columns]
            if '市場' in df_breakout.columns:
                cols.append('市場')
            df_breakout = df_breakout[cols].sort_values('強度評分', ascending=False)
            out_file = f"breakout_signals_{date_str}.xlsx"
            df_breakout.to_excel(out_file, index=False, sheet_name='突破訊號')
            # 加入表格格式（自動篩選 + 斑馬紋 + 欄寬調整 + 凍結首列）
            from openpyxl import load_workbook
            from openpyxl.worksheet.table import Table, TableStyleInfo
            wb = load_workbook(out_file)
            ws = wb['突破訊號']
            last_col = chr(64 + ws.max_column)
            table = Table(displayName='BreakoutSignals', ref=f'A1:{last_col}{ws.max_row}')
            table.tableStyleInfo = TableStyleInfo(
                name='TableStyleMedium9', showFirstColumn=False,
                showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)
            for col_cells in ws.columns:
                max_len = max(len(str(c.value or '')) for c in col_cells)
                ws.column_dimensions[col_cells[0].column_letter].width = min(max_len + 4, 30)
            ws.freeze_panes = 'A2'
            wb.save(out_file)
            print(f"  突破訊號: {out_file} ({len(df_breakout)} 筆)")

    # 取得實際交易日（從 CSV 的「交易日」欄位）
    current_trade_date = ''
    if '交易日' in df.columns and len(df) > 0:
        trade_dates = df['交易日'].dropna()
        if len(trade_dates) > 0:
            current_trade_date = str(trade_dates.mode().iloc[0]).replace('-', '')[:8]
    # （current_trade_date 在歷史收盤價建立後再做 fallback）

    # 歷史收盤價（供前端切換比較基準）
    print("  建立歷史收盤價...")
    hist_prices = {}
    default_compare = ''
    for f in csv_files[1:]:
        try:
            tmp = pd.read_csv(f, encoding="utf-8-sig")
            if len(tmp) < 1:
                continue
            date_str = f.replace('stock_data_', '').replace('.csv', '')
            if not default_compare:
                default_compare = date_str
            prices = {}
            for _, row in tmp.iterrows():
                code = str(int(row['股票代號'])) if isinstance(row['股票代號'], (int, float)) and not pd.isna(row['股票代號']) else str(row['股票代號'])
                if pd.notna(row.get('收盤價')):
                    prices[code] = round(float(row['收盤價']), 2)
            hist_prices[date_str] = prices
        except Exception:
            continue
    hist_json = json.dumps(hist_prices, ensure_ascii=False)
    print(f"  歷史: {len(hist_prices)} 個日期, {len(hist_json)/1024:.1f} KB")

    # fallback: 若沒有「交易日」欄位，用 HIST 推算
    if not current_trade_date:
        from datetime import datetime as dt, timedelta
        csv_date_str = csv_file.replace('stock_data_', '').replace('.csv', '')
        if default_compare:
            # 最近歷史日的下一個工作日就是目前的交易日
            try:
                last_hist = dt.strptime(default_compare, '%Y%m%d')
                candidate = last_hist + timedelta(days=1)
                csv_date = dt.strptime(csv_date_str, '%Y%m%d')
                while candidate.weekday() >= 5:
                    candidate += timedelta(days=1)
                if candidate <= csv_date:
                    current_trade_date = candidate.strftime('%Y%m%d')
                else:
                    current_trade_date = csv_date_str
            except ValueError:
                current_trade_date = csv_date_str
        else:
            current_trade_date = csv_date_str
    print(f"  交易日: {current_trade_date}")

    # 把目前的收盤價也加入 HIST（用實際交易日為 key）
    if current_trade_date and current_trade_date not in hist_prices:
        current_prices = {}
        for _, row in df.iterrows():
            code = str(int(row['股票代號'])) if isinstance(row['股票代號'], (int, float)) and not pd.isna(row['股票代號']) else str(row['股票代號'])
            if pd.notna(row.get('收盤價')):
                current_prices[code] = round(float(row['收盤價']), 2)
        hist_prices[current_trade_date] = current_prices
        hist_json = json.dumps(hist_prices, ensure_ascii=False)
        print(f"  已加入 {current_trade_date} 至歷史 ({len(current_prices)} 筆), 共 {len(hist_prices)} 個日期")

    # OHLC 歷史：由 download_ohlc.py 預先產生 pwa/ohlc/*.json，前端按需載入
    ohlc_dir = os.path.join('pwa', 'ohlc')
    ohlc_count = len(os.listdir(ohlc_dir)) if os.path.isdir(ohlc_dir) else 0
    print(f"  OHLC: {ohlc_count} 支股票（pwa/ohlc/）")

    # 更新時間
    if '更新時間' in df.columns and len(df) > 0:
        update_time = str(df['更新時間'].iloc[0])
    else:
        update_time = datetime.now().strftime('%Y-%m-%d %H:%M')

    # 選取欄位
    desired = [
        '股票代號', '股票名稱', '收盤價',
        'K值', 'D值', 'RSI', 'DIF', 'MACD',
        'BB上軌', 'BB下軌',
        '成交量張', '量比', '市值億', '本益比', '每股配息', '殖利率',
        # 趨勢欄位
        '漲跌幅', '價格趨勢', 'KD趨勢', 'RSI趨勢', '量能趨勢', 'K值變化', 'RSI變化',
        # 突破欄位
        '強度評分', '訊號數量', '訊號',
        # 市場分類
        '市場'
    ]
    cols = [c for c in desired if c in df.columns]
    df_out = df[cols].copy()

    # 數值欄位四捨五入
    for col in df_out.select_dtypes(include='number').columns:
        df_out[col] = df_out[col].round(2)

    json_data = df_out.to_json(orient='records', force_ascii=False)
    print(f"  JSON: {len(json_data)/1024:.1f} KB")

    # 建目錄
    os.makedirs('pwa/icons', exist_ok=True)

    # 每次 build 用新版號，強制清舊快取
    cache_version = 'stock-viewer-' + datetime.now().strftime('%Y%m%d%H%M%S')

    # 產出 index.html
    html = HTML_TEMPLATE
    html = html.replace('__STOCK_DATA__', json_data)
    html = html.replace('__UPDATE_TIME__', update_time)
    html = html.replace('__TOTAL_COUNT__', str(len(df)))
    html = html.replace('__HIST_DATA__', hist_json)
    html = html.replace('__COMPARE_DEFAULT__', default_compare)
    html = html.replace('__CURRENT_TRADE_DATE__', current_trade_date)
    html = html.replace('__CACHE_VERSION__', cache_version)

    with open('pwa/index.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  index.html: {len(html):,} bytes")

    # 產出 manifest.json
    with open('pwa/manifest.json', 'w', encoding='utf-8') as f:
        json.dump(MANIFEST, f, ensure_ascii=False, indent=2)
    print("  manifest.json: OK")
    sw_content = SW_JS_TEMPLATE.replace('__CACHE_VERSION__', cache_version)
    with open('pwa/sw.js', 'w', encoding='utf-8') as f:
        f.write(sw_content)
    print(f"  sw.js: OK (cache: {cache_version})")

    # 產出 _headers（禁止 CDN 快取 sw.js）
    with open('pwa/_headers', 'w', encoding='utf-8') as f:
        f.write("/sw.js\n  Cache-Control: no-cache, no-store, must-revalidate\n")
    print("  _headers: OK")

    # 產出 icons
    generate_icons('pwa/icons')

    print()
    print("PWA 產出完成！")
    print(f"  目錄: pwa/")
    print(f"  預覽: 開啟 pwa/index.html")
    print("=" * 60)


if __name__ == '__main__':
    main()
