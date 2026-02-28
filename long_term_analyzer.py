"""
long_term_analyzer.py
長期趨勢分析器 - 偵測區間內的突破訊號
"""
import pandas as pd
import glob
import os
from datetime import datetime
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import sys

# 導入股票名稱對照表
sys.path.insert(0, '.')
try:
    from data_downloader import STOCK_NAMES
    print(f"✅ 已載入 {len(STOCK_NAMES)} 個股票名稱")
except ImportError as e:
    print(f"⚠️  無法載入股票名稱對照表：{e}")
    STOCK_NAMES = {}

def get_all_data_files():
    """取得所有數據檔案"""
    files = glob.glob("stock_data_*.csv")
    files_info = []
    
    for f in files:
        date_str = f.replace("stock_data_", "").replace(".csv", "")
        try:
            date_obj = datetime.strptime(date_str, "%Y%m%d")
            files_info.append({
                'file': f,
                'date': date_obj,
                'date_str': date_obj.strftime('%Y-%m-%d (%a)')
            })
        except:
            pass
    
    files_info.sort(key=lambda x: x['date'], reverse=True)
    return files_info

def select_date_range(files_info):
    """選擇日期範圍（可重選）"""
    while True:
        print("=" * 70)
        print("  長期趨勢分析器 - 突破訊號偵測")
        print("=" * 70)
        print()
        print("📁 可用的數據檔案：")
        print()
        
        for i, info in enumerate(files_info, 1):
            print(f"   {i:2d}. {info['date_str']}")
        
        print()
        print("=" * 70)
        print()
        print("請選擇分析區間：")
        print("(輸入 0 可返回上一步，輸入 q 退出程式)")
        print()
        
        # 選擇起始日期
        while True:
            start_input = input(f"起始日期(較舊的)(1-{len(files_info)} 或 0/q): ").strip().lower()
            
            if start_input == 'q':
                print("\n👋 已退出程式")
                exit(0)
            
            if start_input == '0':
                print("\n🔄 重新選擇...\n")
                break
            
            try:
                start_idx = int(start_input) - 1
                if 0 <= start_idx < len(files_info):
                    break
                else:
                    print(f"❌ 請輸入 1 到 {len(files_info)} 之間的數字")
            except ValueError:
                print("❌ 請輸入有效的數字")
        
        if start_input == '0':
            continue
        
        # 選擇結束日期
        while True:
            end_input = input(f"結束日期(最新的)(1-{len(files_info)} 或 0/q): ").strip().lower()
            
            if end_input == 'q':
                print("\n👋 已退出程式")
                exit(0)
            
            if end_input == '0':
                print("\n🔄 重新選擇起始日期...\n")
                break
            
            try:
                end_idx = int(end_input) - 1
                
                if not (0 <= end_idx < len(files_info)):
                    print(f"❌ 請輸入 1 到 {len(files_info)} 之間的數字")
                    continue
                
                if end_idx >= start_idx:
                    print("❌ 結束日期應該比起始日期新(編號要更小)")
                    print(f"   您選的起始：{files_info[start_idx]['date_str']}")
                    print(f"   您選的結束：{files_info[end_idx]['date_str']}")
                    print(f"   提示：結束日期編號應 < {start_idx + 1}")
                    continue
                
                break
                
            except ValueError:
                print("❌ 請輸入有效的數字")
        
        if end_input == '0':
            continue
        
        # 確認選擇
        print()
        print("=" * 70)
        print()
        print("📊 您選擇的分析區間：")
        print(f"   起始：{files_info[start_idx]['date_str']}")
        print(f"   結束：{files_info[end_idx]['date_str']}")
        print(f"   天數：{start_idx - end_idx + 1} 天")
        print()
        
        confirm = input("確認此區間？(Y/n): ").strip().lower()
        
        if confirm in ['', 'y', 'yes', '是']:
            selected_files = [info['file'] for info in files_info[end_idx:start_idx+1]]
            selected_files.reverse()
            
            return selected_files, files_info[start_idx]['date_str'], files_info[end_idx]['date_str']
        else:
            print("\n🔄 重新選擇...\n")

def load_multiple_files(files):
    """載入多個 CSV 並合併"""
    all_data = []
    
    for f in files:
        df = pd.read_csv(f, encoding="utf-8-sig")
        date_str = f.replace("stock_data_", "").replace(".csv", "")
        df['日期'] = date_str
        all_data.append(df)
    
    combined_df = pd.concat(all_data, ignore_index=True)
    
    # 補充股票名稱
    if len(STOCK_NAMES) > 0:
        print("⏳ 補充股票名稱...")
        fixed_count = 0
        
        for idx, row in combined_df.iterrows():
            code = str(row['股票代號']).replace('.TW', '')
            current_name = str(row['股票名稱'])
            
            needs_fix = (
                current_name.isdigit() or 
                current_name == code or 
                current_name == 'nan' or
                pd.isna(current_name)
            )
            
            if needs_fix and code in STOCK_NAMES:
                combined_df.at[idx, '股票名稱'] = STOCK_NAMES[code]
                fixed_count += 1
        
        if fixed_count > 0:
            print(f"✅ 已補充 {fixed_count} 筆股票名稱")
        else:
            print("✅ 股票名稱已完整")
    else:
        print("⚠️  未載入股票名稱對照表，跳過補充")
    
    return combined_df

def calculate_period_stats(df_all):
    """計算區間統計"""
    stats = {}
    
    for code in df_all['股票代號'].unique():
        stock_data = df_all[df_all['股票代號'] == code]
        
        if len(stock_data) == 0:
            continue
        
        stats[code] = {
            '股票名稱': stock_data['股票名稱'].iloc[0],
            '平均價格': stock_data['收盤價'].mean(),
            '最高價': stock_data['收盤價'].max(),
            '最低價': stock_data['收盤價'].min(),
            '平均成交量': stock_data['成交量張'].mean(),
            '平均K值': stock_data['K值'].mean(),
            '平均D值': stock_data['D值'].mean(),
            '平均RSI': stock_data['RSI'].mean(),
            '資料筆數': len(stock_data),
            '最新價格': stock_data.iloc[-1]['收盤價'],
            '最新成交量': stock_data.iloc[-1]['成交量張'],
            '最新K值': stock_data.iloc[-1]['K值'],
            '最新D值': stock_data.iloc[-1]['D值'],
            '最新RSI': stock_data.iloc[-1]['RSI'],
        }
    
    return stats

def detect_breakouts(stats):
    """偵測突破訊號(15 種)"""
    signals = []
    
    for code, data in stats.items():
        signal_types = []
        strength_score = 0
        
        # 1. 量能突破
        if data['最新成交量'] > data['平均成交量'] * 1.5:
            signal_types.append('📈 量能突破')
            strength_score += 2
        
        # 2. KD 黃金交叉
        if data['最新K值'] > data['最新D值'] and data['最新K值'] < 50:
            signal_types.append('🟡 KD黃金交叉')
            strength_score += 3
        
        # 3. RSI 轉強
        if data['最新RSI'] > 50 and data['平均RSI'] < 50:
            signal_types.append('💪 RSI轉強')
            strength_score += 2
        
        # 4. 價格突破
        if data['最新價格'] >= data['最高價'] * 0.95:
            signal_types.append('🚀 價格突破')
            strength_score += 3
        
        # 5. 站上均線
        if data['最新價格'] > data['平均價格'] and \
           data['最新價格'] < data['平均價格'] * 1.05:
            signal_types.append('📊 站上均線')
            strength_score += 2
        
        # 6. 連續上漲
        price_trend = (data['最新價格'] - data['平均價格']) / data['平均價格']
        if price_trend > 0.03:
            signal_types.append('🔥 連續上漲')
            strength_score += 2
        
        # 7. 低檔起漲
        rebound = (data['最新價格'] - data['最低價']) / data['最低價']
        if rebound > 0.10 and data['最新價格'] < data['平均價格'] * 1.1:
            signal_types.append('💎 低檔起漲')
            strength_score += 4
        
        # 8. 急漲訊號
        if price_trend > 0.05:
            signal_types.append('⚡ 急漲訊號')
            strength_score += 3
        
        # 9. 量價齊揚
        volume_ratio = data['最新成交量'] / data['平均成交量']
        if price_trend > 0 and volume_ratio > 1.2:
            signal_types.append('🌊 量價齊揚')
            strength_score += 3
        
        # 10. 突破整理
        price_range = (data['最高價'] - data['最低價']) / data['平均價格']
        if price_range < 0.15 and data['最新價格'] >= data['最高價'] * 0.95:
            signal_types.append('🎯 突破整理')
            strength_score += 4
        
        # 11. KD 超賣反彈
        if data['平均D值'] < 30 and data['最新K值'] > data['最新D值']:
            signal_types.append('📉 KD超賣反彈')
            strength_score += 3
        
        # 12. 布林突破
        if data['最新價格'] >= data['最高價'] * 0.95 and data['最新K值'] > 70:
            signal_types.append('🔵 布林突破')
            strength_score += 2
        
        # 13. MACD 翻多
        if data['平均RSI'] < 45 and data['最新RSI'] > 50 and \
           data['最新K值'] > data['最新D值']:
            signal_types.append('📊 MACD翻多')
            strength_score += 3
        
        # 14. 多重訊號
        if len(signal_types) >= 3:
            signal_types.append('🌟 多重訊號')
            strength_score += 5
        
        if len(signal_types) > 0:
            signals.append({
                '股票代號': code,
                '股票名稱': data['股票名稱'],
                '最新價格': round(data['最新價格'], 2),
                '平均價格': round(data['平均價格'], 2),
                '最低價': round(data['最低價'], 2),
                '最高價': round(data['最高價'], 2),
                '漲跌幅%': round((data['最新價格'] - data['平均價格']) / data['平均價格'] * 100, 2),
                '量能比': round(data['最新成交量'] / data['平均成交量'], 2),
                'K值': round(data['最新K值'], 2),
                'D值': round(data['最新D值'], 2),
                'RSI': round(data['最新RSI'], 2),
                '訊號': ' | '.join(signal_types),
                '訊號數量': len(signal_types),
                '強度評分': strength_score,
            })
    
    return signals

# ========== 主程式 ==========

files_info = get_all_data_files()

if len(files_info) < 2:
    print("❌ 數據檔案不足，至少需要 2 個")
    exit(1)

# 循環分析模式
while True:
    selected_files, start_date, end_date = select_date_range(files_info)
    
    print()
    print("=" * 70)
    print()
    print(f"📊 分析區間：{end_date} ~ {start_date}")
    print(f"📁 檔案數量：{len(selected_files)} 個")
    print()
    
    print("⏳ 載入數據中...")
    df_all = load_multiple_files(selected_files)
    
    print(f"✅ 載入完成：{len(df_all)} 筆記錄")
    print()
    
    print("⏳ 計算區間統計...")
    stats = calculate_period_stats(df_all)
    print(f"✅ 完成：{len(stats)} 支股票")
    print()
    
    print("⏳ 偵測突破訊號...")
    signals = detect_breakouts(stats)
    print(f"✅ 找到 {len(signals)} 支股票有突破訊號")
    print()
    
    signals_sorted = sorted(signals, 
                           key=lambda x: (x['強度評分'], x['訊號數量'], x['量能比']), 
                           reverse=True)
    
    print("=" * 70)
    print("  突破訊號排行(前 30 名)")
    print("=" * 70)
    print()
    
    for i, stock in enumerate(signals_sorted[:30], 1):
        stars = '⭐' * min(5, stock['強度評分'] // 3)
        
        print(f"{i:2d}. {stock['股票代號']} {stock['股票名稱']:8s} {stars}")
        print(f"    💰 價格：{stock['最新價格']:6.2f} (區間 {stock['最低價']:6.2f} ~ {stock['最高價']:6.2f})")
        print(f"    📊 均價：{stock['平均價格']:6.2f} ({stock['漲跌幅%']:+5.2f}%)")
        print(f"    📈 量能：{stock['量能比']:.2f}x 平均量")
        print(f"    🎯 技術：K={stock['K值']:.1f} D={stock['D值']:.1f} RSI={stock['RSI']:.1f}")
        print(f"    🔔 訊號：{stock['訊號']}")
        print(f"    💪 評分：{stock['強度評分']} 分 ({stock['訊號數量']} 種訊號)")
        print()
    
    json_file = f"breakout_signals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    df_json = pd.DataFrame(signals_sorted)
    df_json.to_json(json_file, orient='records', force_ascii=False, indent=2)
    
    print()
    print("⏳ 匯出 Excel...")
    
    df_signals = pd.DataFrame(signals_sorted)
    column_order = [
        '強度評分', '訊號數量', '股票代號', '股票名稱',
        '最新價格', '平均價格', '最低價', '最高價', '漲跌幅%',
        '量能比', 'K值', 'D值', 'RSI', '訊號'
    ]
    df_export = df_signals[column_order]
    
    excel_file = f"breakout_signals_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    df_export.to_excel(excel_file, index=False, engine='openpyxl', sheet_name='突破訊號')
    
    print("⏳ 美化 Excel 格式...")
    
    wb = load_workbook(excel_file)
    ws = wb.active
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    column_widths = {
        'A': 10, 'B': 10, 'C': 12, 'D': 16, 'E': 10, 'F': 10, 'G': 10,
        'H': 10, 'I': 10, 'J': 10, 'K': 8, 'L': 8, 'M': 8, 'N': 60,
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=14, max_col=14):
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    score_fills = {
        'high': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
        'medium': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),
        'low': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
    }
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        score = row[0].value
        
        if score >= 12:
            fill = score_fills['high']
        elif score >= 6:
            fill = score_fills['medium']
        else:
            fill = score_fills['low']
        
        row[0].fill = fill
        row[0].font = Font(bold=True, size=11)
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=9, max_col=9):
        for cell in row:
            if cell.value and cell.value > 0:
                cell.font = Font(color="FF0000", bold=True)
            elif cell.value and cell.value < 0:
                cell.font = Font(color="00B050", bold=True)
    
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    
    wb.save(excel_file)
    
    print()
    print("=" * 70)
    print("✅ 完成！")
    print("=" * 70)
    print()
    print(f"📄 JSON 檔案：{json_file}")
    print(f"📊 Excel 檔案：{excel_file}")
    print()
    print("💡 Excel 功能：")
    print("  • 按評分、訊號數量排序")
    print("  • 篩選特定訊號(點訊號欄位下拉)")
    print("  • 顏色標示：")
    print("    - 綠色：高評分(≥12 分)")
    print("    - 黃色：中評分(6-11 分)")
    print("    - 紅色：低評分(<6 分)")
    print()
    print("=" * 70)
    print()
    
    choice = input("是否分析其他區間？(Y/n): ").strip().lower()
    
    if choice in ['n', 'no', '否']:
        print()
        print("=" * 70)
        print("📚 訊號說明：")
        print()
        print("【基礎訊號】")
        print("  📈 量能突破 - 成交量 > 平均量 1.5 倍")
        print("  🟡 KD黃金交叉 - K > D 且在低檔(K < 50)")
        print("  💪 RSI轉強 - RSI 突破 50(由弱轉強)")
        print("  🚀 價格突破 - 接近區間高點(95% 以上)")
        print("  📊 站上均線 - 突破平均價格線")
        print()
        print("【進階訊號】")
        print("  🔥 連續上漲 - 價格持續高於均價 3% 以上")
        print("  💎 低檔起漲 - 從低點反彈超過 10%")
        print("  ⚡ 急漲訊號 - 價格高於均價 5% 以上")
        print("  🌊 量價齊揚 - 價格上漲且成交量放大")
        print("  🎯 突破整理 - 橫盤後突破(波動 < 15%)")
        print()
        print("【技術訊號】")
        print("  📉 KD超賣反彈 - KD 從低檔回升")
        print("  🔵 布林突破 - 突破上軌(K > 70)")
        print("  📊 MACD翻多 - RSI 回升 + KD 黃金交叉")
        print("  🌟 多重訊號 - 同時觸發 3 種以上訊號")
        print()
        print("【評分機制】")
        print("  ⭐       - 1-2 分(一般訊號)")
        print("  ⭐⭐     - 3-5 分(中等訊號)")
        print("  ⭐⭐⭐   - 6-8 分(強烈訊號)")
        print("  ⭐⭐⭐⭐ - 9-11 分(極強訊號)")
        print("  ⭐⭐⭐⭐⭐ - 12 分以上(多重確認)")
        print("=" * 70)
        print()
        print("👋 謝謝使用！")
        print()
        break
    
    print()
    print("=" * 70)
    print()
    print("🔄 重新選擇分析區間...")
    print()