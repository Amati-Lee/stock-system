"""
quick_filter.py - 快速篩選股票（支援單選/複選/組合）
"""
import pandas as pd
import glob
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

def get_latest_data_file():
    files = glob.glob("stock_data_*.csv")
    return max(files) if files else None

def load_data(filename):
    try:
        df = pd.read_csv(filename, encoding="utf-8-sig")
        print(f"✅ 載入數據：{filename}")
        print(f"   總筆數：{len(df)}")
        return df
    except Exception as e:
        print(f"❌ 載入失敗：{e}")
        return None

def show_menu():
    print("\n" + "=" * 70)
    print("  快速篩選器")
    print("=" * 70 + "\n")
    print("【長線投資】")
    print("  1. 高殖利率股")
    print("  2. 價值投資\n")
    print("【短線操作】")
    print("  3. KD 黃金交叉")
    print("  4. 布林通道上軌")
    print("  5. MACD 黃金交叉\n")
    print("【超賣反彈】")
    print("  6. KD 超賣")
    print("  7. RSI 超賣")
    print("  8. 布林下軌\n")
    print("【綜合策略】")
    print("  9. 五指標共振")
    print(" 10. 短線雙確認\n")
    print(" 11. 自訂條件（支援單選/複選/組合）\n")
    print("=" * 70)

def custom_filter(df):
    print("\n" + "=" * 70)
    print("  自訂條件（可單選或複選組合）")
    print("=" * 70 + "\n")
    
    mode = input("模式 [1=單一條件(AND), 2=複選組合(OR)]: ").strip() or "1"
    
    if mode == "2":
        return multi_select_mode(df)
    else:
        return single_condition_mode(df)

def single_condition_mode(df):
    print("\n單一條件模式（所有條件必須同時符合）\n")
    result = df.copy()
    
    # 價格
    if input("篩選價格？ (y/n) [n]: ").lower() == 'y':
        min_p = float(input("  最低（元）[0]: ") or 0)
        max_p = float(input("  最高（元）[9999]: ") or 9999)
        result = result[(result["收盤價"] >= min_p) & (result["收盤價"] <= max_p)]
        print(f"  ✅ 剩餘 {len(result)} 筆")
    
    # 殖利率
    if input("篩選殖利率？ (y/n) [n]: ").lower() == 'y':
        min_d = float(input("  最低（%）[5]: ") or 5)
        result = result[result["殖利率"] >= min_d]
        print(f"  ✅ 剩餘 {len(result)} 筆")
    
    # KD
    if input("篩選 KD？ (y/n) [n]: ").lower() == 'y':
        kd_type = input("  [1=超賣, 2=黃金交叉]: ") or "1"
        if kd_type == "1":
            max_k = float(input("    K< [20]: ") or 20)
            max_d = float(input("    D< [20]: ") or 20)
            result = result[(result["K值"] < max_k) & (result["D值"] < max_d)]
        else:
            result = result[result["KD狀態"].str.contains("黃金交叉|K>D", na=False)]
        print(f"  ✅ 剩餘 {len(result)} 筆")
    
    # RSI
    if input("篩選 RSI？ (y/n) [n]: ").lower() == 'y':
        max_r = float(input("  上限 [30]: ") or 30)
        result = result[result["RSI"] < max_r]
        print(f"  ✅ 剩餘 {len(result)} 筆")
    
    # 布林通道
    if input("篩選布林通道？ (y/n) [n]: ").lower() == 'y':
        bb_type = input("  [1=下軌, 2=上軌]: ") or "1"
        if bb_type == "1":
            result = result[result["BB位置"] <= 30]
        else:
            result = result[result["BB位置"] >= 70]
        print(f"  ✅ 剩餘 {len(result)} 筆")
    
    # 量能
    if input("篩選量能？ (y/n) [n]: ").lower() == 'y':
        min_v = float(input("  最低量比 [1.5]: ") or 1.5)
        result = result[result["量比"] >= min_v]
        print(f"  ✅ 剩餘 {len(result)} 筆")
    
    # 市值
    if input("篩選市值？ (y/n) [n]: ").lower() == 'y':
        min_c = float(input("  最低（億）[10]: ") or 10)
        max_c = float(input("  最高（億）[9999]: ") or 9999)
        result = result[(result["市值億"] >= min_c) & (result["市值億"] <= max_c)]
        print(f"  ✅ 剩餘 {len(result)} 筆")
    
    # 本益比
    if input("篩選本益比？ (y/n) [n]: ").lower() == 'y':
        min_pe = float(input("  最低 [5]: ") or 5)
        max_pe = float(input("  最高 [30]: ") or 30)
        result = result[(result["本益比"] >= min_pe) & (result["本益比"] <= max_pe)]
        print(f"  ✅ 剩餘 {len(result)} 筆")
    
    return result, "自訂條件(AND)"

def multi_select_mode(df):
    print("\n複選模式（符合任一條件即可）\n")
    
    result_sets = []
    
    # 長線投資
    print("【長線投資】")
    choices = input("  選擇（可多選，逗號分隔）[1=高殖利率, 2=價值投資]: ").strip()
    if choices:
        for c in choices.split(','):
            c = c.strip()
            if c == "1":
                min_d = float(input("    殖利率>= [6]: ") or 6)
                subset = df[df["殖利率"] >= min_d].copy()
                result_sets.append((f"高殖利率>={min_d}%", subset))
                print(f"    ✅ {len(subset)} 筆")
            elif c == "2":
                min_d = float(input("    殖利率>= [5]: ") or 5)
                max_pe = float(input("    本益比<= [20]: ") or 20)
                subset = df[(df["殖利率"] >= min_d) & (df["本益比"] <= max_pe)].copy()
                result_sets.append(("價值投資", subset))
                print(f"    ✅ {len(subset)} 筆")
    
    # 短線操作
    print("\n【短線操作】")
    choices = input("  選擇（可多選）[3=KD黃金交叉, 4=布林上軌, 5=MACD黃金交叉]: ").strip()
    if choices:
        for c in choices.split(','):
            c = c.strip()
            if c == "3":
                subset = df[df["KD狀態"].str.contains("黃金交叉|K>D", na=False)].copy()
                result_sets.append(("KD黃金交叉", subset))
                print(f"    ✅ {len(subset)} 筆")
            elif c == "4":
                subset = df[df["BB位置"] >= 70].copy()
                result_sets.append(("布林上軌", subset))
                print(f"    ✅ {len(subset)} 筆")
            elif c == "5":
                subset = df[df["MACD狀態"].str.contains("黃金交叉|DIF>MACD", na=False)].copy()
                result_sets.append(("MACD黃金交叉", subset))
                print(f"    ✅ {len(subset)} 筆")
    
    # 超賣反彈
    print("\n【超賣反彈】")
    choices = input("  選擇（可多選）[6=KD超賣, 7=RSI超賣, 8=布林下軌]: ").strip()
    if choices:
        for c in choices.split(','):
            c = c.strip()
            if c == "6":
                max_k = float(input("    K< [20]: ") or 20)
                max_d = float(input("    D< [20]: ") or 20)
                subset = df[(df["K值"] < max_k) & (df["D值"] < max_d)].copy()
                result_sets.append(("KD超賣", subset))
                print(f"    ✅ {len(subset)} 筆")
            elif c == "7":
                max_r = float(input("    RSI< [30]: ") or 30)
                subset = df[df["RSI"] < max_r].copy()
                result_sets.append(("RSI超賣", subset))
                print(f"    ✅ {len(subset)} 筆")
            elif c == "8":
                subset = df[df["BB位置"] <= 30].copy()
                result_sets.append(("布林下軌", subset))
                print(f"    ✅ {len(subset)} 筆")
    
    # 合併
    if result_sets:
        combined = pd.concat([s for _, s in result_sets]).drop_duplicates()
        name = "+".join([n for n, _ in result_sets])
        print(f"\n✅ 合併結果：{len(combined)} 筆")
        
        # 額外條件
        if input("\n加入額外條件（價格/市值/量比）？ (y/n) [n]: ").lower() == 'y':
            combined = apply_extra_filters(combined)
        
        return combined, name
    else:
        print("\n❌ 未選擇任何條件")
        return None, None

def apply_extra_filters(df):
    result = df.copy()
    
    if input("  價格範圍？ (y/n) [n]: ").lower() == 'y':
        min_p = float(input("    最低 [0]: ") or 0)
        max_p = float(input("    最高 [9999]: ") or 9999)
        result = result[(result["收盤價"] >= min_p) & (result["收盤價"] <= max_p)]
        print(f"    ✅ 剩餘 {len(result)} 筆")
    
    if input("  市值範圍？ (y/n) [n]: ").lower() == 'y':
        min_c = float(input("    最低（億）[10]: ") or 10)
        max_c = float(input("    最高（億）[9999]: ") or 9999)
        result = result[(result["市值億"] >= min_c) & (result["市值億"] <= max_c)]
        print(f"    ✅ 剩餘 {len(result)} 筆")
    
    if input("  量比？ (y/n) [n]: ").lower() == 'y':
        min_v = float(input("    最低 [1.5]: ") or 1.5)
        result = result[result["量比"] >= min_v]
        print(f"    ✅ 剩餘 {len(result)} 筆")
    
    return result

def apply_preset_strategy(df, strategy):
    result = df.copy()
    
    if strategy == "1":
        min_d = float(input("  最低殖利率（%）[6]: ") or 6)
        result = result[result["殖利率"] >= min_d]
        return result, f"高殖利率>={min_d}%"
    
    elif strategy == "2":
        min_d = float(input("  最低殖利率（%）[5]: ") or 5)
        max_pe = float(input("  最高本益比 [20]: ") or 20)
        result = result[(result["殖利率"] >= min_d) & (result["本益比"] <= max_pe)]
        return result, "價值投資"
    
    elif strategy == "3":
        result = result[result["KD狀態"].str.contains("黃金交叉|K>D", na=False)]
        return result, "KD黃金交叉"
    
    elif strategy == "4":
        result = result[result["BB位置"] >= 70]
        return result, "布林上軌"
    
    elif strategy == "5":
        result = result[result["MACD狀態"].str.contains("黃金交叉|DIF>MACD", na=False)]
        return result, "MACD黃金交叉"
    
    elif strategy == "6":
        max_k = float(input("  K值上限 [20]: ") or 20)
        max_d = float(input("  D值上限 [20]: ") or 20)
        result = result[(result["K值"] < max_k) & (result["D值"] < max_d)]
        return result, "KD超賣"
    
    elif strategy == "7":
        max_r = float(input("  RSI上限 [30]: ") or 30)
        result = result[result["RSI"] < max_r]
        return result, "RSI超賣"
    
    elif strategy == "8":
        result = result[result["BB位置"] <= 30]
        return result, "布林下軌"
    
    elif strategy == "9":
        result = result[
            (result["K值"] < 20) & (result["D值"] < 20) &
            (result["RSI"] < 30) & (result["BB位置"] <= 30) &
            (result["量比"] >= 1.5) &
            (result["MACD狀態"].str.contains("黃金交叉", na=False))
        ]
        return result, "五指標共振"
    
    elif strategy == "10":
        result = result[
            (result["KD狀態"].str.contains("黃金交叉", na=False)) &
            (result["BB位置"] >= 70)
        ]
        return result, "短線雙確認"
    
    else:
        return None, None

def export_result(df, strategy_name):
    if df is None or len(df) == 0:
        print("\n❌ 無符合條件的股票")
        return
    
    print(f"\n✅ 篩選完成：{len(df)} 筆")
    
    # 排序
    print("\n排序：[1=殖利率↓, 2=本益比↑, 3=市值↓, 4=量比↓, 5=K值↑, 6=不排序]")
    sort = input("選擇 [1]: ").strip() or "1"
    
    if sort == "1": df = df.sort_values("殖利率", ascending=False)
    elif sort == "2": df = df.sort_values("本益比", ascending=True)
    elif sort == "3": df = df.sort_values("市值億", ascending=False)
    elif sort == "4": df = df.sort_values("量比", ascending=False)
    elif sort == "5": df = df.sort_values("K值", ascending=True)
    
    # 顯示前 20 筆
    print("\n前 20 筆：")
    cols = ["股票代號", "股票名稱", "收盤價", "K值", "D值", "RSI", "BB位置", "殖利率", "本益比", "量比"]
    print(df[[c for c in cols if c in df.columns]].head(20).to_string(index=False))
    
    # 儲存
    if input("\n儲存 Excel？ (y/n) [y]: ").strip().lower() != 'n':
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        fn = f"filter_{strategy_name.replace(' ', '_').replace('+', '_')}_{ts}.xlsx"
        df.to_excel(fn, index=False, engine='openpyxl')
        
        # 格式化
        wb = load_workbook(fn)
        ws = wb.active
        
        # 標題
        hf = Font(bold=True, color="FFFFFF", size=11)
        hp = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ha = Alignment(horizontal="center", vertical="center")
        for cell in ws[1]:
            cell.font = hf
            cell.fill = hp
            cell.alignment = ha
        
        # 欄寬（+5，最大35）
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    val = str(cell.value)
                    length = sum(2 if '\u4e00' <= c <= '\u9fff' else 1 for c in val)
                    if length > max_len:
                        max_len = length
                except:
                    pass
            ws.column_dimensions[col_letter].width = min(max_len + 5, 35)
        
        # 凍結標題
        ws.freeze_panes = "A2"
        
        # 使用 Excel 表格樣式（自動包含篩選功能和間隔色）
        from openpyxl.worksheet.table import Table, TableStyleInfo
        
        # 創建表格範圍
        tab = Table(displayName="Table1", ref=ws.dimensions)
        
        # 設定表格樣式（自動間隔色）
        style = TableStyleInfo(
            name="TableStyleLight1",  # 淡色樣式
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,      # 自動間隔行
            showColumnStripes=False
        )
        tab.tableStyleInfo = style
        
        # 套用表格
        ws.add_table(tab)
        
        wb.save(fn)
        print(f"\n📊 已儲存：{fn}")
        print("✅ 功能：自動篩選、凍結標題、間隔顏色、加寬欄位")

if __name__ == "__main__":
    data_file = get_latest_data_file()
    if not data_file:
        print("❌ 找不到數據檔，請先執行 data_downloader.py")
        exit(1)
    
    df = load_data(data_file)
    if df is None:
        exit(1)
    
    show_menu()
    strategy = input("選擇策略 [1]: ").strip() or "1"
    
    if strategy == "11":
        result, name = custom_filter(df)
    else:
        result, name = apply_preset_strategy(df, strategy)
    
    if result is not None and name is not None:
        export_result(result, name)
